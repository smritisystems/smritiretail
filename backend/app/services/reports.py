"""
Project      : SMRITI Retail OS
Repository   : SMRITIRetailNX
Organization : AITDL NETWORKS

Founders

* Pushpa Devi Jawahar Mallah — Founder & Chairperson
* Jawahar Ramkripal Mallah  — Founder, CEO & Chief Software Architect
* Websites: aitdl.com | erpnbook.com | smritibooks.com

* Version    : 3.23.0
* Created    : 2026-07-11
* Modified   : 2026-08-15
* Copyright  : © AITDL.com and SMRITIBooks.com. All Rights Reserved.
* License    : Proprietary Commercial Software
"""

from decimal import Decimal, ROUND_HALF_UP
from datetime import date, datetime, timezone
from typing import List, Dict, Any, Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import or_
from fastapi import HTTPException

from ..models.inventory import Product
from ..models.sales import (
    SalesInvoice, SalesInvoiceItem, SalesReturn, SalesReturnItem,
    SalesOrder, SalesOrderItem, SalesOrderInvoiceAllocation
)
from ..models.purchase import Supplier, PurchaseOrder, PurchaseReceipt
from ..models.supplier_payment import SupplierPayment
from ..models.report_schedule import ReportSchedule
from ..models.crm import Customer
from ..models.loyalty import LoyaltyMember, LoyaltyPointsLedger
from ..models.promotions import PromotionCampaign, PromotionRedemption
from ..models.commission import CommissionParticipant, CommissionLedger
from ..models.fulfillment import PackingSlip, Dispatch
from ..models.profitability import InvoiceProfitabilityLedger, ProductCostValuation
from ..api.deps import TenantContext
from ..schemas.reports import (
    StockValuationLine, StockValuationReport,
    DailySalesSummary,
    SupplierLedgerEntry, SupplierLedger,
    PurchaseSummaryLine,
    SalesOrderSummaryLine, SalesOrderSummaryReport,
    PendingOrderLine, PendingOrdersReport,
    BilledVsPendingOrderLine, BilledVsPendingOrdersReport,
    CustomerWiseOrderLine, CustomerWiseOrdersReport,
    ProductWiseOrderedQuantityLine, ProductWiseOrderedQuantityReport,
    OrderFulfillmentStatusGroup, OrderFulfillmentStatusReport,
    InvoiceAllocationReportLine, InvoiceAllocationReportModel,
    SalesOrderDetailLine, SalesOrderDetailReport,
)

class ReportsService:
    def __init__(self, db: AsyncSession, tenant: TenantContext):
        self.db = db
        self.tenant = tenant

    async def stock_valuation(self) -> StockValuationReport:
        stmt = select(Product).where(
            Product.is_deleted == False,
            Product.is_active == True,
        )
        if self.tenant and self.tenant.company_id:
            stmt = stmt.where(Product.company_id == self.tenant.company_id)
        stmt = stmt.order_by(Product.name)
        res = await self.db.execute(stmt)
        products = res.scalars().all()

        lines = []
        total_value = Decimal("0.00")
        for p in products:
            stock      = Decimal(str(p.stock or "0"))
            cost_price = Decimal(str(p.cost_price or "0"))
            value      = (stock * cost_price).quantize(Decimal("0.01"))
            total_value += value
            lines.append(StockValuationLine(
                product_id=p.id,
                code=p.code,
                name=p.name,
                stock=stock,
                cost_price=cost_price,
                stock_value=value,
            ))

        return StockValuationReport(
            generated_at=datetime.now(timezone.utc).isoformat(),
            total_items=len(lines),
            total_value=total_value.quantize(Decimal("0.01")),
            lines=lines,
        )

    async def daily_sales(self, report_date: Optional[date] = None) -> DailySalesSummary:
        stmt = select(SalesInvoice).where(
            SalesInvoice.is_deleted == False,
            or_(SalesInvoice.status.is_(None), SalesInvoice.status != "CANCELLED"),
        )
        if report_date:
            stmt = stmt.where(SalesInvoice.date == report_date)

        if self.tenant and self.tenant.company_id:
            stmt = stmt.where(SalesInvoice.company_id == self.tenant.company_id)

        if self.tenant and self.tenant.branch_id:
            stmt = stmt.where(SalesInvoice.branch_id == self.tenant.branch_id)

        res = await self.db.execute(stmt)
        invoices = res.scalars().all()

        total_sales = Decimal("0.00")
        tax_total = Decimal("0.00")
        cash_sales = Decimal("0.00")
        card_sales = Decimal("0.00")
        upi_sales = Decimal("0.00")
        credit_sales = Decimal("0.00")

        for inv in invoices:
            net = Decimal(str(getattr(inv, "net_amount", None) or getattr(inv, "total_amount", None) or getattr(inv, "grand_total", None) or "0"))
            total_sales += net
            tax_total += Decimal(str(getattr(inv, "tax_total", None) or "0"))

            pm = (inv.payment_mode or "").upper()
            if "CASH" in pm:
                cash_sales += net
            elif "CARD" in pm:
                card_sales += net
            elif "UPI" in pm:
                upi_sales += net
            elif "CREDIT" in pm:
                credit_sales += net
            else:
                cash_sales += net

        return DailySalesSummary(
            report_date=report_date or date.today(),
            total_invoices=len(invoices),
            total_sales=total_sales,
            tax_total=tax_total,
            cash_sales=cash_sales,
            card_sales=card_sales,
            upi_sales=upi_sales,
            credit_sales=credit_sales,
            shift_breakdown=[],
        )

    async def supplier_ledger(self, supplier_id: str) -> SupplierLedger:
        sup_res = await self.db.execute(
            select(Supplier).where(
                Supplier.id         == supplier_id,
                Supplier.company_id == self.tenant.company_id,
            )
        )
        supplier = sup_res.scalar_one_or_none()
        if not supplier:
            raise HTTPException(status_code=404, detail="Supplier not found")

        po_res = await self.db.execute(
            select(PurchaseReceipt).where(
                PurchaseReceipt.supplier_id == supplier_id,
                PurchaseReceipt.company_id  == self.tenant.company_id,
            )
        )
        receipts = po_res.scalars().all()

        pay_res = await self.db.execute(
            select(SupplierPayment).where(
                SupplierPayment.supplier_id == supplier_id,
                SupplierPayment.company_id  == self.tenant.company_id,
            )
        )
        payments = pay_res.scalars().all()

        entries = []
        for r in receipts:
            amt = float(getattr(r, "total_amount", None) or getattr(r, "grand_total", 0.0) or 0.0)
            entries.append({
                "date": str(getattr(r, "receipt_date", None) or getattr(r, "created_at", "")),
                "type": "PURCHASE",
                "reference": getattr(r, "receipt_number", None) or getattr(r, "receipt_no", "") or r.id,
                "debit": amt,
                "credit": 0.0,
            })
        for p in payments:
            amt = float(getattr(p, "amount", 0.0) or 0.0)
            entries.append({
                "date": str(getattr(p, "payment_date", None) or getattr(p, "created_at", "")),
                "type": "PAYMENT",
                "reference": getattr(p, "payment_number", None) or getattr(p, "payment_no", "") or p.id,
                "debit": 0.0,
                "credit": amt,
            })

        entries.sort(key=lambda x: x["date"])

        running_balance = Decimal("0.00")
        total_purchased = Decimal("0.00")
        total_paid = Decimal("0.00")
        typed_entries = []
        for e in entries:
            deb = Decimal(str(e["debit"]))
            cred = Decimal(str(e["credit"]))
            total_purchased += deb
            total_paid += cred
            running_balance += deb - cred
            typed_entries.append(SupplierLedgerEntry(
                entry_type=e["type"],
                date=e["date"],
                reference=e["reference"],
                amount=deb if deb > 0 else cred,
                balance_after=running_balance,
            ))

        return SupplierLedger(
            supplier_id=supplier.id,
            supplier_name=supplier.name,
            opening_balance=Decimal("0.00"),
            total_purchased=total_purchased,
            total_paid=total_paid,
            closing_balance=running_balance,
            entries=typed_entries,
        )

    async def purchase_summary(self, from_date: Optional[date] = None, to_date: Optional[date] = None) -> List[PurchaseSummaryLine]:
        sup_res = await self.db.execute(
            select(Supplier).where(
                Supplier.company_id == self.tenant.company_id,
                Supplier.is_deleted == False,
            )
        )
        suppliers = sup_res.scalars().all()
        result = []
        for sup in suppliers:
            po_res = await self.db.execute(
                select(PurchaseOrder).where(
                    PurchaseOrder.supplier_id == sup.id,
                    PurchaseOrder.company_id == self.tenant.company_id,
                    PurchaseOrder.is_deleted == False,
                )
            )
            pos = po_res.scalars().all()

            grn_res = await self.db.execute(
                select(PurchaseReceipt).where(
                    PurchaseReceipt.supplier_id == sup.id,
                    PurchaseReceipt.company_id == self.tenant.company_id,
                    PurchaseReceipt.is_deleted == False,
                )
            )
            grns = grn_res.scalars().all()

            total_ordered = sum((getattr(p, "total_amount", None) or getattr(p, "grand_total", Decimal("0.00")) or Decimal("0.00")) for p in pos)
            total_received = sum((getattr(g, "total_amount", None) or getattr(g, "grand_total", Decimal("0.00")) or Decimal("0.00")) for g in grns)

            result.append(PurchaseSummaryLine(
                supplier_id=sup.id,
                supplier_name=sup.name,
                po_count=len(pos),
                grn_count=len(grns),
                total_ordered=Decimal(str(total_ordered)),
                total_received=Decimal(str(total_received)),
                outstanding=Decimal(str(sup.outstanding or "0.00")),
            ))
        return result


    # ──────────────────────────────────────────────────────────────
    # 5. Profitability Waterfall Report
    # ──────────────────────────────────────────────────────────────
    async def profitability_report(self, cost_basis: str = "WAC") -> Dict[str, Any]:
        res = await self.db.execute(select(InvoiceProfitabilityLedger))
        ledgers = res.scalars().all()
        total_gross = sum(float(l.gross_sales_amount or 0) for l in ledgers)
        total_cogs = sum(float(l.total_cogs or 0) for l in ledgers)
        total_net_contrib = sum(float(l.net_contribution or 0) for l in ledgers)

        return {
            "cost_basis_selected": cost_basis,
            "total_invoices_analyzed": len(ledgers),
            "total_gross_sales": total_gross,
            "total_cogs": total_cogs,
            "total_net_contribution": total_net_contrib,
            "net_margin_percent": (total_net_contrib / total_gross * 100.0) if total_gross > 0 else 0.0,
            "ledgers": ledgers
        }

    # ──────────────────────────────────────────────────────────────
    # 6. Report Schedules
    # ──────────────────────────────────────────────────────────────
    async def list_schedules(self) -> List[ReportSchedule]:
        res = await self.db.execute(
            select(ReportSchedule).where(
                ReportSchedule.company_id == self.tenant.company_id,
                ReportSchedule.is_deleted == False,
            ).order_by(ReportSchedule.created_at.desc())
        )
        return res.scalars().all()

    async def create_schedule(self, payload: Any, created_by_id: str) -> ReportSchedule:
        import uuid
        cron_expr = getattr(payload, "cron_expression", None)
        if not cron_expr and getattr(payload, "frequency", None):
            freq = payload.frequency.upper()
            exec_time = getattr(payload, "execution_time", "08:00") or "08:00"
            try:
                hour, minute = exec_time.split(":")
                hour_i = int(hour)
                min_i = int(minute)
            except Exception:
                hour_i, min_i = 8, 0
            if freq == "DAILY":
                cron_expr = f"{min_i} {hour_i} * * *"
            elif freq == "WEEKLY":
                cron_expr = f"{min_i} {hour_i} * * 1"
            elif freq == "MONTHLY":
                cron_expr = f"{min_i} {hour_i} 1 * *"

        sched = ReportSchedule(
            id=f"SCH-{uuid.uuid4().hex[:12].upper()}",
            company_id=self.tenant.company_id,
            branch_id=self.tenant.branch_id,
            report_id=payload.report_id,
            report_name=payload.report_name,
            frequency=payload.frequency,
            execution_time=getattr(payload, "execution_time", "08:00"),
            delivery_format=payload.delivery_format,
            delivery_channel=payload.delivery_channel,
            delivery_target=payload.delivery_target,
            cron_expression=cron_expr,
            created_by_id=created_by_id,
        )
        self.db.add(sched)
        await self.db.commit()
        await self.db.refresh(sched)
        return sched

    async def delete_schedule(self, schedule_id: str) -> None:
        res = await self.db.execute(
            select(ReportSchedule).where(
                ReportSchedule.id == schedule_id,
                ReportSchedule.company_id == self.tenant.company_id,
            )
        )
        sched = res.scalar_one_or_none()
        if not sched:
            raise HTTPException(status_code=404, detail="Report schedule not found")
        sched.is_deleted = True
        sched.is_active = False
        self.db.add(sched)
        await self.db.commit()


    # ──────────────────────────────────────────────────────────────
    # Sprint 8a P1 Report Methods -- Shoper9 parity Tax & Compliance
    # ──────────────────────────────────────────────────────────────

    def _date_filter(self, stmt, model, from_date, to_date):
        if from_date:
            stmt = stmt.where(model.date >= from_date)
        if to_date:
            stmt = stmt.where(model.date <= to_date)
        return stmt

    def _tenant_filter(self, stmt, model):
        if self.tenant and self.tenant.company_id:
            stmt = stmt.where(model.company_id == self.tenant.company_id)
        if self.tenant and self.tenant.branch_id:
            stmt = stmt.where(model.branch_id == self.tenant.branch_id)
        return stmt

    async def bill_wise_sales(self, from_date=None, to_date=None):
        """RPT-TAX-002 -- Shoper9 SR202400 Bill-wise Sales."""
        from ..schemas.reports import BillWiseSalesLine, BillWiseSalesReport
        from sqlalchemy.orm import selectinload
        stmt = (
            select(SalesInvoice)
            .options(selectinload(SalesInvoice.items))
            .where(SalesInvoice.is_deleted == False, SalesInvoice.status != "CANCELLED")
        )
        stmt = self._tenant_filter(stmt, SalesInvoice)
        stmt = self._date_filter(stmt, SalesInvoice, from_date, to_date)
        invoices = (await self.db.execute(stmt.order_by(SalesInvoice.date.desc()))).scalars().all()
        lines, tg, td, tn, tt = [], Decimal("0.00"), Decimal("0.00"), Decimal("0.00"), Decimal("0.00")
        for inv in invoices:
            g = Decimal(str(getattr(inv, "gross_amount", None) or getattr(inv, "total_amount", None) or getattr(inv, "grand_total", "0") or "0"))
            d = Decimal(str(getattr(inv, "discount_amount", None) or getattr(inv, "discount", "0") or "0"))
            n = Decimal(str(getattr(inv, "net_amount", None) or getattr(inv, "grand_total", "0") or "0"))
            t = Decimal(str(getattr(inv, "tax_amount", None) or getattr(inv, "tax_total", None) or getattr(inv, "gst_amount", "0") or "0"))
            tg += g
            td += d
            tn += n
            tt += t
            lines.append(
                BillWiseSalesLine(
                    invoice_id=inv.id,
                    invoice_number=getattr(inv, "invoice_number", None) or getattr(inv, "invoice_no", None) or inv.id,
                    invoice_date=str(getattr(inv, "date", "") or ""),
                    customer_name=getattr(inv, "customer_name", None),
                    payment_mode=getattr(inv, "payment_mode", None),
                    gross_amount=g,
                    discount=d,
                    net_amount=n,
                    tax_amount=t,
                    items_count=len(inv.items or []),
                )
            )
        return BillWiseSalesReport(
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            total_bills=len(lines),
            total_gross=tg,
            total_discount=td,
            total_net=tn,
            total_tax=tt,
            lines=lines,
        )

    async def item_wise_sales(self, from_date=None, to_date=None):
        """RPT-TAX-003 -- Shoper9 SR202200 Item-wise Sales."""
        from ..schemas.reports import ItemWiseSalesLine, ItemWiseSalesReport
        stmt = (
            select(SalesInvoice, SalesInvoiceItem)
            .join(SalesInvoiceItem, SalesInvoiceItem.invoice_id == SalesInvoice.id)
            .where(SalesInvoice.is_deleted == False, SalesInvoice.status != "CANCELLED")
        )
        stmt = self._tenant_filter(stmt, SalesInvoice)
        stmt = self._date_filter(stmt, SalesInvoice, from_date, to_date)
        rows = (await self.db.execute(stmt)).all()
        
        agg: Dict[str, dict] = {}
        for inv, item in rows:
            pid = getattr(item, "product_id", None) or getattr(item, "code", None) or "UNKNOWN"
            qty = Decimal(str(getattr(item, "quantity", 0) or 0))
            net = Decimal(str(getattr(item, "total_amount", None) or getattr(item, "amount", 0) or 0))
            tax = Decimal(str(getattr(item, "tax_amount", 0) or 0))
            disc = Decimal(str(getattr(item, "discount_amount", 0) or 0))
            gross = net + disc
            
            if pid not in agg:
                agg[pid] = {
                    "code": getattr(item, "code", "") or getattr(item, "product_code", ""),
                    "name": getattr(item, "name", "") or getattr(item, "product_name", pid),
                    "hsn": getattr(item, "hsn_code", None),
                    "qty": Decimal("0.0000"),
                    "gross": Decimal("0.00"),
                    "disc": Decimal("0.00"),
                    "net": Decimal("0.00"),
                    "tax": Decimal("0.00"),
                    "rqty": Decimal("0.0000"),
                }
            agg[pid]["qty"] += qty
            agg[pid]["gross"] += gross
            agg[pid]["net"] += net
            agg[pid]["tax"] += tax
            agg[pid]["disc"] += disc
            
        lines = [
            ItemWiseSalesLine(
                product_id=pid,
                product_code=d["code"],
                product_name=d["name"],
                hsn_code=d["hsn"],
                qty_sold=d["qty"],
                gross_amount=d["gross"],
                discount=d["disc"],
                net_amount=d["net"],
                tax_amount=d["tax"],
                return_qty=d["rqty"],
            )
            for pid, d in sorted(agg.items(), key=lambda x: -x[1]["net"])
        ]
        return ItemWiseSalesReport(
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            total_items=len(lines),
            total_qty=sum(l.qty_sold for l in lines),
            total_net=sum(l.net_amount for l in lines),
            lines=lines,
        )

    async def tax_register(self, from_date=None, to_date=None):
        """RPT-TAX-001 -- Shoper9 SR202300 Tax Register."""
        from ..schemas.reports import TaxRegisterLine, TaxRegisterReport
        from sqlalchemy.orm import selectinload
        stmt = (
            select(SalesInvoice)
            .options(selectinload(SalesInvoice.items))
            .where(SalesInvoice.is_deleted == False, SalesInvoice.status != "CANCELLED")
        )
        stmt = self._tenant_filter(stmt, SalesInvoice)
        stmt = self._date_filter(stmt, SalesInvoice, from_date, to_date)
        invoices = (await self.db.execute(stmt.order_by(SalesInvoice.date.desc()))).scalars().all()
        
        lines: List[TaxRegisterLine] = []
        t_taxable = Decimal("0.00")
        t_cgst = Decimal("0.00")
        t_sgst = Decimal("0.00")
        t_igst = Decimal("0.00")
        t_tax = Decimal("0.00")
        
        for inv in invoices:
            taxable = Decimal(str(getattr(inv, "taxable_value", None) or getattr(inv, "taxable_amount", None) or getattr(inv, "net_amount", None) or getattr(inv, "grand_total", "0") or "0"))
            item_cgst = sum((Decimal(str(item.cgst_amount or "0")) for item in (inv.items or [])), Decimal("0.00"))
            item_sgst = sum((Decimal(str(item.sgst_amount or "0")) for item in (inv.items or [])), Decimal("0.00"))
            item_igst = sum((Decimal(str(item.igst_amount or "0")) for item in (inv.items or [])), Decimal("0.00"))
            header_cgst = Decimal(str(getattr(inv, "cgst_amount", "0") or "0"))
            header_sgst = Decimal(str(getattr(inv, "sgst_amount", "0") or "0"))
            header_igst = Decimal(str(getattr(inv, "igst_amount", "0") or "0"))
            cgst_a = item_cgst or header_cgst
            sgst_a = item_sgst or header_sgst
            igst_a = item_igst or header_igst
            tax_tot = (cgst_a + sgst_a + igst_a)
            if tax_tot <= 0:
                tax_tot = Decimal(str(getattr(inv, "tax_total", None) or getattr(inv, "tax_amount", "0") or "0"))
                if tax_tot > 0:
                    if getattr(inv, "is_interstate", False):
                        igst_a = tax_tot
                    else:
                        cgst_a = (tax_tot / Decimal("2")).quantize(Decimal("0.01"))
                        sgst_a = tax_tot - cgst_a

            effective_rate = (tax_tot / taxable * Decimal("100.00")).quantize(Decimal("0.01")) if taxable > 0 else Decimal("0.00")
            is_interstate = bool(getattr(inv, "is_interstate", False))
            
            t_taxable += taxable
            t_cgst += cgst_a
            t_sgst += sgst_a
            t_igst += igst_a
            t_tax += tax_tot
            
            lines.append(
                TaxRegisterLine(
                    invoice_number=getattr(inv, "invoice_number", None) or getattr(inv, "invoice_no", None) or inv.id,
                    invoice_date=str(getattr(inv, "date", "") or ""),
                    customer_name=getattr(inv, "customer_name", None),
                    taxable_amount=taxable,
                    cgst_rate=(effective_rate / Decimal("2")).quantize(Decimal("0.01")) if cgst_a > 0 and not is_interstate else Decimal("0.00"),
                    cgst_amount=cgst_a,
                    sgst_rate=(effective_rate / Decimal("2")).quantize(Decimal("0.01")) if sgst_a > 0 and not is_interstate else Decimal("0.00"),
                    sgst_amount=sgst_a,
                    igst_rate=effective_rate if igst_a > 0 else Decimal("0.00"),
                    igst_amount=igst_a,
                    total_tax=tax_tot,
                    net_amount=taxable + tax_tot,
                )
            )
        return TaxRegisterReport(
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            total_invoices=len(lines),
            total_taxable=t_taxable,
            total_cgst=t_cgst,
            total_sgst=t_sgst,
            total_igst=t_igst,
            total_tax=t_tax,
            lines=lines,
        )

    async def cancelled_bills(self, from_date=None, to_date=None):
        """RPT-TAX-004 -- Shoper9 SR210200 Cancelled Bills."""
        from ..schemas.reports import CancelledBillLine, CancelledBillsReport
        stmt = select(SalesInvoice).where(SalesInvoice.is_deleted == False, SalesInvoice.status == "CANCELLED")
        stmt = self._tenant_filter(stmt, SalesInvoice)
        stmt = self._date_filter(stmt, SalesInvoice, from_date, to_date)
        invoices = (await self.db.execute(stmt.order_by(SalesInvoice.date))).scalars().all()
        lines, total_voided = [], Decimal(0)
        for inv in invoices:
            amt = Decimal(str(getattr(inv,"grand_total",None) or getattr(inv,"total_amount","0") or "0"))
            total_voided += amt
            lines.append(CancelledBillLine(invoice_number=getattr(inv,"invoice_number",None) or inv.id,
                invoice_date=str(getattr(inv,"date","") or ""),
                cancelled_at=str(getattr(inv,"modified_at","") or ""),
                cancelled_by=getattr(inv,"updated_by",None),
                cancel_reason=getattr(inv,"cancel_reason",None) or getattr(inv,"remarks",None),
                original_amount=amt,customer_name=getattr(inv,"customer_name",None)))
        return CancelledBillsReport(from_date=str(from_date or ""),to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            total_cancelled=len(lines),total_value_voided=total_voided,lines=lines)

    async def salesperson_discount(self, from_date=None, to_date=None):
        """RPT-MIS-005 -- Shoper9 SR238400 Salesperson-wise Discount."""
        from ..schemas.reports import SalespersonDiscountLine, SalespersonDiscountReport
        stmt = select(SalesInvoice).where(SalesInvoice.is_deleted == False, SalesInvoice.status != "CANCELLED")
        stmt = self._tenant_filter(stmt, SalesInvoice)
        stmt = self._date_filter(stmt, SalesInvoice, from_date, to_date)
        invoices = (await self.db.execute(stmt)).scalars().all()
        agg: Dict[str, dict] = {}
        for inv in invoices:
            sp   = getattr(inv,"salesperson_name",None) or getattr(inv,"cashier_name",None) or "Unknown"
            disc = Decimal(str(getattr(inv,"discount_amount",None) or getattr(inv,"discount","0") or "0"))
            net  = Decimal(str(getattr(inv,"net_amount",None) or getattr(inv,"grand_total","0") or "0"))
            if sp not in agg:
                agg[sp]={"bills":0,"sales":Decimal(0),"disc":Decimal(0)}
            agg[sp]["bills"]+=1; agg[sp]["sales"]+=net; agg[sp]["disc"]+=disc
        lines=[]
        for name,d in sorted(agg.items(),key=lambda x:-x[1]["disc"]):
            pct=(d["disc"]/d["sales"]*100).quantize(Decimal("0.01")) if d["sales"]>0 else Decimal(0)
            lines.append(SalespersonDiscountLine(salesperson_name=name,total_bills=d["bills"],
                total_sales=d["sales"],total_discount=d["disc"],discount_pct=pct))
        return SalespersonDiscountReport(from_date=str(from_date or ""),to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            total_salespersons=len(lines),total_discount=sum(l.total_discount for l in lines),lines=lines)

    async def bill_wise_items(self, from_date=None, to_date=None):
        """RPT-TAX-005 -- Shoper9 SR202000 Bill-wise Items Detail."""
        from ..schemas.reports import BillWiseItemsLine, BillWiseItemsReport
        stmt = (
            select(SalesInvoice, SalesInvoiceItem)
            .join(SalesInvoiceItem, SalesInvoiceItem.invoice_id == SalesInvoice.id)
            .where(SalesInvoice.is_deleted == False, SalesInvoice.status != "CANCELLED")
        )
        stmt = self._tenant_filter(stmt, SalesInvoice)
        stmt = self._date_filter(stmt, SalesInvoice, from_date, to_date)
        stmt = stmt.order_by(SalesInvoice.date.desc(), SalesInvoice.id, SalesInvoiceItem.line_no)
        
        res = await self.db.execute(stmt)
        rows = res.all()
        
        lines: List[BillWiseItemsLine] = []
        unique_invs = set()
        total_qty = Decimal("0.0000")
        total_amt = Decimal("0.00")
        
        for inv, item in rows:
            unique_invs.add(inv.id)
            qty = Decimal(str(getattr(item, "quantity", 0) or 0))
            price = Decimal(str(getattr(item, "price", 0) or 0))
            line_tot = Decimal(str(getattr(item, "total_amount", None) or getattr(item, "amount", 0) or (qty * price) or 0))
            tax_amt = Decimal(str(getattr(item, "tax_amount", 0) or 0))
            gst = Decimal(str(getattr(item, "gst_rate", 18.00) or 18.00))
            disc = Decimal(str(getattr(item, "disc_pct", 0) or 0))
            
            total_qty += qty
            total_amt += line_tot
            
            lines.append(
                BillWiseItemsLine(
                    invoice_number=getattr(inv, "invoice_number", None) or getattr(inv, "invoice_no", None) or inv.id,
                    invoice_date=str(getattr(inv, "date", "") or ""),
                    customer_name=getattr(inv, "customer_name", None),
                    line_no=int(getattr(item, "line_no", None) or len(lines) + 1),
                    product_code=getattr(item, "code", "") or getattr(item, "product_code", ""),
                    product_name=getattr(item, "name", "") or getattr(item, "product_name", ""),
                    hsn_code=getattr(item, "hsn_code", None),
                    quantity=qty,
                    unit_price=price,
                    discount=disc,
                    gst_rate=gst,
                    tax_amount=tax_amt,
                    line_total=line_tot,
                )
            )
            
        return BillWiseItemsReport(
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            total_invoices=len(unique_invs),
            total_lines=len(lines),
            total_quantity=total_qty,
            total_amount=total_amt,
            lines=lines,
        )

    async def discount_summary(self, from_date=None, to_date=None):
        """RPT-OPS-001 -- Shoper9 SR202100 Discount Given Summary."""
        from ..schemas.reports import DiscountSummaryLine, DiscountSummaryReport
        stmt = select(SalesInvoice).where(SalesInvoice.is_deleted == False, SalesInvoice.status != "CANCELLED")
        stmt = self._tenant_filter(stmt, SalesInvoice)
        stmt = self._date_filter(stmt, SalesInvoice, from_date, to_date)
        invoices = (await self.db.execute(stmt.order_by(SalesInvoice.date.desc()))).scalars().all()
        
        lines: List[DiscountSummaryLine] = []
        tot_gross = Decimal("0.00")
        tot_disc = Decimal("0.00")
        tot_net = Decimal("0.00")
        
        for inv in invoices:
            gross = Decimal(str(getattr(inv, "gross_amount", None) or getattr(inv, "total_amount", None) or getattr(inv, "grand_total", "0") or "0"))
            disc = Decimal(str(getattr(inv, "discount_amount", None) or getattr(inv, "discount", "0") or "0"))
            net = Decimal(str(getattr(inv, "net_amount", None) or getattr(inv, "grand_total", "0") or "0"))
            
            tot_gross += gross
            tot_disc += disc
            tot_net += net
            
            disc_pct = (disc / gross * Decimal("100.00")).quantize(Decimal("0.01")) if gross > 0 else Decimal("0.00")
            
            lines.append(
                DiscountSummaryLine(
                    invoice_number=getattr(inv, "invoice_number", None) or getattr(inv, "invoice_no", None) or inv.id,
                    invoice_date=str(getattr(inv, "date", "") or ""),
                    salesperson_name=getattr(inv, "salesperson_name", None) or getattr(inv, "cashier_name", None),
                    customer_name=getattr(inv, "customer_name", None),
                    gross_amount=gross,
                    discount_amount=disc,
                    net_amount=net,
                    discount_pct=disc_pct,
                    remarks=getattr(inv, "remarks", None),
                )
            )
            
        overall_pct = (tot_disc / tot_gross * Decimal("100.00")).quantize(Decimal("0.01")) if tot_gross > 0 else Decimal("0.00")
        
        return DiscountSummaryReport(
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            total_bills=len(lines),
            total_gross=tot_gross,
            total_discount=tot_disc,
            total_net=tot_net,
            overall_discount_pct=overall_pct,
            lines=lines,
        )

    async def item_wise_returns(self, from_date=None, to_date=None):
        """RPT-MRC-003 -- Shoper9 SR214100 Item-wise Sales Returns."""
        from ..schemas.reports import ItemWiseReturnsLine, ItemWiseReturnsReport
        stmt = (
            select(SalesReturn, SalesReturnItem, SalesInvoice)
            .join(SalesReturnItem, SalesReturnItem.return_id == SalesReturn.id)
            .outerjoin(SalesInvoice, SalesInvoice.id == SalesReturn.original_invoice_id)
            .where(SalesReturn.is_deleted == False, SalesReturn.status != "CANCELLED")
        )
        stmt = self._tenant_filter(stmt, SalesReturn)
        stmt = self._date_filter(stmt, SalesReturn, from_date, to_date)
        stmt = stmt.order_by(SalesReturn.date.desc())
        
        res = await self.db.execute(stmt)
        rows = res.all()
        
        lines: List[ItemWiseReturnsLine] = []
        tot_qty = Decimal("0.0000")
        tot_amt = Decimal("0.00")
        
        for ret, item, orig_inv in rows:
            qty = Decimal(str(getattr(item, "quantity", 0) or 0))
            price = Decimal(str(getattr(item, "price", 0) or 0))
            amt = Decimal(str(getattr(item, "total_amount", None) or (qty * price) or 0))
            tax = Decimal(str(getattr(item, "tax_amount", 0) or 0))
            
            tot_qty += qty
            tot_amt += amt
            
            lines.append(
                ItemWiseReturnsLine(
                    return_number=getattr(ret, "return_no", None) or ret.id,
                    return_date=str(getattr(ret, "date", "") or ""),
                    original_inv_no=getattr(orig_inv, "invoice_no", None) or getattr(orig_inv, "invoice_number", None) or ret.original_invoice_id,
                    product_code=getattr(item, "code", ""),
                    product_name=getattr(item, "name", ""),
                    quantity=qty,
                    unit_price=price,
                    tax_amount=tax,
                    total_amount=amt,
                    reason=getattr(ret, "reason", None),
                )
            )
            
        return ItemWiseReturnsReport(
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            total_returns=len(lines),
            total_qty=tot_qty,
            total_amount=tot_amt,
            lines=lines,
        )

    async def attribute_size_sales(self, from_date=None, to_date=None):
        """RPT-MRC-001 -- Shoper9 SR236300 Attribute+Size wise Sales."""
        from ..schemas.reports import AttributeSizeSalesLine, AttributeSizeSalesReport
        stmt = (
            select(SalesInvoiceItem, Product)
            .join(SalesInvoice, SalesInvoice.id == SalesInvoiceItem.invoice_id)
            .outerjoin(Product, Product.id == SalesInvoiceItem.product_id)
            .where(SalesInvoice.is_deleted == False, SalesInvoice.status != "CANCELLED")
        )
        stmt = self._tenant_filter(stmt, SalesInvoice)
        stmt = self._date_filter(stmt, SalesInvoice, from_date, to_date)
        
        res = await self.db.execute(stmt)
        rows = res.all()
        
        agg: Dict[tuple, dict] = {}
        for item, prod in rows:
            cat = (prod.category if prod else None) or "General"
            brand = (prod.brand if prod else None) or "Standard"
            color = (prod.color if prod else None) or "N/A"
            size = (prod.size if prod else None) or "Standard"
            
            key = (cat, brand, color, size)
            qty = Decimal(str(getattr(item, "quantity", 0) or 0))
            gross = Decimal(str(getattr(item, "total_amount", None) or getattr(item, "amount", 0) or 0))
            disc = Decimal(str(getattr(item, "discount_amount", 0) or 0))
            net = gross - disc if gross >= disc else gross
            
            if key not in agg:
                agg[key] = {
                    "category": cat,
                    "brand": brand,
                    "color": color,
                    "size": size,
                    "qty_sold": Decimal("0.0000"),
                    "gross_revenue": Decimal("0.00"),
                    "discount": Decimal("0.00"),
                    "net_revenue": Decimal("0.00"),
                }
            agg[key]["qty_sold"] += qty
            agg[key]["gross_revenue"] += gross
            agg[key]["discount"] += disc
            agg[key]["net_revenue"] += net
            
        lines = [
            AttributeSizeSalesLine(**data)
            for key, data in sorted(agg.items(), key=lambda x: -x[1]["net_revenue"])
        ]
        
        return AttributeSizeSalesReport(
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            total_groups=len(lines),
            total_qty=sum(l.qty_sold for l in lines),
            total_net=sum(l.net_revenue for l in lines),
            lines=lines,
        )

    # ──────────────────────────────────────────────────────────────
    # SMRITI Standard Statutory Tax Invoices & Footwear Matrix Services
    # ──────────────────────────────────────────────────────────────

    @staticmethod
    def _parse_article_color_size(code: str, name: str) -> tuple:
        code_s = str(code or "").strip()
        name_s = str(name or "").strip()
        parts = code_s.split('-')
        if len(parts) >= 3 and parts[-1].isdigit():
            return '-'.join(parts[:-2]), parts[-2], parts[-1]
        name_parts = name_s.split()
        if len(name_parts) >= 3 and name_parts[-1].isdigit():
            return ' '.join(name_parts[:-2]), name_parts[-2], name_parts[-1]
        return code_s, "STANDARD", "FREE"

    @staticmethod
    def _format_pos(pos_state: str, customer_gstin: str, is_interstate: bool) -> tuple:
        from .invoice_pdf_service import GST_STATE_MAP
        code = None
        if customer_gstin and len(customer_gstin) >= 2 and customer_gstin[:2].isdigit() and customer_gstin[:2] in GST_STATE_MAP:
            code = customer_gstin[:2]
        elif pos_state:
            import re
            m = re.search(r'\(?(\d{2})\)?', str(pos_state))
            if m and m.group(1) in GST_STATE_MAP:
                code = m.group(1)
            else:
                for c, n in GST_STATE_MAP.items():
                    if n.lower() in str(pos_state).lower() or str(pos_state).lower() in n.lower():
                        code = c
                        break
        if not code:
            code = "27" if not is_interstate else "18"
        state_name = GST_STATE_MAP.get(code, str(pos_state) if pos_state else "Assam")
        supply_type = "Inter-State" if is_interstate else "Intra-State"
        return f"{state_name} ({code})", supply_type

    @staticmethod
    def _number_to_indian_words(num: float) -> str:
        from .invoice_pdf_service import number_to_indian_words
        return number_to_indian_words(num)

    async def tax_invoices_master_register(self, from_date=None, to_date=None, bill_from: Optional[int] = None, bill_to: Optional[int] = None, status_filter: Optional[str] = None):
        """RPT-TAX-006 -- Statutory GST Tax Invoices Master Register."""
        from ..schemas.reports import TaxInvoiceMasterRegisterLine, TaxInvoiceMasterRegisterReport
        from sqlalchemy.orm import selectinload
        import re

        stmt = (
            select(SalesInvoice)
            .options(selectinload(SalesInvoice.items))
            .where(SalesInvoice.is_deleted == False)
        )
        stmt = self._tenant_filter(stmt, SalesInvoice)
        stmt = self._date_filter(stmt, SalesInvoice, from_date, to_date)
        if status_filter:
            stmt = stmt.where(SalesInvoice.status == status_filter)

        invoices = (await self.db.execute(stmt.order_by(SalesInvoice.date, SalesInvoice.invoice_no))).scalars().all()

        lines: List[TaxInvoiceMasterRegisterLine] = []
        tot_qty = Decimal("0.00")
        tot_taxable = Decimal("0.00")
        tot_cgst = Decimal("0.00")
        tot_sgst = Decimal("0.00")
        tot_igst = Decimal("0.00")
        tot_tax = Decimal("0.00")
        tot_grand = Decimal("0.00")
        completed_cnt = 0
        cancelled_cnt = 0

        for inv in invoices:
            inv_no = getattr(inv, "invoice_no", None) or getattr(inv, "invoice_number", None) or inv.id
            m = re.search(r'TT2026-2027/([0-9]+)', inv_no)
            b_no = int(m.group(1)) if m else None
            if bill_from and (b_no is None or b_no < bill_from):
                continue
            if bill_to and (b_no is None or b_no > bill_to):
                continue

            inv_status = str(getattr(inv, "status", "COMPLETED") or "COMPLETED").upper()
            if inv_status == "CANCELLED":
                cancelled_cnt += 1
            else:
                completed_cnt += 1

            # Aggregate items if header taxable is 0
            items = inv.items or []
            item_sum_qty = Decimal("0.00")
            item_sum_taxable = Decimal("0.00")
            item_sum_tax = Decimal("0.00")
            for it in items:
                q = Decimal(str(it.quantity or 0))
                p = Decimal(str(it.price or 0))
                tx = Decimal(str(it.taxable_value or (q * p) or 0))
                # Reconcile if rate was MRP (> 1500)
                if p > 1500 and (it.disc_pct or Decimal(0)) > 0:
                    tx = (q * (p * (Decimal("1.00") - (Decimal(str(it.disc_pct)) / Decimal("100.00"))))).quantize(Decimal("0.01"))
                elif tx == 0 and p > 0:
                    tx = (q * p).quantize(Decimal("0.01"))
                
                t_amt = Decimal(str(it.tax_amount or (tx * (Decimal(str(it.gst_rate or 5.00)) / Decimal("100.00"))))).quantize(Decimal("0.01"))
                item_sum_qty += q
                item_sum_taxable += tx
                item_sum_tax += t_amt

            h_taxable = Decimal(str(getattr(inv, "taxable_value", None) or 0))
            taxable_val = h_taxable if h_taxable > 0 else item_sum_taxable
            grand_val = Decimal(str(getattr(inv, "grand_total", None) or getattr(inv, "net_amount", None) or (taxable_val + item_sum_tax) or 0))

            is_inter = bool(getattr(inv, "is_interstate", True))
            pos_disp, sup_type = self._format_pos(inv.pos_state, inv.customer_gstin, is_inter)

            if is_inter:
                cgst_a = Decimal("0.00")
                sgst_a = Decimal("0.00")
                igst_a = Decimal(str(getattr(inv, "tax_total", None) or item_sum_tax or (taxable_val * Decimal("0.05")))).quantize(Decimal("0.01"))
            else:
                h_tax = Decimal(str(getattr(inv, "tax_total", None) or item_sum_tax or (taxable_val * Decimal("0.05")))).quantize(Decimal("0.01"))
                cgst_a = (h_tax / Decimal("2.00")).quantize(Decimal("0.01"))
                sgst_a = h_tax - cgst_a
                igst_a = Decimal("0.00")

            t_tax = cgst_a + sgst_a + igst_a
            rnd = Decimal(str(getattr(inv, "rounding_amount", None) or (grand_val - (taxable_val + t_tax)))).quantize(Decimal("0.01"))
            words = getattr(inv, "amount_in_words", None) or self._number_to_indian_words(float(grand_val))

            tot_qty += item_sum_qty
            tot_taxable += taxable_val
            tot_cgst += cgst_a
            tot_sgst += sgst_a
            tot_igst += igst_a
            tot_tax += t_tax
            tot_grand += grand_val

            lines.append(
                TaxInvoiceMasterRegisterLine(
                    invoice_id=inv.id,
                    bill_no=b_no,
                    invoice_number=inv_no,
                    invoice_date=str(getattr(inv, "date", "") or ""),
                    status=inv_status,
                    document_type="TAX INVOICE",
                    sis_code=getattr(inv, "sis_code", None),
                    supplier_name="Tattly Threads",
                    supplier_gstin="27AAXFT2508H1ZR",
                    supplier_state="Maharashtra (27)",
                    customer_name=getattr(inv, "customer_name", "Reliance Retail Limited"),
                    customer_gstin=getattr(inv, "customer_gstin", None),
                    place_of_supply=pos_disp,
                    supply_type=sup_type,
                    reverse_charge="No",
                    po_reference=getattr(inv, "po_reference", None),
                    eway_bill_no=getattr(inv, "eway_bill_no", None),
                    irn=getattr(inv, "irn", None),
                    site_name=getattr(inv, "site_name", None),
                    billing_address=getattr(inv, "billing_address", None),
                    shipping_address=getattr(inv, "shipping_address", None),
                    items_count=len(items),
                    total_quantity=item_sum_qty,
                    taxable_value=taxable_val,
                    gst_rate=Decimal("5.00"),
                    cgst_amount=cgst_a,
                    sgst_amount=sgst_a,
                    igst_amount=igst_a,
                    total_tax=t_tax,
                    round_off=rnd,
                    grand_total=grand_val,
                    amount_in_words=words,
                )
            )

        return TaxInvoiceMasterRegisterReport(
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            total_invoices=len(lines),
            completed_count=completed_cnt,
            cancelled_count=cancelled_cnt,
            total_quantity=tot_qty,
            total_taxable=tot_taxable,
            total_cgst=tot_cgst,
            total_sgst=tot_sgst,
            total_igst=tot_igst,
            total_tax=tot_tax,
            total_grand_total=tot_grand,
            lines=lines,
        )

    async def article_color_size_matrix(self, from_date=None, to_date=None, article_filter: Optional[str] = None, color_filter: Optional[str] = None):
        """RPT-MRC-005 -- Article, Color & Size Variant Curve Matrix."""
        from ..schemas.reports import ArticleColorSizeMatrixRow, ArticleColorSizeMatrixReport

        stmt = (
            select(SalesInvoiceItem, SalesInvoice)
            .join(SalesInvoice, SalesInvoice.id == SalesInvoiceItem.invoice_id)
            .where(SalesInvoice.is_deleted == False)
        )
        stmt = self._tenant_filter(stmt, SalesInvoice)
        stmt = self._date_filter(stmt, SalesInvoice, from_date, to_date)

        rows = (await self.db.execute(stmt)).all()

        agg: Dict[tuple, dict] = {}
        for item, inv in rows:
            art, col, sz = self._parse_article_color_size(item.code, item.name)
            if article_filter and article_filter.upper() not in art.upper():
                continue
            if color_filter and color_filter.upper() not in col.upper():
                continue

            key = (art, col)
            qty = Decimal(str(item.quantity or 0))
            p = Decimal(str(item.price or 0))
            disc = Decimal(str(item.disc_pct or 0))
            if p > 1500 and disc > 0:
                net_rate = (p * (Decimal("1.00") - (disc / Decimal("100.00")))).quantize(Decimal("0.01"))
                tx = (qty * net_rate).quantize(Decimal("0.01"))
            else:
                tx = Decimal(str(item.taxable_value or (qty * p) or 0))
                if tx == 0 and p > 0:
                    tx = (qty * p).quantize(Decimal("0.01"))

            tax = Decimal(str(item.tax_amount or (tx * Decimal("0.05")))).quantize(Decimal("0.01"))
            tot = tx + tax

            if key not in agg:
                agg[key] = {
                    "article": art,
                    "color": col,
                    "size_36": Decimal("0"),
                    "size_37": Decimal("0"),
                    "size_38": Decimal("0"),
                    "size_39": Decimal("0"),
                    "size_40": Decimal("0"),
                    "size_41": Decimal("0"),
                    "size_42": Decimal("0"),
                    "total_units": Decimal("0"),
                    "taxable_value": Decimal("0.00"),
                    "tax_amount": Decimal("0.00"),
                    "gross_total": Decimal("0.00"),
                }

            agg[key]["total_units"] += qty
            agg[key]["taxable_value"] += tx
            agg[key]["tax_amount"] += tax
            agg[key]["gross_total"] += tot

            sz_key = f"size_{sz}"
            if sz_key in agg[key]:
                agg[key][sz_key] += qty

        matrix_rows = [
            ArticleColorSizeMatrixRow(**data)
            for key, data in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1]))
        ]

        return ArticleColorSizeMatrixReport(
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            total_variants=len(matrix_rows),
            total_units=sum(r.total_units for r in matrix_rows),
            total_taxable=sum(r.taxable_value for r in matrix_rows),
            total_tax=sum(r.tax_amount for r in matrix_rows),
            total_gross=sum(r.gross_total for r in matrix_rows),
            rows=matrix_rows,
        )

    async def store_wise_summary(self, from_date=None, to_date=None):
        """RPT-OPS-006 -- Store-Wise SIS Tax Invoice & Distribution Register."""
        from ..schemas.reports import StoreWiseSummaryLine, StoreWiseSummaryReport

        stmt = (
            select(SalesInvoice)
            .where(SalesInvoice.is_deleted == False)
        )
        stmt = self._tenant_filter(stmt, SalesInvoice)
        stmt = self._date_filter(stmt, SalesInvoice, from_date, to_date)

        invoices = (await self.db.execute(stmt)).scalars().all()

        agg: Dict[str, dict] = {}
        for inv in invoices:
            sis = getattr(inv, "sis_code", None) or "UNKNOWN"
            st = str(getattr(inv, "status", "COMPLETED") or "COMPLETED").upper()
            site = getattr(inv, "site_name", None) or getattr(inv, "shipping_address", "") or sis
            grand = Decimal(str(getattr(inv, "grand_total", None) or getattr(inv, "net_amount", "0") or 0))
            taxable = Decimal(str(getattr(inv, "taxable_value", None) or (grand / Decimal("1.05")) or 0)).quantize(Decimal("0.01"))
            tax = Decimal(str(getattr(inv, "tax_total", None) or (grand - taxable) or 0)).quantize(Decimal("0.01"))

            if sis not in agg:
                agg[sis] = {
                    "sis_code": sis,
                    "site_name": site,
                    "total_invoices": 0,
                    "completed_count": 0,
                    "cancelled_count": 0,
                    "total_quantity": Decimal("0"),
                    "taxable_value": Decimal("0.00"),
                    "tax_amount": Decimal("0.00"),
                    "grand_total": Decimal("0.00"),
                }

            agg[sis]["total_invoices"] += 1
            if st == "CANCELLED":
                agg[sis]["cancelled_count"] += 1
            else:
                agg[sis]["completed_count"] += 1

            agg[sis]["taxable_value"] += taxable
            agg[sis]["tax_amount"] += tax
            agg[sis]["grand_total"] += grand

        lines = [
            StoreWiseSummaryLine(**data)
            for sis, data in sorted(agg.items(), key=lambda x: x[0])
        ]

        return StoreWiseSummaryReport(
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            total_stores=len(lines),
            total_invoices=sum(l.total_invoices for l in lines),
            total_units=sum(l.total_quantity for l in lines),
            total_taxable=sum(l.taxable_value for l in lines),
            total_tax=sum(l.tax_amount for l in lines),
            total_grand=sum(l.grand_total for l in lines),
            lines=lines,
        )

    async def export_tax_invoices_master_excel(self, from_date=None, to_date=None, bill_from=None, bill_to=None, status=None) -> bytes:
        """Exports full 6-sheet statutory Tax Invoices workbook as Excel bytes."""
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Load data via reports methods
        reg_report = await self.tax_invoices_master_register(from_date, to_date, bill_from, bill_to, status)
        matrix_report = await self.article_color_size_matrix(from_date, to_date)
        store_report = await self.store_wise_summary(from_date, to_date)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Style tokens
        f_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        f_body = Font(name="Calibri", size=10, color="1E293B")
        f_body_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")
        fill_h = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_tot = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        fill_can = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        b_thin = Border(left=Side(style='thin', color="CBD5E1"), right=Side(style='thin', color="CBD5E1"), top=Side(style='thin', color="CBD5E1"), bottom=Side(style='thin', color="CBD5E1"))
        b_tot = Border(left=Side(style='thin', color="CBD5E1"), right=Side(style='thin', color="CBD5E1"), top=Side(style='thin', color="0F172A"), bottom=Side(style='double', color="0F172A"))

        # Sheet 1: Register
        ws1 = wb.create_sheet(title="Tax Invoices Register")
        ws1.views.sheetView[0].showGridLines = True
        headers1 = [
            "Bill No", "Invoice No", "Date", "Status", "Document Type", "Store Code", 
            "Supplier Name", "Supplier GSTIN", "Supplier State", "Customer Name", "Customer GSTIN", 
            "Place of Supply", "Supply Type", "Reverse Charge", "PO Reference", "E-Way Bill", 
            "Delivery Site", "Billing Address", "Shipping Address", "Items Count", "Quantity", 
            "Taxable Value (₹)", "GST Rate", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Total Tax (₹)", 
            "Round Off (₹)", "Grand Total (₹)", "Amount in Words"
        ]
        for c_idx, h in enumerate(headers1, 1):
            c = ws1.cell(row=1, column=c_idx, value=h)
            c.font, c.fill, c.alignment, c.border = f_header, fill_h, Alignment(horizontal='center', vertical='center', wrap_text=True), b_thin

        for r_idx, line in enumerate(reg_report.lines, 2):
            is_c = line.status == "CANCELLED"
            r_fill = fill_can if is_c else PatternFill(start_color="F8FAFC" if r_idx % 2 == 0 else "FFFFFF", end_color="F8FAFC" if r_idx % 2 == 0 else "FFFFFF", fill_type="solid")
            vals = [
                line.bill_no, line.invoice_number, line.invoice_date, line.status, line.document_type,
                line.sis_code, line.supplier_name, line.supplier_gstin, line.supplier_state,
                line.customer_name, line.customer_gstin, line.place_of_supply, line.supply_type,
                line.reverse_charge, line.po_reference, line.eway_bill_no, line.site_name,
                line.billing_address, line.shipping_address, line.items_count, float(line.total_quantity),
                float(line.taxable_value), "5.00%", float(line.cgst_amount), float(line.sgst_amount),
                float(line.igst_amount), float(line.total_tax), float(line.round_off), float(line.grand_total),
                line.amount_in_words
            ]
            for c_idx, v in enumerate(vals, 1):
                cell = ws1.cell(row=r_idx, column=c_idx, value=v)
                cell.font, cell.fill, cell.border = f_body, r_fill, b_thin
                if c_idx in [1, 2, 3, 4, 5, 6, 8, 9, 11, 13, 14, 15, 16, 23]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_idx in [20, 21, 22, 24, 25, 26, 27, 28, 29]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if c_idx >= 22:
                        cell.number_format = '₹ #,##0.00'

        # Total row
        last_r = len(reg_report.lines) + 1
        tot_r = last_r + 1
        ws1.cell(row=tot_r, column=1, value="TOTAL").alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=tot_r, column=2, value=f"{len(reg_report.lines)} Bills").alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=tot_r, column=21, value=float(reg_report.total_quantity)).number_format = '#,##0.00'
        ws1.cell(row=tot_r, column=22, value=float(reg_report.total_taxable)).number_format = '₹ #,##0.00'
        ws1.cell(row=tot_r, column=24, value=float(reg_report.total_cgst)).number_format = '₹ #,##0.00'
        ws1.cell(row=tot_r, column=25, value=float(reg_report.total_sgst)).number_format = '₹ #,##0.00'
        ws1.cell(row=tot_r, column=26, value=float(reg_report.total_igst)).number_format = '₹ #,##0.00'
        ws1.cell(row=tot_r, column=27, value=float(reg_report.total_tax)).number_format = '₹ #,##0.00'
        ws1.cell(row=tot_r, column=29, value=float(reg_report.total_grand_total)).number_format = '₹ #,##0.00'
        for c_idx in range(1, len(headers1) + 1):
            c = ws1.cell(row=tot_r, column=c_idx)
            c.font, c.fill, c.border = f_body_bold, fill_tot, b_tot

        # Sheet 2: Article & Size Matrix
        ws2 = wb.create_sheet(title="Article & Size Matrix")
        ws2.views.sheetView[0].showGridLines = True
        headers2 = ["Article", "Color", "Size 36", "Size 37", "Size 38", "Size 39", "Size 40", "Size 41", "Size 42", "Total Units", "Taxable Value (₹)", "IGST 5% (₹)", "Gross Total (₹)"]
        for c_idx, h in enumerate(headers2, 1):
            c = ws2.cell(row=1, column=c_idx, value=h)
            c.font, c.fill, c.alignment, c.border = f_header, fill_h, Alignment(horizontal='center', vertical='center', wrap_text=True), b_thin

        for r_idx, r in enumerate(matrix_report.rows, 2):
            r_fill = PatternFill(start_color="F8FAFC" if r_idx % 2 == 0 else "FFFFFF", end_color="F8FAFC" if r_idx % 2 == 0 else "FFFFFF", fill_type="solid")
            m_vals = [
                r.article, r.color, float(r.size_36), float(r.size_37), float(r.size_38),
                float(r.size_39), float(r.size_40), float(r.size_41), float(r.size_42),
                float(r.total_units), float(r.taxable_value), float(r.tax_amount), float(r.gross_total)
            ]
            for c_idx, v in enumerate(m_vals, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=v)
                cell.font, cell.fill, cell.border = f_body, r_fill, b_thin
                if c_idx <= 2:
                    cell.alignment, cell.font = Alignment(horizontal='center', vertical='center'), f_body_bold
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if c_idx in [11, 12, 13]:
                        cell.number_format = '₹ #,##0.00'
                    else:
                        cell.number_format = '#,##0.00'

        # Sheet 3: Store Wise Summary
        ws3 = wb.create_sheet(title="Store-Wise Summary")
        ws3.views.sheetView[0].showGridLines = True
        headers3 = ["Store Code", "Site / Store Name", "Total Invoices", "Completed", "Cancelled", "Taxable Value (₹)", "IGST 5% (₹)", "Gross Total (₹)"]
        for c_idx, h in enumerate(headers3, 1):
            c = ws3.cell(row=1, column=c_idx, value=h)
            c.font, c.fill, c.alignment, c.border = f_header, fill_h, Alignment(horizontal='center', vertical='center', wrap_text=True), b_thin

        for r_idx, s in enumerate(store_report.lines, 2):
            r_fill = PatternFill(start_color="F8FAFC" if r_idx % 2 == 0 else "FFFFFF", end_color="F8FAFC" if r_idx % 2 == 0 else "FFFFFF", fill_type="solid")
            s_vals = [s.sis_code, s.site_name, s.total_invoices, s.completed_count, s.cancelled_count, float(s.taxable_value), float(s.tax_amount), float(s.grand_total)]
            for c_idx, v in enumerate(s_vals, 1):
                cell = ws3.cell(row=r_idx, column=c_idx, value=v)
                cell.font, cell.fill, cell.border = f_body, r_fill, b_thin
                if c_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_idx >= 3:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if c_idx >= 6:
                        cell.number_format = '₹ #,##0.00'

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    # ─────────────────────────────────────────────────────────────────────────────
    # Sales Order BI Reports (RPT-SO-001 to RPT-SO-007)
    # ─────────────────────────────────────────────────────────────────────────────

    def _so_tenant_filter(self, stmt):
        if self.tenant and self.tenant.company_id:
            stmt = stmt.where(
                (SalesOrder.company_id == self.tenant.company_id) | 
                (SalesOrder.company_id == "COMP-001") | 
                (SalesOrder.company_id == "smriti001") | 
                (SalesOrder.company_id.is_(None))
            )
        if self.tenant and self.tenant.branch_id:
            stmt = stmt.where(
                (SalesOrder.branch_id == self.tenant.branch_id) | 
                (SalesOrder.branch_id == "MAIN") | 
                (SalesOrder.branch_id.is_(None))
            )
        return stmt

    async def sales_order_summary(
        self,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
        customer_id: Optional[str] = None,
        status: Optional[str] = None,
    ) -> SalesOrderSummaryReport:
        stmt = select(SalesOrder).where(SalesOrder.is_deleted == False)
        stmt = self._so_tenant_filter(stmt)
        if from_date:
            stmt = stmt.where(SalesOrder.date >= from_date)
        if to_date:
            stmt = stmt.where(SalesOrder.date <= to_date)
        if customer_id:
            stmt = stmt.where((SalesOrder.customer_id == customer_id) | (SalesOrder.customer_name.ilike(f"%{customer_id}%")))
        if status:
            stmt = stmt.where(SalesOrder.fulfillment_status == status)
        stmt = stmt.order_by(SalesOrder.date.desc(), SalesOrder.order_no.asc())

        res = await self.db.execute(stmt)
        orders = res.scalars().all()

        lines = []
        tot_qty = Decimal("0.0000")
        tot_basic = Decimal("0.00")
        tot_tax = Decimal("0.00")
        tot_val = Decimal("0.00")
        tot_billed = Decimal("0.00")
        tot_pending = Decimal("0.00")
        status_counts: Dict[str, int] = {}

        for o in orders:
            qty = Decimal(str(o.total_qty or "0.0000"))
            basic = Decimal(str(o.basic_total or "0.00"))
            tax = Decimal(str(o.tax_total or "0.00"))
            grand = Decimal(str(o.grand_total or "0.00"))
            billed = Decimal(str(o.billed_value or "0.00"))
            pending = Decimal(str(o.pending_value or (grand - billed)))
            f_status = o.fulfillment_status or "UNFULFILLED"

            tot_qty += qty
            tot_basic += basic
            tot_tax += tax
            tot_val += grand
            tot_billed += billed
            tot_pending += pending
            status_counts[f_status] = status_counts.get(f_status, 0) + 1

            lines.append(SalesOrderSummaryLine(
                order_no=o.order_no,
                po_number=o.po_number,
                customer_name=o.customer_name or "Reliance Retail Limited",
                date=str(o.date) if o.date else "",
                delivery_date=str(o.delivery_date) if o.delivery_date else None,
                site_code=o.site_code,
                total_qty=qty,
                basic_total=basic,
                tax_total=tax,
                grand_total=grand,
                billed_value=billed,
                pending_value=pending,
                fulfillment_status=f_status,
            ))

        filters_payload = {
            "company_id": self.tenant.company_id if self.tenant else None,
            "branch_id": self.tenant.branch_id if self.tenant else None,
            "from_date": str(from_date) if from_date else None,
            "to_date": str(to_date) if to_date else None,
            "customer_id": customer_id,
            "status": status,
        }
        summary_payload = {
            "total_orders": len(lines),
            "total_ordered_qty": tot_qty,
            "total_order_value": tot_val,
            "total_billed_value": tot_billed,
            "total_pending_value": tot_pending,
            "status_counts": status_counts,
        }
        totals_payload = {
            "total_qty": tot_qty,
            "basic_total": tot_basic,
            "tax_total": tot_tax,
            "grand_total": tot_val,
            "billed_value": tot_billed,
            "pending_value": tot_pending,
        }

        return SalesOrderSummaryReport(
            report_id="RPT-SO-001",
            report_name="Sales Order Summary",
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            filters=filters_payload,
            summary=summary_payload,
            totals=totals_payload,
            total_orders=len(lines),
            total_ordered_qty=tot_qty,
            total_order_value=tot_val,
            total_billed_value=tot_billed,
            total_pending_value=tot_pending,
            status_counts=status_counts,
            lines=lines,
            rows=lines,
        )

    async def pending_orders(
        self,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
        customer_id: Optional[str] = None,
    ) -> PendingOrdersReport:
        stmt = select(SalesOrder).where(
            SalesOrder.is_deleted == False,
            SalesOrder.fulfillment_status != "FULLY_BILLED"
        )
        stmt = self._so_tenant_filter(stmt)
        if from_date:
            stmt = stmt.where(SalesOrder.date >= from_date)
        if to_date:
            stmt = stmt.where(SalesOrder.date <= to_date)
        if customer_id:
            stmt = stmt.where((SalesOrder.customer_id == customer_id) | (SalesOrder.customer_name.ilike(f"%{customer_id}%")))
        stmt = stmt.order_by(SalesOrder.date.desc(), SalesOrder.order_no.asc())

        res = await self.db.execute(stmt)
        orders = res.scalars().all()

        lines = []
        tot_ordered_qty = Decimal("0.0000")
        tot_billed_qty = Decimal("0.0000")
        tot_pending_qty = Decimal("0.0000")
        tot_grand_val = Decimal("0.00")
        tot_billed_val = Decimal("0.00")
        tot_pending_val = Decimal("0.00")

        for o in orders:
            qty = Decimal(str(o.total_qty or "0.0000"))
            billed_qty = Decimal(str(o.billed_qty or "0.0000"))
            pending_qty = Decimal(str(o.pending_qty or (qty - billed_qty)))
            grand = Decimal(str(o.grand_total or "0.00"))
            billed = Decimal(str(o.billed_value or "0.00"))
            pending_val = Decimal(str(o.pending_value or (grand - billed)))
            f_status = o.fulfillment_status or "UNFULFILLED"

            tot_ordered_qty += qty
            tot_billed_qty += billed_qty
            tot_pending_qty += pending_qty
            tot_grand_val += grand
            tot_billed_val += billed
            tot_pending_val += pending_val

            lines.append(PendingOrderLine(
                order_no=o.order_no,
                po_number=o.po_number,
                customer_name=o.customer_name or "Reliance Retail Limited",
                po_date=str(o.po_date) if o.po_date else str(o.date) if o.date else None,
                delivery_date=str(o.delivery_date) if o.delivery_date else None,
                site_code=o.site_code,
                total_qty=qty,
                billed_qty=billed_qty,
                pending_qty=pending_qty,
                grand_total=grand,
                billed_value=billed,
                pending_value=pending_val,
                fulfillment_status=f_status,
            ))

        filters_payload = {
            "company_id": self.tenant.company_id if self.tenant else None,
            "branch_id": self.tenant.branch_id if self.tenant else None,
            "from_date": str(from_date) if from_date else None,
            "to_date": str(to_date) if to_date else None,
            "customer_id": customer_id,
        }
        summary_payload = {
            "total_pending_orders": len(lines),
            "total_pending_qty": tot_pending_qty,
            "total_pending_value": tot_pending_val,
        }
        totals_payload = {
            "total_qty": tot_ordered_qty,
            "billed_qty": tot_billed_qty,
            "pending_qty": tot_pending_qty,
            "grand_total": tot_grand_val,
            "billed_value": tot_billed_val,
            "pending_value": tot_pending_val,
        }

        return PendingOrdersReport(
            report_id="RPT-SO-002",
            report_name="Pending Orders",
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            filters=filters_payload,
            summary=summary_payload,
            totals=totals_payload,
            total_pending_orders=len(lines),
            total_pending_qty=tot_pending_qty,
            total_pending_value=tot_pending_val,
            lines=lines,
            rows=lines,
        )

    async def billed_vs_pending_orders(
        self,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
        customer_id: Optional[str] = None,
    ) -> BilledVsPendingOrdersReport:
        stmt = select(SalesOrder).where(SalesOrder.is_deleted == False)
        stmt = self._so_tenant_filter(stmt)
        if from_date:
            stmt = stmt.where(SalesOrder.date >= from_date)
        if to_date:
            stmt = stmt.where(SalesOrder.date <= to_date)
        if customer_id:
            stmt = stmt.where((SalesOrder.customer_id == customer_id) | (SalesOrder.customer_name.ilike(f"%{customer_id}%")))
        stmt = stmt.order_by(SalesOrder.date.desc(), SalesOrder.order_no.asc())

        res = await self.db.execute(stmt)
        orders = res.scalars().all()

        lines = []
        tot_val = Decimal("0.00")
        tot_billed = Decimal("0.00")
        tot_pending = Decimal("0.00")

        for o in orders:
            grand = Decimal(str(o.grand_total or "0.00"))
            billed = Decimal(str(o.billed_value or "0.00"))
            pending = Decimal(str(o.pending_value or (grand - billed)))
            pct = (billed / grand * Decimal("100.00")).quantize(Decimal("0.01")) if grand > 0 else Decimal("0.00")
            pending_pct = (pending / grand * Decimal("100.00")).quantize(Decimal("0.01")) if grand > 0 else Decimal("0.00")
            f_status = o.fulfillment_status or "UNFULFILLED"

            tot_val += grand
            tot_billed += billed
            tot_pending += pending

            lines.append(BilledVsPendingOrderLine(
                order_no=o.order_no,
                po_number=o.po_number,
                customer_name=o.customer_name or "Reliance Retail Limited",
                date=str(o.date) if o.date else "",
                grand_total=grand,
                billed_value=billed,
                pending_value=pending,
                billing_pct=pct,
                pending_pct=pending_pct,
                fulfillment_status=f_status,
            ))

        overall_pct = (tot_billed / tot_val * Decimal("100.00")).quantize(Decimal("0.01")) if tot_val > 0 else Decimal("0.00")

        filters_payload = {
            "company_id": self.tenant.company_id if self.tenant else None,
            "branch_id": self.tenant.branch_id if self.tenant else None,
            "from_date": str(from_date) if from_date else None,
            "to_date": str(to_date) if to_date else None,
            "customer_id": customer_id,
        }
        summary_payload = {
            "total_orders": len(lines),
            "total_order_value": tot_val,
            "total_billed_value": tot_billed,
            "total_pending_value": tot_pending,
            "overall_billing_pct": overall_pct,
        }
        totals_payload = {
            "grand_total": tot_val,
            "billed_value": tot_billed,
            "pending_value": tot_pending,
            "overall_billing_pct": overall_pct,
        }

        return BilledVsPendingOrdersReport(
            report_id="RPT-SO-003",
            report_name="Billed vs Pending Orders",
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            filters=filters_payload,
            summary=summary_payload,
            totals=totals_payload,
            total_orders=len(lines),
            total_order_value=tot_val,
            total_billed_value=tot_billed,
            total_pending_value=tot_pending,
            overall_billing_pct=overall_pct,
            lines=lines,
            rows=lines,
        )

    async def customer_wise_orders(
        self,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
    ) -> CustomerWiseOrdersReport:
        stmt = select(SalesOrder).where(SalesOrder.is_deleted == False)
        stmt = self._so_tenant_filter(stmt)
        if from_date:
            stmt = stmt.where(SalesOrder.date >= from_date)
        if to_date:
            stmt = stmt.where(SalesOrder.date <= to_date)
        stmt = stmt.order_by(SalesOrder.customer_name.asc())

        res = await self.db.execute(stmt)
        orders = res.scalars().all()

        cust_map: Dict[str, dict] = {}
        for o in orders:
            cname = o.customer_name or "Reliance Retail Limited"
            gstin = o.customer_gstin or ""
            qty = Decimal(str(o.total_qty or "0.0000"))
            grand = Decimal(str(o.grand_total or "0.00"))
            billed = Decimal(str(o.billed_value or "0.00"))
            pending = Decimal(str(o.pending_value or (grand - billed)))

            if cname not in cust_map:
                cust_map[cname] = {
                    "customer_name": cname,
                    "customer_gstin": gstin,
                    "order_count": 0,
                    "total_qty": Decimal("0.0000"),
                    "total_value": Decimal("0.00"),
                    "billed_value": Decimal("0.00"),
                    "pending_value": Decimal("0.00"),
                }
            entry = cust_map[cname]
            entry["order_count"] += 1
            entry["total_qty"] += qty
            entry["total_value"] += grand
            entry["billed_value"] += billed
            entry["pending_value"] += pending

        lines = []
        tot_val = Decimal("0.00")
        tot_billed = Decimal("0.00")
        tot_pending = Decimal("0.00")
        tot_orders = 0
        tot_qty = Decimal("0.0000")

        for cname, data in cust_map.items():
            cnt = data["order_count"]
            tval = data["total_value"]
            avg_val = (tval / Decimal(cnt)).quantize(Decimal("0.01")) if cnt > 0 else Decimal("0.00")
            tot_val += tval
            tot_billed += data["billed_value"]
            tot_pending += data["pending_value"]
            tot_orders += cnt
            tot_qty += data["total_qty"]

            lines.append(CustomerWiseOrderLine(
                customer_name=cname,
                customer_gstin=data["customer_gstin"],
                order_count=cnt,
                total_qty=data["total_qty"],
                total_value=tval,
                billed_value=data["billed_value"],
                pending_value=data["pending_value"],
                avg_order_value=avg_val,
            ))

        lines.sort(key=lambda x: x.total_value, reverse=True)

        filters_payload = {
            "company_id": self.tenant.company_id if self.tenant else None,
            "branch_id": self.tenant.branch_id if self.tenant else None,
            "from_date": str(from_date) if from_date else None,
            "to_date": str(to_date) if to_date else None,
        }
        summary_payload = {
            "total_customers": len(lines),
            "total_orders": tot_orders,
            "total_value": tot_val,
            "total_billed_value": tot_billed,
            "total_pending_value": tot_pending,
        }
        totals_payload = {
            "order_count": tot_orders,
            "total_qty": tot_qty,
            "total_value": tot_val,
            "billed_value": tot_billed,
            "pending_value": tot_pending,
        }

        return CustomerWiseOrdersReport(
            report_id="RPT-SO-004",
            report_name="Customer-wise Orders",
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            filters=filters_payload,
            summary=summary_payload,
            totals=totals_payload,
            total_customers=len(lines),
            total_orders=tot_orders,
            total_value=tot_val,
            total_billed_value=tot_billed,
            total_pending_value=tot_pending,
            lines=lines,
            rows=lines,
        )

    async def product_wise_ordered_qty(
        self,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
        product_id: Optional[str] = None,
    ) -> ProductWiseOrderedQuantityReport:
        stmt = (
            select(SalesOrderItem, SalesOrder)
            .join(SalesOrder, SalesOrderItem.order_id == SalesOrder.id)
            .where(SalesOrder.is_deleted == False)
        )
        if self.tenant and self.tenant.company_id:
            stmt = stmt.where((SalesOrder.company_id == self.tenant.company_id) | (SalesOrder.company_id.is_(None)))
        if from_date:
            stmt = stmt.where(SalesOrder.date >= from_date)
        if to_date:
            stmt = stmt.where(SalesOrder.date <= to_date)
        if product_id:
            stmt = stmt.where((SalesOrderItem.product_id == product_id) | (SalesOrderItem.article_no == product_id) | (SalesOrderItem.code == product_id))

        res = await self.db.execute(stmt)
        rows_db = res.all()

        prod_map: Dict[str, dict] = {}
        for item, order in rows_db:
            key = f"{item.article_no or item.code}_{item.vendor_style or ''}_{item.color or ''}_{item.size or ''}"
            qty = Decimal(str(item.quantity or "0.0000"))
            val = Decimal(str(item.total_amount or "0.00"))
            cost = Decimal(str(item.price or "0.00"))

            # Calculate item-level billing ratio based on parent order
            ord_total_qty = Decimal(str(order.total_qty or "0.0000"))
            ord_billed_qty = Decimal(str(order.billed_qty or "0.0000"))
            ratio = (ord_billed_qty / ord_total_qty) if ord_total_qty > 0 else Decimal("0.00")
            item_billed = (qty * ratio).quantize(Decimal("0.0001"))
            item_pending = qty - item_billed

            if key not in prod_map:
                prod_map[key] = {
                    "product_id": item.product_id,
                    "article_no": item.article_no or item.code,
                    "vendor_style": item.vendor_style or item.code,
                    "name": item.name,
                    "color": item.color,
                    "size": item.size,
                    "uom": item.uom or "EA",
                    "ordered_qty": Decimal("0.0000"),
                    "billed_qty": Decimal("0.0000"),
                    "pending_qty": Decimal("0.0000"),
                    "total_value": Decimal("0.00"),
                    "costs": [],
                    "order_ids": set(),
                }
            entry = prod_map[key]
            entry["ordered_qty"] += qty
            entry["billed_qty"] += item_billed
            entry["pending_qty"] += item_pending
            entry["total_value"] += val
            entry["costs"].append(cost)
            entry["order_ids"].add(order.id)

        lines = []
        tot_qty = Decimal("0.0000")
        tot_billed_qty = Decimal("0.0000")
        tot_pending_qty = Decimal("0.0000")
        tot_val = Decimal("0.00")

        for key, d in prod_map.items():
            tot_qty += d["ordered_qty"]
            tot_billed_qty += d["billed_qty"]
            tot_pending_qty += d["pending_qty"]
            tot_val += d["total_value"]
            avg_c = (sum(d["costs"]) / Decimal(len(d["costs"]))).quantize(Decimal("0.01")) if d["costs"] else Decimal("0.00")

            lines.append(ProductWiseOrderedQuantityLine(
                product_id=d["product_id"],
                article_no=d["article_no"],
                vendor_style=d["vendor_style"],
                name=d["name"],
                color=d["color"],
                size=d["size"],
                uom=d["uom"],
                ordered_qty=d["ordered_qty"],
                billed_qty=d["billed_qty"],
                pending_qty=d["pending_qty"],
                avg_cost=avg_c,
                total_value=d["total_value"],
                order_count=len(d["order_ids"]),
            ))

        lines.sort(key=lambda x: x.ordered_qty, reverse=True)

        filters_payload = {
            "company_id": self.tenant.company_id if self.tenant else None,
            "branch_id": self.tenant.branch_id if self.tenant else None,
            "from_date": str(from_date) if from_date else None,
            "to_date": str(to_date) if to_date else None,
            "product_id": product_id,
        }
        summary_payload = {
            "total_products": len(lines),
            "total_ordered_qty": tot_qty,
            "total_billed_qty": tot_billed_qty,
            "total_pending_qty": tot_pending_qty,
            "total_value": tot_val,
        }
        totals_payload = {
            "ordered_qty": tot_qty,
            "billed_qty": tot_billed_qty,
            "pending_qty": tot_pending_qty,
            "total_value": tot_val,
        }

        return ProductWiseOrderedQuantityReport(
            report_id="RPT-SO-005",
            report_name="Product-wise Ordered Quantity",
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            filters=filters_payload,
            summary=summary_payload,
            totals=totals_payload,
            total_products=len(lines),
            total_ordered_qty=tot_qty,
            total_value=tot_val,
            lines=lines,
            rows=lines,
        )

    async def order_fulfillment_status(
        self,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
    ) -> OrderFulfillmentStatusReport:
        summary_rep = await self.sales_order_summary(from_date=from_date, to_date=to_date)
        
        group_map: Dict[str, dict] = {}
        for l in summary_rep.lines:
            st = l.fulfillment_status
            if st not in group_map:
                group_map[st] = {
                    "status": st,
                    "order_count": 0,
                    "total_qty": Decimal("0.0000"),
                    "total_value": Decimal("0.00"),
                    "billed_value": Decimal("0.00"),
                    "pending_value": Decimal("0.00"),
                }
            g = group_map[st]
            g["order_count"] += 1
            g["total_qty"] += l.total_qty
            g["total_value"] += l.grand_total
            g["billed_value"] += l.billed_value
            g["pending_value"] += l.pending_value

        groups = [
            OrderFulfillmentStatusGroup(
                status=g["status"],
                order_count=g["order_count"],
                total_qty=g["total_qty"],
                total_value=g["total_value"],
                billed_value=g["billed_value"],
                pending_value=g["pending_value"],
            )
            for g in group_map.values()
        ]

        filters_payload = {
            "company_id": self.tenant.company_id if self.tenant else None,
            "branch_id": self.tenant.branch_id if self.tenant else None,
            "from_date": str(from_date) if from_date else None,
            "to_date": str(to_date) if to_date else None,
        }
        summary_payload = {
            "total_orders": summary_rep.total_orders,
            "total_value": summary_rep.total_order_value,
            "groups_count": len(groups),
        }
        totals_payload = {
            "order_count": summary_rep.total_orders,
            "total_qty": summary_rep.total_ordered_qty,
            "total_value": summary_rep.total_order_value,
            "billed_value": summary_rep.total_billed_value,
            "pending_value": summary_rep.total_pending_value,
        }

        return OrderFulfillmentStatusReport(
            report_id="RPT-SO-006",
            report_name="Order Fulfillment Status",
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            filters=filters_payload,
            summary=summary_payload,
            totals=totals_payload,
            total_orders=summary_rep.total_orders,
            total_value=summary_rep.total_order_value,
            groups=groups,
            lines=summary_rep.lines,
            rows=summary_rep.lines,
        )

    async def invoice_allocations(
        self,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
        order_id: Optional[str] = None,
    ) -> InvoiceAllocationReportModel:
        stmt = (
            select(SalesOrderInvoiceAllocation)
            .where(SalesOrderInvoiceAllocation.is_deleted == False)
        )
        if self.tenant and self.tenant.company_id:
            stmt = stmt.where((SalesOrderInvoiceAllocation.company_id == self.tenant.company_id) | (SalesOrderInvoiceAllocation.company_id.is_(None)))
        if from_date:
            stmt = stmt.where(SalesOrderInvoiceAllocation.invoice_date >= from_date)
        if to_date:
            stmt = stmt.where(SalesOrderInvoiceAllocation.invoice_date <= to_date)
        if order_id:
            stmt = stmt.where(
                (SalesOrderInvoiceAllocation.order_id == order_id) |
                (SalesOrderInvoiceAllocation.order_no == order_id) |
                (SalesOrderInvoiceAllocation.po_number == order_id)
            )
        stmt = stmt.order_by(SalesOrderInvoiceAllocation.invoice_date.desc(), SalesOrderInvoiceAllocation.invoice_no.asc())

        res = await self.db.execute(stmt)
        allocs = res.scalars().all()

        lines = []
        tot_po_qty = Decimal("0.0000")
        tot_po_val = Decimal("0.00")
        tot_billed_qty = Decimal("0.0000")
        tot_billed_val = Decimal("0.00")
        tot_pending_qty = Decimal("0.0000")
        tot_pending_val = Decimal("0.00")

        for a in allocs:
            p_qty = Decimal(str(a.po_quantity or "0.0000"))
            p_val = Decimal(str(a.po_value or "0.00"))
            b_qty = Decimal(str(a.billed_quantity or "0.0000"))
            b_val = Decimal(str(a.billed_value or "0.00"))
            pen_qty = Decimal(str(a.pending_quantity or "0.0000"))
            pen_val = Decimal(str(a.pending_value or "0.00"))

            tot_po_qty += p_qty
            tot_po_val += p_val
            tot_billed_qty += b_qty
            tot_billed_val += b_val
            tot_pending_qty += pen_qty
            tot_pending_val += pen_val

            lines.append(InvoiceAllocationReportLine(
                id=a.id,
                order_no=a.order_no,
                po_number=a.po_number,
                invoice_no=a.invoice_no,
                invoice_date=str(a.invoice_date) if a.invoice_date else "",
                po_quantity=p_qty,
                po_value=p_val,
                billed_quantity=b_qty,
                billed_value=b_val,
                pending_quantity=pen_qty,
                pending_value=pen_val,
                status=a.status or "ALLOCATED",
            ))

        filters_payload = {
            "company_id": self.tenant.company_id if self.tenant else None,
            "branch_id": self.tenant.branch_id if self.tenant else None,
            "from_date": str(from_date) if from_date else None,
            "to_date": str(to_date) if to_date else None,
            "order_id": order_id,
        }
        summary_payload = {
            "total_allocations": len(lines),
            "total_po_value": tot_po_val,
            "total_billed_value": tot_billed_val,
            "total_pending_value": tot_pending_val,
        }
        totals_payload = {
            "total_po_quantity": tot_po_qty,
            "total_po_value": tot_po_val,
            "total_billed_qty": tot_billed_qty,
            "total_billed_value": tot_billed_val,
            "total_pending_qty": tot_pending_qty,
            "total_pending_value": tot_pending_val,
        }

        return InvoiceAllocationReportModel(
            report_id="RPT-SO-007",
            report_name="Invoice Allocation Report",
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            filters=filters_payload,
            summary=summary_payload,
            totals=totals_payload,
            total_allocations=len(lines),
            total_po_quantity=tot_po_qty,
            total_po_value=tot_po_val,
            total_billed_qty=tot_billed_qty,
            total_billed_value=tot_billed_val,
            total_pending_qty=tot_pending_qty,
            total_pending_value=tot_pending_val,
            lines=lines,
            rows=lines,
        )

    async def sales_order_detailed(
        self,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
        customer_id: Optional[str] = None,
        status: Optional[str] = None,
        site_code: Optional[str] = None,
        search: Optional[str] = None,
    ) -> SalesOrderDetailReport:
        """RPT-SO-008 -- Detailed Line-Item Sales Orders Register & Fulfillment Trace."""
        from sqlalchemy.orm import selectinload

        stmt = select(SalesOrder).options(selectinload(SalesOrder.items)).where(SalesOrder.is_deleted == False)
        stmt = self._so_tenant_filter(stmt)
        if from_date:
            stmt = stmt.where(SalesOrder.date >= from_date)
        if to_date:
            stmt = stmt.where(SalesOrder.date <= to_date)
        if customer_id:
            stmt = stmt.where((SalesOrder.customer_id == customer_id) | (SalesOrder.customer_name.ilike(f"%{customer_id}%")))
        if status:
            stmt = stmt.where(SalesOrder.fulfillment_status == status)
        if site_code:
            stmt = stmt.where((SalesOrder.site_code == site_code) | (SalesOrder.site_name.ilike(f"%{site_code}%")))
        if search:
            s_term = f"%{search}%"
            stmt = stmt.where(
                (SalesOrder.order_no.ilike(s_term)) |
                (SalesOrder.po_number.ilike(s_term)) |
                (SalesOrder.customer_name.ilike(s_term)) |
                (SalesOrder.site_name.ilike(s_term))
            )
        stmt = stmt.order_by(SalesOrder.date.desc(), SalesOrder.order_no.asc())

        res = await self.db.execute(stmt)
        orders = res.scalars().all()

        # Load all allocations for linked invoice traceability
        alloc_res = await self.db.execute(select(SalesOrderInvoiceAllocation))
        all_allocs = alloc_res.scalars().all()
        alloc_map: Dict[str, List[SalesOrderInvoiceAllocation]] = {}
        for al in all_allocs:
            alloc_map.setdefault(al.order_id, []).append(al)

        lines: List[SalesOrderDetailLine] = []
        tot_ordered_qty = Decimal("0.0000")
        tot_taxable = Decimal("0.00")
        tot_tax = Decimal("0.00")
        tot_grand = Decimal("0.00")
        tot_billed_qty = Decimal("0.0000")
        tot_pending_qty = Decimal("0.0000")
        tot_billed_val = Decimal("0.00")
        tot_pending_val = Decimal("0.00")

        order_ids_seen = set()

        for o in orders:
            order_ids_seen.add(o.id)
            f_status = (o.fulfillment_status or "UNFULFILLED").upper()
            
            dest_state = "TELANGANA"
            if o.po_metadata and isinstance(o.po_metadata, dict):
                dest_state = o.po_metadata.get("site_state") or o.po_metadata.get("destination_state") or "TELANGANA"
            if o.delivery_address:
                addr_u = o.delivery_address.upper()
                if "TELANGANA" in addr_u or "HYDERABAD" in addr_u:
                    dest_state = "TELANGANA"
                elif "ASSAM" in addr_u or "DIBRUGARH" in addr_u:
                    dest_state = "ASSAM"
                elif "MAHARASHTRA" in addr_u or "MUMBAI" in addr_u:
                    dest_state = "MAHARASHTRA"
                elif "KARNATAKA" in addr_u or "BANGALORE" in addr_u:
                    dest_state = "KARNATAKA"
                elif "GUJARAT" in addr_u or "AHMEDABAD" in addr_u:
                    dest_state = "GUJARAT"

            pos = f"{dest_state} (36)" if "TELANGANA" in dest_state.upper() else f"{dest_state} (18)" if "ASSAM" in dest_state.upper() else dest_state
            is_interstate = bool(o.is_interstate) if o.is_interstate is not None else ("MAHARASHTRA" not in dest_state.upper())

            ord_allocs = alloc_map.get(o.id, [])
            inv_nos = list(set([a.invoice_no for a in ord_allocs if a.invoice_no]))
            inv_dates = list(set([str(a.invoice_date) for a in ord_allocs if a.invoice_date]))

            ord_tot_qty = Decimal(str(o.total_qty or "0.0000"))
            ord_billed_qty = Decimal(str(o.billed_qty or "0.0000"))
            billing_ratio = (ord_billed_qty / ord_tot_qty) if ord_tot_qty > 0 else (Decimal("1.00") if f_status == "FULFILLED" else Decimal("0.00"))

            items = getattr(o, "items", []) or []
            if not items:
                # If order has no discrete items recorded, generate a virtual header line
                q = Decimal(str(o.total_qty or "0.0000"))
                basic = Decimal(str(o.basic_total or "0.00"))
                tax = Decimal(str(o.tax_total or "0.00"))
                if f_status == "FULFILLED":
                    b_qty = q
                elif f_status == "UNFULFILLED":
                    b_qty = Decimal("0.0000")
                else:
                    b_qty = ord_billed_qty if ord_billed_qty > 0 else (q * billing_ratio).quantize(Decimal("1"), rounding=ROUND_HALF_UP).quantize(Decimal("0.0000"))
                p_qty = max(Decimal("0.0000"), q - b_qty)
                b_val = (grand * billing_ratio).quantize(Decimal("0.01"))
                p_val = grand - b_val

                igst = tax if is_interstate else Decimal("0.00")
                cgst = (tax / 2).quantize(Decimal("0.01")) if not is_interstate else Decimal("0.00")
                sgst = (tax - cgst) if not is_interstate else Decimal("0.00")

                tot_ordered_qty += q
                tot_taxable += basic
                tot_tax += tax
                tot_grand += grand
                tot_billed_qty += b_qty
                tot_pending_qty += p_qty
                tot_billed_val += b_val
                tot_pending_val += p_val

                lines.append(SalesOrderDetailLine(
                    order_id=o.id,
                    item_id=f"{o.id}_summary",
                    order_no=o.order_no,
                    po_number=o.po_number,
                    po_date=str(o.po_date or o.date) if (o.po_date or o.date) else "",
                    order_date=str(o.date) if o.date else "",
                    delivery_date=str(o.delivery_date) if o.delivery_date else None,
                    customer_id=o.customer_id,
                    customer_name=o.customer_name or "Reliance Retail Limited",
                    customer_gstin=o.customer_gstin or "27AABCR1718E1ZL",
                    site_code=o.site_code,
                    destination_state=dest_state,
                    place_of_supply=pos,
                    item_code="FOOTWEAR-LOT",
                    item_description="Footwear Assorted Footwear Pairs",
                    hsn_code="64041990",
                    category="FOOTWEAR",
                    ordered_qty=q,
                    unit_price=basic / q if q > 0 else basic,
                    mrp=grand / q if q > 0 else grand,
                    discount_pct=Decimal("0.00"),
                    taxable_value=basic,
                    gst_rate=Decimal("5.00"),
                    cgst_amount=cgst,
                    sgst_amount=sgst,
                    igst_amount=igst,
                    total_tax=tax,
                    total_amount=grand,
                    billed_qty=b_qty,
                    pending_qty=p_qty,
                    billed_value=b_val,
                    pending_value=p_val,
                    fulfillment_status=f_status,
                    linked_invoice_nos=inv_nos,
                    linked_invoice_dates=inv_dates,
                    eway_bill_no=None,
                ))
            else:
                for itm in items:
                    q = Decimal(str(itm.quantity or "0.0000"))
                    rate = Decimal(str(itm.price or "0.00"))
                    disc = Decimal("0.00")
                    taxable = Decimal(str(itm.taxable_value or (q * rate)))
                    tax = Decimal(str(itm.tax_amount or (taxable * Decimal("0.05"))))
                    grand = Decimal(str(itm.line_total or itm.total_amount or (taxable + tax)))

                    if f_status == "FULFILLED":
                        b_qty = q
                    elif f_status == "UNFULFILLED":
                        b_qty = Decimal("0.0000")
                    else:
                        b_qty = (q * billing_ratio).quantize(Decimal("1"), rounding=ROUND_HALF_UP).quantize(Decimal("0.0000"))
                    p_qty = max(Decimal("0.0000"), q - b_qty)
                    b_val = (grand * billing_ratio).quantize(Decimal("0.01"))
                    p_val = grand - b_val

                    igst = Decimal(str(itm.igst_amount or (tax if is_interstate else 0.00)))
                    cgst = Decimal(str(itm.cgst_amount or ((tax / 2).quantize(Decimal("0.01")) if not is_interstate else 0.00)))
                    sgst = Decimal(str(itm.sgst_amount or ((tax - cgst) if not is_interstate else 0.00)))

                    tot_ordered_qty += q
                    tot_taxable += taxable
                    tot_tax += tax
                    tot_grand += grand
                    tot_billed_qty += b_qty
                    tot_pending_qty += p_qty
                    tot_billed_val += b_val
                    tot_pending_val += p_val

                    lines.append(SalesOrderDetailLine(
                        order_id=o.id,
                        item_id=str(itm.id),
                        order_no=o.order_no,
                        po_number=o.po_number,
                        po_date=str(o.po_date or o.date) if (o.po_date or o.date) else "",
                        order_date=str(o.date) if o.date else "",
                        delivery_date=str(o.delivery_date) if o.delivery_date else None,
                        customer_id=o.customer_id,
                        customer_name=o.customer_name or "Reliance Retail Limited",
                        customer_gstin=o.customer_gstin or "27AABCR1718E1ZL",
                        site_code=o.site_code,
                        destination_state=dest_state,
                        place_of_supply=pos,
                        item_code=itm.code or itm.article_no or "FOOTWEAR",
                        item_description=itm.name or itm.article_no or "Footwear Pair",
                        hsn_code=itm.hsn_code or "64041990",
                        category="FOOTWEAR",
                        ordered_qty=q,
                        unit_price=rate,
                        mrp=Decimal(str(itm.mrp or (rate * Decimal("1.75")))),
                        discount_pct=disc,
                        taxable_value=taxable,
                        gst_rate=Decimal(str(itm.gst_rate or "5.00")),
                        cgst_amount=cgst,
                        sgst_amount=sgst,
                        igst_amount=igst,
                        total_tax=tax,
                        total_amount=grand,
                        billed_qty=b_qty,
                        pending_qty=p_qty,
                        billed_value=b_val,
                        pending_value=p_val,
                        fulfillment_status=f_status,
                        linked_invoice_nos=inv_nos,
                        linked_invoice_dates=inv_dates,
                        eway_bill_no=None,
                    ))

        fulfillment_rate = ((tot_billed_qty / tot_ordered_qty) * 100).quantize(Decimal("0.01")) if tot_ordered_qty > 0 else Decimal("0.00")

        filters_payload = {
            "company_id": self.tenant.company_id if self.tenant else None,
            "branch_id": self.tenant.branch_id if self.tenant else None,
            "from_date": str(from_date) if from_date else None,
            "to_date": str(to_date) if to_date else None,
            "customer_id": customer_id,
            "status": status,
            "site_code": site_code,
            "search": search,
        }
        summary_payload = {
            "total_orders": len(order_ids_seen),
            "total_lines": len(lines),
            "total_ordered_qty": tot_ordered_qty,
            "total_grand_amount": tot_grand,
            "total_billed_value": tot_billed_val,
            "total_pending_value": tot_pending_val,
            "fulfillment_rate_pct": fulfillment_rate,
        }
        totals_payload = {
            "total_ordered_qty": tot_ordered_qty,
            "total_taxable_value": tot_taxable,
            "total_tax_amount": tot_tax,
            "total_grand_amount": tot_grand,
            "total_billed_qty": tot_billed_qty,
            "total_pending_qty": tot_pending_qty,
            "total_billed_value": tot_billed_val,
            "total_pending_value": tot_pending_val,
        }

        return SalesOrderDetailReport(
            report_id="RPT-SO-008",
            report_name="Detailed Sales Orders Register",
            from_date=str(from_date or ""),
            to_date=str(to_date or ""),
            generated_at=datetime.now(timezone.utc).isoformat(),
            filters=filters_payload,
            summary=summary_payload,
            totals=totals_payload,
            total_orders=len(order_ids_seen),
            total_lines=len(lines),
            total_ordered_qty=tot_ordered_qty,
            total_taxable_value=tot_taxable,
            total_tax_amount=tot_tax,
            total_grand_amount=tot_grand,
            total_billed_qty=tot_billed_qty,
            total_pending_qty=tot_pending_qty,
            total_billed_value=tot_billed_val,
            total_pending_value=tot_pending_val,
            fulfillment_rate_pct=fulfillment_rate,
            lines=lines,
            rows=lines,
        )

    async def export_sales_orders_master_excel(
        self,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
        customer_id: Optional[str] = None,
        status: Optional[str] = None,
    ) -> bytes:
        """Generates full 6-sheet Master Sales Orders Excel workbook."""
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Load reports datasets
        detail_rep = await self.sales_order_detailed(from_date=from_date, to_date=to_date, customer_id=customer_id, status=status)
        summary_rep = await self.sales_order_summary(from_date=from_date, to_date=to_date, customer_id=customer_id, status=status)
        prod_rep = await self.product_wise_ordered_qty(from_date=from_date, to_date=to_date)
        alloc_rep = await self.invoice_allocations(from_date=from_date, to_date=to_date)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Style tokens
        f_title = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
        f_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        f_body = Font(name="Calibri", size=10, color="1E293B")
        f_body_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")
        fill_h = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_tot = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        fill_kpi = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        b_thin = Border(left=Side(style='thin', color="CBD5E1"), right=Side(style='thin', color="CBD5E1"), top=Side(style='thin', color="CBD5E1"), bottom=Side(style='thin', color="CBD5E1"))
        b_tot = Border(left=Side(style='thin', color="CBD5E1"), right=Side(style='thin', color="CBD5E1"), top=Side(style='thin', color="0F172A"), bottom=Side(style='double', color="0F172A"))

        # ─────────────────────────────────────────────────────────────────
        # Sheet 1: Executive Summary
        # ─────────────────────────────────────────────────────────────────
        ws1 = wb.create_sheet(title="Executive Summary")
        ws1.views.sheetView[0].showGridLines = True
        ws1.cell(row=1, column=1, value="SMRITI RETAIL OS — SALES ORDERS EXECUTIVE SUMMARY").font = f_title
        ws1.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')} | Supplier: TATTLY THREADS (27AAXFT2508H1ZR)").font = Font(size=9, italic=True, color="64748B")

        kpis = [
            ("Total Sales Orders", f"{detail_rep.total_orders} Orders"),
            ("Total Ordered Pairs", f"{float(detail_rep.total_ordered_qty):,.0f} Pairs"),
            ("Total Order Value", float(detail_rep.total_grand_amount)),
            ("Total Billed Value", float(detail_rep.total_billed_value)),
            ("Total Pending Value", float(detail_rep.total_pending_value)),
            ("Fulfillment Rate", f"{float(detail_rep.fulfillment_rate_pct):.1f}%"),
        ]
        ws1.cell(row=4, column=1, value="KPI METRIC").font = f_header
        ws1.cell(row=4, column=1).fill = fill_h
        ws1.cell(row=4, column=1).alignment = Alignment(horizontal="center")
        ws1.cell(row=4, column=2, value="VALUE").font = f_header
        ws1.cell(row=4, column=2).fill = fill_h
        ws1.cell(row=4, column=2).alignment = Alignment(horizontal="center")

        for r_i, (k, val) in enumerate(kpis, 5):
            c1 = ws1.cell(row=r_i, column=1, value=k)
            c2 = ws1.cell(row=r_i, column=2, value=val)
            c1.font, c1.border, c1.fill = f_body_bold, b_thin, fill_kpi
            c2.font, c2.border = f_body, b_thin
            if isinstance(val, (int, float)):
                c2.number_format = '₹ #,##0.00'
                c2.alignment = Alignment(horizontal="right")
            else:
                c2.alignment = Alignment(horizontal="center")

        # ─────────────────────────────────────────────────────────────────
        # Sheet 2: Sales Order Register (Header Level)
        # ─────────────────────────────────────────────────────────────────
        ws2 = wb.create_sheet(title="Sales Order Register")
        ws2.views.sheetView[0].showGridLines = True
        h2 = ["Order No", "PO Number", "Date", "Customer Name", "Site Code", "Delivery Date", "Total Qty", "Taxable Value (₹)", "Tax Total (₹)", "Grand Total (₹)", "Billed Value (₹)", "Pending Value (₹)", "Status"]
        for c_i, h in enumerate(h2, 1):
            c = ws2.cell(row=1, column=c_i, value=h)
            c.font, c.fill, c.alignment, c.border = f_header, fill_h, Alignment(horizontal='center', vertical='center', wrap_text=True), b_thin

        for r_i, o in enumerate(summary_rep.lines, 2):
            r_fill = PatternFill(start_color="F8FAFC" if r_i % 2 == 0 else "FFFFFF", end_color="F8FAFC" if r_i % 2 == 0 else "FFFFFF", fill_type="solid")
            row_vals = [
                o.order_no, o.po_number, o.date, o.customer_name, o.site_code or "-", o.delivery_date or "-",
                float(o.total_qty), float(o.basic_total), float(o.tax_total), float(o.grand_total),
                float(o.billed_value), float(o.pending_value), o.fulfillment_status
            ]
            for c_i, v in enumerate(row_vals, 1):
                cell = ws2.cell(row=r_i, column=c_i, value=v)
                cell.font, cell.fill, cell.border = f_body, r_fill, b_thin
                if c_i in [1, 2, 3, 5, 6, 13]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_i == 7:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0.00'
                elif c_i in [8, 9, 10, 11, 12]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '₹ #,##0.00'

        # Totals row for Sheet 2
        tot_r2 = len(summary_rep.lines) + 2
        ws2.cell(row=tot_r2, column=1, value="TOTAL").alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=tot_r2, column=7, value=float(summary_rep.total_ordered_qty)).number_format = '#,##0.00'
        ws2.cell(row=tot_r2, column=8, value=float(summary_rep.totals.get("basic_total", 0))).number_format = '₹ #,##0.00'
        ws2.cell(row=tot_r2, column=9, value=float(summary_rep.totals.get("tax_total", 0))).number_format = '₹ #,##0.00'
        ws2.cell(row=tot_r2, column=10, value=float(summary_rep.total_order_value)).number_format = '₹ #,##0.00'
        ws2.cell(row=tot_r2, column=11, value=float(summary_rep.total_billed_value)).number_format = '₹ #,##0.00'
        ws2.cell(row=tot_r2, column=12, value=float(summary_rep.total_pending_value)).number_format = '₹ #,##0.00'
        for c_i in range(1, len(h2) + 1):
            c = ws2.cell(row=tot_r2, column=c_i)
            c.font, c.fill, c.border = f_body_bold, fill_tot, b_tot

        # ─────────────────────────────────────────────────────────────────
        # Sheet 3: Detailed Line Items (Item Level)
        # ─────────────────────────────────────────────────────────────────
        ws3 = wb.create_sheet(title="Detailed Line Items")
        ws3.views.sheetView[0].showGridLines = True
        h3 = [
            "Order No", "PO Number", "Order Date", "Customer", "Site Code", "Destination State",
            "Item Code", "Item Description", "HSN Code", "Ordered Qty", "Unit Rate (₹)", "MRP (₹)", "Disc %",
            "Taxable Value (₹)", "GST %", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Total Amount (₹)",
            "Billed Qty", "Pending Qty", "Billed Value (₹)", "Pending Value (₹)", "Status", "Linked Invoices"
        ]
        for c_i, h in enumerate(h3, 1):
            c = ws3.cell(row=1, column=c_i, value=h)
            c.font, c.fill, c.alignment, c.border = f_header, fill_h, Alignment(horizontal='center', vertical='center', wrap_text=True), b_thin

        for r_i, l in enumerate(detail_rep.lines, 2):
            r_fill = PatternFill(start_color="F8FAFC" if r_i % 2 == 0 else "FFFFFF", end_color="F8FAFC" if r_i % 2 == 0 else "FFFFFF", fill_type="solid")
            row_vals = [
                l.order_no, l.po_number, l.order_date, l.customer_name, l.site_code or "-", l.destination_state or "-",
                l.item_code or "-", l.item_description, l.hsn_code, float(l.ordered_qty), float(l.unit_price), float(l.mrp),
                float(l.discount_pct), float(l.taxable_value), "5.00%", float(l.cgst_amount), float(l.sgst_amount),
                float(l.igst_amount), float(l.total_amount), float(l.billed_qty), float(l.pending_qty),
                float(l.billed_value), float(l.pending_value), l.fulfillment_status, ", ".join(l.linked_invoice_nos) if l.linked_invoice_nos else "-"
            ]
            for c_i, v in enumerate(row_vals, 1):
                cell = ws3.cell(row=r_i, column=c_i, value=v)
                cell.font, cell.fill, cell.border = f_body, r_fill, b_thin
                if c_i in [1, 2, 3, 5, 6, 7, 9, 15, 24, 25]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_i in [10, 20, 21]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0.00'
                elif c_i in [11, 12, 13, 14, 16, 17, 18, 19, 22, 23]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if c_i != 13:
                        cell.number_format = '₹ #,##0.00'

        # Totals row for Sheet 3
        tot_r3 = len(detail_rep.lines) + 2
        ws3.cell(row=tot_r3, column=1, value="TOTAL").alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(row=tot_r3, column=10, value=float(detail_rep.total_ordered_qty)).number_format = '#,##0.00'
        ws3.cell(row=tot_r3, column=14, value=float(detail_rep.total_taxable_value)).number_format = '₹ #,##0.00'
        ws3.cell(row=tot_r3, column=19, value=float(detail_rep.total_grand_amount)).number_format = '₹ #,##0.00'
        ws3.cell(row=tot_r3, column=20, value=float(detail_rep.total_billed_qty)).number_format = '#,##0.00'
        ws3.cell(row=tot_r3, column=21, value=float(detail_rep.total_pending_qty)).number_format = '#,##0.00'
        ws3.cell(row=tot_r3, column=22, value=float(detail_rep.total_billed_value)).number_format = '₹ #,##0.00'
        ws3.cell(row=tot_r3, column=23, value=float(detail_rep.total_pending_value)).number_format = '₹ #,##0.00'
        for c_i in range(1, len(h3) + 1):
            c = ws3.cell(row=tot_r3, column=c_i)
            c.font, c.fill, c.border = f_body_bold, fill_tot, b_tot

        # ─────────────────────────────────────────────────────────────────
        # Sheet 4: Store & Site Summary
        # ─────────────────────────────────────────────────────────────────
        ws4 = wb.create_sheet(title="Store & Site Summary")
        ws4.views.sheetView[0].showGridLines = True
        h4 = ["Site Code", "Destination State", "Orders Count", "Ordered Pairs", "Total Value (₹)", "Billed Pairs", "Billed Value (₹)", "Pending Value (₹)"]
        for c_i, h in enumerate(h4, 1):
            c = ws4.cell(row=1, column=c_i, value=h)
            c.font, c.fill, c.alignment, c.border = f_header, fill_h, Alignment(horizontal='center', vertical='center'), b_thin

        store_map: Dict[str, dict] = {}
        for l in detail_rep.lines:
            s_k = l.site_code or "UNSPECIFIED"
            if s_k not in store_map:
                store_map[s_k] = {
                    "site_code": s_k,
                    "dest_state": l.destination_state or "-",
                    "orders": set(),
                    "qty": Decimal("0.0000"),
                    "val": Decimal("0.00"),
                    "b_qty": Decimal("0.0000"),
                    "b_val": Decimal("0.00"),
                    "p_val": Decimal("0.00")
                }
            store_map[s_k]["orders"].add(l.order_no)
            store_map[s_k]["qty"] += l.ordered_qty
            store_map[s_k]["val"] += l.total_amount
            store_map[s_k]["b_qty"] += l.billed_qty
            store_map[s_k]["b_val"] += l.billed_value
            store_map[s_k]["p_val"] += l.pending_value

        for r_i, (k, d) in enumerate(store_map.items(), 2):
            r_fill = PatternFill(start_color="F8FAFC" if r_i % 2 == 0 else "FFFFFF", end_color="F8FAFC" if r_i % 2 == 0 else "FFFFFF", fill_type="solid")
            row_vals = [
                d["site_code"], d["dest_state"], len(d["orders"]), float(d["qty"]),
                float(d["val"]), float(d["b_qty"]), float(d["b_val"]), float(d["p_val"])
            ]
            for c_i, v in enumerate(row_vals, 1):
                cell = ws4.cell(row=r_i, column=c_i, value=v)
                cell.font, cell.fill, cell.border = f_body, r_fill, b_thin
                if c_i in [1, 2, 3]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_i in [4, 6]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0.00'
                elif c_i in [5, 7, 8]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '₹ #,##0.00'

        # ─────────────────────────────────────────────────────────────────
        # Sheet 5: Product & Style Matrix
        # ─────────────────────────────────────────────────────────────────
        ws5 = wb.create_sheet(title="Product & Style Matrix")
        ws5.views.sheetView[0].showGridLines = True
        h5 = ["Article / Code", "Description", "Color", "Size", "Ordered Qty", "Billed Qty", "Pending Qty", "Total Value (₹)"]
        for c_i, h in enumerate(h5, 1):
            c = ws5.cell(row=1, column=c_i, value=h)
            c.font, c.fill, c.alignment, c.border = f_header, fill_h, Alignment(horizontal='center', vertical='center'), b_thin

        for r_i, p in enumerate(prod_rep.lines, 2):
            r_fill = PatternFill(start_color="F8FAFC" if r_i % 2 == 0 else "FFFFFF", end_color="F8FAFC" if r_i % 2 == 0 else "FFFFFF", fill_type="solid")
            row_vals = [
                p.article_no or p.vendor_style or "-", p.name, p.color or "-", p.size or "-",
                float(p.ordered_qty), float(p.billed_qty), float(p.pending_qty), float(p.total_value)
            ]
            for c_i, v in enumerate(row_vals, 1):
                cell = ws5.cell(row=r_i, column=c_i, value=v)
                cell.font, cell.fill, cell.border = f_body, r_fill, b_thin
                if c_i in [1, 3, 4]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_i in [5, 6, 7]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0.00'
                elif c_i == 8:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '₹ #,##0.00'

        # ─────────────────────────────────────────────────────────────────
        # Sheet 6: Invoice Allocations
        # ─────────────────────────────────────────────────────────────────
        ws6 = wb.create_sheet(title="Invoice Allocations")
        ws6.views.sheetView[0].showGridLines = True
        h6 = ["Order No", "PO Number", "Invoice No", "Invoice Date", "PO Qty", "PO Value (₹)", "Billed Qty", "Billed Value (₹)", "Pending Qty", "Pending Value (₹)", "Status"]
        for c_i, h in enumerate(h6, 1):
            c = ws6.cell(row=1, column=c_i, value=h)
            c.font, c.fill, c.alignment, c.border = f_header, fill_h, Alignment(horizontal='center', vertical='center'), b_thin

        for r_i, al in enumerate(alloc_rep.lines, 2):
            r_fill = PatternFill(start_color="F8FAFC" if r_i % 2 == 0 else "FFFFFF", end_color="F8FAFC" if r_i % 2 == 0 else "FFFFFF", fill_type="solid")
            row_vals = [
                al.order_no, al.po_number or "-", al.invoice_no, al.invoice_date,
                float(al.po_quantity), float(al.po_value), float(al.billed_quantity), float(al.billed_value),
                float(al.pending_quantity), float(al.pending_value), al.status
            ]
            for c_i, v in enumerate(row_vals, 1):
                cell = ws6.cell(row=r_i, column=c_i, value=v)
                cell.font, cell.fill, cell.border = f_body, r_fill, b_thin
                if c_i in [1, 2, 3, 4, 11]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_i in [5, 7, 9]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0.00'
                elif c_i in [6, 8, 10]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '₹ #,##0.00'

        # Auto-adjust column widths across all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    async def export_sales_orders_csv(
        self,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
        customer_id: Optional[str] = None,
        status: Optional[str] = None,
    ) -> str:
        """Generates flat CSV string for Sales Order detailed line items."""
        import io
        import csv

        rep = await self.sales_order_detailed(from_date=from_date, to_date=to_date, customer_id=customer_id, status=status)
        output = io.StringIO()
        writer = csv.writer(output)

        headers = [
            "Order No", "PO Number", "Order Date", "Customer Name", "Customer GSTIN",
            "Site Code", "Destination State", "Place of Supply", "Item Code", "Item Description",
            "HSN Code", "Ordered Qty", "Unit Price", "MRP", "Discount %", "Taxable Value",
            "GST Rate", "CGST Amount", "SGST Amount", "IGST Amount", "Total Tax", "Total Amount",
            "Billed Qty", "Pending Qty", "Billed Value", "Pending Value", "Fulfillment Status", "Linked Invoices"
        ]
        writer.writerow(headers)

        for l in rep.lines:
            writer.writerow([
                l.order_no, l.po_number, l.order_date, l.customer_name, l.customer_gstin,
                l.site_code or "", l.destination_state or "", l.place_of_supply or "",
                l.item_code or "", l.item_description, l.hsn_code, float(l.ordered_qty),
                float(l.unit_price), float(l.mrp), float(l.discount_pct), float(l.taxable_value),
                float(l.gst_rate), float(l.cgst_amount), float(l.sgst_amount), float(l.igst_amount),
                float(l.total_tax), float(l.total_amount), float(l.billed_qty), float(l.pending_qty),
                float(l.billed_value), float(l.pending_value), l.fulfillment_status,
                "; ".join(l.linked_invoice_nos) if l.linked_invoice_nos else ""
            ])

        return output.getvalue()

    async def sales_order_fulfillment_variance(
        self,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
        customer_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Automated Fulfillment Variance & Backorder Analytics.
        Computes SLA aging buckets, store shortages, style variances, and actionable order queues.
        """
        today = datetime.now(timezone.utc).date()
        rep = await self.sales_order_detailed(from_date=from_date, to_date=to_date, customer_id=customer_id)

        # 1. Aging Buckets
        aging_buckets = {
            "0_7_days": {"count": 0, "ordered_qty": Decimal("0.00"), "pending_qty": Decimal("0.00"), "pending_val": Decimal("0.00")},
            "8_14_days": {"count": 0, "ordered_qty": Decimal("0.00"), "pending_qty": Decimal("0.00"), "pending_val": Decimal("0.00")},
            "15_30_days": {"count": 0, "ordered_qty": Decimal("0.00"), "pending_qty": Decimal("0.00"), "pending_val": Decimal("0.00")},
            "over_30_days": {"count": 0, "ordered_qty": Decimal("0.00"), "pending_qty": Decimal("0.00"), "pending_val": Decimal("0.00")},
        }

        # 2. Store-Level Variance
        store_map: Dict[str, Dict[str, Any]] = {}
        # 3. Style-Level Variance
        style_map: Dict[str, Dict[str, Any]] = {}
        # 4. Order-Level Queue
        order_map: Dict[str, Dict[str, Any]] = {}

        for l in rep.lines:
            # Order grouping
            o_no = l.order_no
            if o_no not in order_map:
                o_date = today
                if l.order_date:
                    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
                        try:
                            o_date = datetime.strptime(str(l.order_date).strip(), fmt).date()
                            break
                        except Exception:
                            pass
                age_days = (today - o_date).days
                order_map[o_no] = {
                    "order_no": o_no,
                    "po_number": l.po_number,
                    "order_date": l.order_date,
                    "age_days": max(0, age_days),
                    "customer_name": l.customer_name,
                    "site_code": l.site_code or "1977",
                    "destination_state": l.destination_state or "Maharashtra",
                    "ordered_qty": Decimal("0.00"),
                    "billed_qty": Decimal("0.00"),
                    "pending_qty": Decimal("0.00"),
                    "ordered_val": Decimal("0.00"),
                    "billed_val": Decimal("0.00"),
                    "pending_val": Decimal("0.00"),
                    "linked_invoices": l.linked_invoice_nos,
                    "fulfillment_status": l.fulfillment_status,
                }
            
            order_map[o_no]["ordered_qty"] += l.ordered_qty
            order_map[o_no]["billed_qty"] += l.billed_qty
            order_map[o_no]["pending_qty"] += l.pending_qty
            order_map[o_no]["ordered_val"] += l.total_amount
            order_map[o_no]["billed_val"] += l.billed_value
            order_map[o_no]["pending_val"] += l.pending_value

            # Store grouping
            st_code = l.site_code or "Unassigned"
            if st_code not in store_map:
                store_map[st_code] = {
                    "site_code": st_code,
                    "destination_state": l.destination_state or "Unknown",
                    "total_orders": set(),
                    "ordered_qty": Decimal("0.00"),
                    "billed_qty": Decimal("0.00"),
                    "pending_qty": Decimal("0.00"),
                    "ordered_val": Decimal("0.00"),
                    "pending_val": Decimal("0.00"),
                }
            store_map[st_code]["total_orders"].add(o_no)
            store_map[st_code]["ordered_qty"] += l.ordered_qty
            store_map[st_code]["billed_qty"] += l.billed_qty
            store_map[st_code]["pending_qty"] += l.pending_qty
            store_map[st_code]["ordered_val"] += l.total_amount
            store_map[st_code]["pending_val"] += l.pending_value

            # Style grouping
            # Extract base style code (e.g. CH-24-G BLACK)
            desc_parts = l.item_description.split(" ")
            style_key = " ".join(desc_parts[:3]) if len(desc_parts) >= 3 else l.item_description
            if style_key not in style_map:
                style_map[style_key] = {
                    "style_name": style_key,
                    "hsn_code": l.hsn_code,
                    "ordered_qty": Decimal("0.00"),
                    "billed_qty": Decimal("0.00"),
                    "pending_qty": Decimal("0.00"),
                    "ordered_val": Decimal("0.00"),
                    "pending_val": Decimal("0.00"),
                }
            style_map[style_key]["ordered_qty"] += l.ordered_qty
            style_map[style_key]["billed_qty"] += l.billed_qty
            style_map[style_key]["pending_qty"] += l.pending_qty
            style_map[style_key]["ordered_val"] += l.total_amount
            style_map[style_key]["pending_val"] += l.pending_value

        # Aggregate aging from orders
        for o in order_map.values():
            age = o["age_days"]
            if age <= 7:
                bucket = "0_7_days"
            elif age <= 14:
                bucket = "8_14_days"
            elif age <= 30:
                bucket = "15_30_days"
            else:
                bucket = "over_30_days"

            aging_buckets[bucket]["count"] += 1
            aging_buckets[bucket]["ordered_qty"] += o["ordered_qty"]
            aging_buckets[bucket]["pending_qty"] += o["pending_qty"]
            aging_buckets[bucket]["pending_val"] += o["pending_val"]

        # Convert structures to serializable formats
        stores_list = []
        for s in store_map.values():
            o_qty = int(round(float(s["ordered_qty"])))
            b_qty = int(round(float(s["billed_qty"])))
            p_qty = max(0, o_qty - b_qty)
            rate = round((b_qty / o_qty * 100), 1) if o_qty > 0 else 0.0
            stores_list.append({
                "site_code": s["site_code"],
                "destination_state": s["destination_state"],
                "orders_count": len(s["total_orders"]),
                "ordered_qty": o_qty,
                "billed_qty": b_qty,
                "pending_qty": p_qty,
                "ordered_val": float(s["ordered_val"]),
                "pending_val": float(s["pending_val"]),
                "fulfillment_rate": rate,
            })
        stores_list.sort(key=lambda x: x["pending_qty"], reverse=True)

        styles_list = []
        for st in style_map.values():
            o_qty = int(round(float(st["ordered_qty"])))
            b_qty = int(round(float(st["billed_qty"])))
            p_qty = max(0, o_qty - b_qty)
            rate = round((b_qty / o_qty * 100), 1) if o_qty > 0 else 0.0
            styles_list.append({
                "style_name": st["style_name"],
                "hsn_code": st["hsn_code"],
                "ordered_qty": o_qty,
                "billed_qty": b_qty,
                "pending_qty": p_qty,
                "ordered_val": float(st["ordered_val"]),
                "pending_val": float(st["pending_val"]),
                "fulfillment_rate": rate,
            })
        styles_list.sort(key=lambda x: x["pending_qty"], reverse=True)

        orders_list = []
        for o in order_map.values():
            o_qty = int(round(float(o["ordered_qty"])))
            b_qty = int(round(float(o["billed_qty"])))
            p_qty = max(0, o_qty - b_qty)
            rate = round((b_qty / o_qty * 100), 1) if o_qty > 0 else 0.0
            orders_list.append({
                "order_no": o["order_no"],
                "po_number": o["po_number"],
                "order_date": o["order_date"],
                "age_days": o["age_days"],
                "customer_name": o["customer_name"],
                "site_code": o["site_code"],
                "destination_state": o["destination_state"],
                "ordered_qty": o_qty,
                "billed_qty": b_qty,
                "pending_qty": p_qty,
                "ordered_val": float(o["ordered_val"]),
                "billed_val": float(o["billed_val"]),
                "pending_val": float(o["pending_val"]),
                "fulfillment_rate": rate,
                "fulfillment_status": o["fulfillment_status"],
                "linked_invoices": o["linked_invoices"],
            })
        orders_list.sort(key=lambda x: (x["pending_qty"], x["age_days"]), reverse=True)

        total_booked_pairs = sum(o["ordered_qty"] for o in orders_list)
        total_billed_pairs = sum(o["billed_qty"] for o in orders_list)
        total_pending_pairs = sum(o["pending_qty"] for o in orders_list)
        total_pending_val = sum(o["pending_val"] for o in orders_list)
        overall_rate = round((total_billed_pairs / total_booked_pairs * 100), 1) if total_booked_pairs > 0 else 0.0

        return {
            "summary": {
                "total_orders": len(orders_list),
                "total_booked_pairs": total_booked_pairs,
                "total_billed_pairs": total_billed_pairs,
                "total_pending_pairs": total_pending_pairs,
                "total_pending_value": total_pending_val,
                "overall_fulfillment_rate": overall_rate,
            },
            "aging_buckets": {
                k: {
                    "count": v["count"],
                    "ordered_qty": int(round(float(v["ordered_qty"]))),
                    "pending_qty": int(round(float(v["pending_qty"]))),
                    "pending_val": float(v["pending_val"]),
                }
                for k, v in aging_buckets.items()
            },
            "stores": stores_list,
            "styles": styles_list,
            "orders": orders_list,
        }





