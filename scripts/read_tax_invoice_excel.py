#!/usr/bin/env python3
"""
Read Tax Invoices Excel file to analyze items for import
"""
import openpyxl

file_path = 'Tax_Invoices_TT_18_to_137_Article_Color_Size_Split.xlsx'

print("=" * 80)
print(f"READING: {file_path}")
print("=" * 80)

wb = openpyxl.load_workbook(file_path)
print(f"\n✓ Worksheets found: {wb.sheetnames}")

# Read the first sheet
ws = wb.active
print(f"\n✓ Active sheet: {ws.title}")
print(f"✓ Dimensions: {ws.dimensions}")

# Get headers
headers = []
for cell in ws[1]:
    headers.append(cell.value)
    
print(f"\n✓ Headers ({len(headers)}):")
for i, h in enumerate(headers, 1):
    print(f"   {i}. {h}")

# Show first 5 rows
print(f"\n✓ First 5 data rows:")
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 2):
    print(f"   Row {row_idx}: {row[:10]}")  # First 10 columns

# Count total rows
total_rows = ws.max_row - 1
print(f"\n✓ Total data rows: {total_rows}")

wb.close()
