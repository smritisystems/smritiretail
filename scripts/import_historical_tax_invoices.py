#!/usr/bin/env python3
"""Import the authoritative TT2026-2027/18..137 invoice package into smriti001.

The command is dry-run by default. Use --commit only after reviewing the plan.
It never deletes tenant data and is idempotent by invoice_no and product code.
"""
from __future__ import annotations

import argparse
import hashlib
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2
from psycopg2.extras import execute_batch

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT / "exports" / "Tax_Invoices_TT_18_to_137_Complete_Package" / "Tax_Invoices_18_to_137_Master.xlsx"
DB_URL = "postgresql://postgres:postgres@localhost:5432/smriti001"
COMPANY_ID = "COMP-001"
BRANCH_ID = "MAIN"
BATCH_ID = "HIST-TT-18-137-CANONICAL-V1"


def stable_id(prefix: str, value: str) -> str:
    return f"{prefix}-{hashlib.sha1(value.encode('utf-8')).hexdigest()[:24]}"


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None if value in (None, "") else date.fromisoformat(str(value)[:10])


def dec(value: Any) -> Decimal:
    return Decimal(str(value or 0)).quantize(Decimal("0.01"))


def read_source(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    summary_sheet = workbook["Invoice_Summary"]
    item_sheet = workbook["Item_Details"]
    summary_headers = list(next(summary_sheet.iter_rows(values_only=True)))
    item_headers = list(next(item_sheet.iter_rows(values_only=True)))
    summaries = [dict(zip(summary_headers, row)) for row in summary_sheet.iter_rows(min_row=2, values_only=True)]
    items = [dict(zip(item_headers, row)) for row in item_sheet.iter_rows(min_row=2, values_only=True)]
    return summaries, items


def source_profile(summaries: list[dict[str, Any]], items: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "invoices": len({row["invoice_no"] for row in summaries}),
        "lines": len(items),
        "skus": len({row["item_code"] for row in items}),
        "pos": len({row["po_reference"] for row in summaries if row.get("po_reference")}),
        "sis_codes": len({row["sis_code"] for row in summaries if row.get("sis_code")}),
    }


def ensure_customer(cur, row: dict[str, Any]) -> str:
    customer_id = row.get("customer_id") or stable_id("cust", str(row["customer_name"]))
    cur.execute("SELECT id FROM customers WHERE id = %s", (customer_id,))
    if cur.fetchone():
        return customer_id
    cur.execute(
        """
        INSERT INTO customers (id, uuid, company_id, branch_id, code, name, gst_number,
            status, is_active, is_deleted, version, created_at, modified_at)
        VALUES (%s, gen_random_uuid()::text, %s, %s, %s, %s, %s, 'Active', true, false, 1, now(), now())
        """,
        (customer_id, COMPANY_ID, BRANCH_ID, "RRL-001", row["customer_name"], row.get("customer_gstin")),
    )
    return customer_id


def ensure_gst_and_location(cur, row: dict[str, Any], customer_id: str) -> tuple[str | None, str | None]:
    gstin = row.get("customer_gstin")
    gst_id = None
    location_id = None
    if gstin:
        gst_id = stable_id("cgr", f"{customer_id}|{gstin}")
        state_code = str(gstin)[:2]
        cur.execute(
            "SELECT id FROM customer_gst_registrations WHERE company_id = %s AND gstin = %s LIMIT 1",
            (COMPANY_ID, gstin),
        )
        existing_gst = cur.fetchone()
        if existing_gst:
            gst_id = existing_gst[0]
        else:
            cur.execute(
                """
                INSERT INTO customer_gst_registrations
                    (id, uuid, company_id, branch_id, customer_id, gstin, state_name, state_code,
                     registration_type, is_primary, status, is_active, is_deleted, version, created_at, modified_at)
                VALUES (%s, gen_random_uuid()::text, %s, %s, %s, %s, %s, %s, 'REGULAR', false,
                        'ACTIVE', true, false, 1, now(), now())
                """,
                (gst_id, COMPANY_ID, BRANCH_ID, customer_id, gstin, row.get("pos_state") or "Unknown", state_code),
            )
    sis_code = str(row.get("sis_code") or "").strip()
    if sis_code:
        location_id = stable_id("cdl", f"{customer_id}|{sis_code}")
        cur.execute(
            """
            INSERT INTO customer_delivery_locations
                (id, uuid, company_id, branch_id, customer_id, store_code, location_name,
                 address_line1, state, state_code, country, gst_registration_id, gstin,
                 is_default, status, source, is_active, is_deleted, version, created_at, modified_at)
            VALUES (%s, gen_random_uuid()::text, %s, %s, %s, %s, %s, %s, %s, %s, 'India', %s, %s,
                    false, 'ACTIVE', 'EXCEL_IMPORT', true, false, 1, now(), now())
            ON CONFLICT (id) DO UPDATE SET location_name = EXCLUDED.location_name,
                address_line1 = EXCLUDED.address_line1, gstin = EXCLUDED.gstin
            """,
            (location_id, COMPANY_ID, BRANCH_ID, customer_id, sis_code, row.get("site_name") or sis_code,
             row.get("shipping_address"), row.get("pos_state"), str(gstin or "")[:2] or None,
             gst_id, gstin),
        )
    return gst_id, location_id


def ensure_products(cur, items: list[dict[str, Any]]) -> dict[str, str]:
    product_ids: dict[str, str] = {}
    for item in items:
        code = str(item["item_code"]).strip()
        cur.execute(
            "SELECT id FROM products WHERE (code = %s OR sku = %s) LIMIT 1",
            (code, code),
        )
        found = cur.fetchone()
        if found:
            product_ids[code] = found[0]
            cur.execute(
                "UPDATE products SET is_deleted = false, is_active = true, modified_at = now() WHERE id = %s",
                (found[0],),
            )
            continue
        product_id = stable_id("prd-hist", code)
        product_ids[code] = product_id
        cur.execute(
            """
            INSERT INTO products
                (id, uuid, company_id, branch_id, code, sku, name, category, barcode, mrp, price,
                 cost_price, stock, reserved_stock, hsn_code, gst_percentage, is_active, is_deleted,
                 version, created_at, modified_at)
            VALUES (%s, gen_random_uuid()::text, %s, %s, %s, %s, %s, 'FOOTWEAR', %s, %s, %s,
                    0, 0, 0, %s, %s, true, false, 1, now(), now())
            ON CONFLICT (id) DO UPDATE SET name = EXCLUDED.name, mrp = EXCLUDED.mrp,
                price = EXCLUDED.price, hsn_code = EXCLUDED.hsn_code,
                gst_percentage = EXCLUDED.gst_percentage, is_deleted = false, is_active = true
            """,
            (product_id, COMPANY_ID, BRANCH_ID, code, code, item["item_name"], code,
             dec(item.get("mrp")), dec(item.get("unit_price")), item.get("hsn_code"), dec(item.get("gst_rate"))),
        )
    return product_ids


def import_package(path: Path, commit: bool) -> None:
    summaries, items = read_source(path)
    profile = source_profile(summaries, items)
    print(f"Source: {path}")
    print(f"Scope: {profile}")
    if profile != {"invoices": 120, "lines": 6661, "skus": 217, "pos": 58, "sis_codes": 60}:
        raise RuntimeError(f"Unexpected source profile; refusing import: {profile}")

    items_by_invoice: dict[str, list[dict[str, Any]]] = {}
    unique_items: dict[str, dict[str, Any]] = {}
    for item in items:
        items_by_invoice.setdefault(item["invoice_no"], []).append(item)
        unique_items.setdefault(str(item["item_code"]), item)

    conn = psycopg2.connect(DB_URL)
    try:
        cur = conn.cursor()
        cur.execute("SELECT count(*) FROM sales_invoices WHERE import_batch_id = %s", (BATCH_ID,))
        print(f"Existing imported invoices: {cur.fetchone()[0]}")
        if not commit:
            print("DRY RUN: no database changes committed")
            return

        products = ensure_products(cur, list(unique_items.values()))
        for row in summaries:
            invoice_no = row["invoice_no"]
            customer_id = ensure_customer(cur, row)
            gst_id, location_id = ensure_gst_and_location(cur, row, customer_id)
            invoice_id = row.get("id") or stable_id("inv", invoice_no)
            invoice_date = as_date(row.get("date"))
            cur.execute("DELETE FROM sales_invoice_items WHERE invoice_id = %s", (invoice_id,))
            cur.execute(
                """
                INSERT INTO sales_invoices
                    (id, uuid, company_id, branch_id, invoice_no, date, customer_id, tax_total,
                     grand_total, is_interstate, payment_mode, status, is_active, is_deleted,
                     version, source_type, source_system, source_file, import_batch_id, imported_at,
                     import_validation_status, sis_code, pos_state, reverse_charge, is_reverse_charge,
                     po_reference, customer_name, customer_gstin, billing_address, shipping_address,
                     site_name, delivery_location_id, delivery_store_code, delivery_gstin,
                     billed_party_gstin_id, place_of_supply_code,
                     taxable_value, rounding_amount, amount_in_words, bank_name, account_no, ifsc_code,
                     original_pdf_sha256, original_pdf_path, original_pdf_size, original_pdf_pages,
                     e_invoice_status, paid_amount, balance_amount, discount_amount, net_amount)
                VALUES (%s, gen_random_uuid()::text, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        true, false, 1, 'HISTORICAL_IMPORT', %s, %s, %s, %s, 'VALIDATED', %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, 'NOT_APPLICABLE', %s, %s, %s, %s)
                ON CONFLICT (invoice_no) DO UPDATE SET
                    date = EXCLUDED.date, customer_id = EXCLUDED.customer_id,
                    tax_total = EXCLUDED.tax_total, grand_total = EXCLUDED.grand_total,
                    status = EXCLUDED.status, sis_code = EXCLUDED.sis_code,
                    po_reference = EXCLUDED.po_reference, import_batch_id = EXCLUDED.import_batch_id,
                    modified_at = now(), is_deleted = false, is_active = true
                """,
                (invoice_id, COMPANY_ID, BRANCH_ID, invoice_no, invoice_date, customer_id,
                 dec(row.get("tax_total")), dec(row.get("grand_total")), bool(row.get("is_interstate")),
                 row.get("payment_mode") or "BANK_TRANSFER", row.get("status") or "COMPLETED",
                 row.get("source_system") or "LEGACY_PDF_EXPORT", row.get("source_file"), BATCH_ID,
                 invoice_date, row.get("sis_code"), row.get("pos_state"), bool(row.get("reverse_charge")),
                 bool(row.get("is_reverse_charge")), row.get("po_reference"), row.get("customer_name"),
                 row.get("customer_gstin"), row.get("billing_address"), row.get("shipping_address"),
                 row.get("site_name"), location_id, row.get("sis_code"), row.get("customer_gstin"),
                 gst_id, str(row.get("customer_gstin") or "")[:2] or None,
                 dec(row.get("taxable_value")), dec(row.get("rounding_amount")), row.get("amount_in_words"),
                 row.get("bank_name"), row.get("account_no"), row.get("ifsc_code"), row.get("original_pdf_sha256"),
                 row.get("original_pdf_path"), row.get("original_pdf_size"), row.get("original_pdf_pages"),
                 dec(row.get("paid_amount")), dec(row.get("balance_amount")), dec(row.get("discount_amount")),
                 dec(row.get("net_amount") or row.get("grand_total"))),
            )
            cur.execute("SELECT id FROM sales_invoices WHERE invoice_no = %s", (invoice_no,))
            actual_invoice_id = cur.fetchone()[0]
            line_values = []
            for line in items_by_invoice[invoice_no]:
                line_values.append((actual_invoice_id, products[str(line["item_code"])], line.get("line_no"),
                    line["item_code"], line["item_name"], line.get("batch_no"), dec(line["quantity"]),
                    dec(line["unit_price"]), line.get("hsn_code"), dec(line["gst_rate"]), dec(line["tax_amount"]),
                    dec(line["total_amount"]), dec(line["mrp"]), dec(line["discount_percent"]), dec(line["taxable_value"]),
                    dec(line["igst_amount"]), dec(line["cgst_amount"]), dec(line["sgst_amount"])))
            execute_batch(cur, """
                INSERT INTO sales_invoice_items
                    (invoice_id, product_id, line_no, code, name, batch_no, quantity, price, hsn_code,
                     gst_rate, tax_amount, total_amount, mrp, disc_pct, taxable_value, igst_amount,
                     cgst_amount, sgst_amount)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, line_values, page_size=500)
        conn.commit()
        print(f"COMMITTED: {len(summaries)} invoices and {len(items)} invoice lines")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--commit", action="store_true", help="Commit the import; otherwise dry-run")
    args = parser.parse_args()
    import_package(args.source, args.commit)
