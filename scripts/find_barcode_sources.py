#!/usr/bin/env python3
"""
Search for barcode data in all available Excel files
"""
import openpyxl
import os

files_to_check = [
    'Tax_Invoices_TT_18_to_137_Article_Color_Size_Split.xlsx',
    'Tax_Invoices_Full_Details_TT_18_to_137.xlsx',
    'Tax_Invoices_TT_18_to_137_Including_Cancelled.xlsx',
    'Tax_Invoices_TT_18_to_137_Updated_Bill136.xlsx',
]

print("=" * 80)
print("BARCODE SOURCE SEARCH - All Available Excel Files")
print("=" * 80)

barcode_found = False

for file_path in files_to_check:
    if not os.path.exists(file_path):
        print(f"\n❌ {file_path}: NOT FOUND")
        continue
    
    print(f"\n📄 {file_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        sheets = wb.sheetnames
        print(f"   Sheets: {sheets}")
        
        for sheet in sheets:
            ws = wb[sheet]
            headers = [cell.value for cell in ws[1]]
            
            # Check for barcode column
            barcode_col = None
            for i, header in enumerate(headers, 1):
                if header and 'barcode' in str(header).lower():
                    barcode_col = i
                    barcode_found = True
                    print(f"   ✓✓✓ Sheet '{sheet}' has BARCODE at column {i}!")
                    
                    # Show first few rows
                    print(f"       Sample barcodes:")
                    for row_idx in range(2, min(7, ws.max_row + 1)):
                        bc = ws.cell(row=row_idx, column=barcode_col).value
                        print(f"         Row {row_idx}: {bc}")
                    break
            
            if not barcode_col:
                print(f"   Sheet '{sheet}': {len(headers)} columns, no barcode")
        
        wb.close()
    except Exception as e:
        print(f"   ❌ Error: {e}")

print("\n" + "=" * 80)
if barcode_found:
    print("✓ BARCODE COLUMNS FOUND - Can use for import")
else:
    print("❌ NO BARCODE COLUMNS FOUND - Will use generated format (SKU-barcode)")
    print("   Status: Proceeding with barcode generation from SKU codes")
print("=" * 80)
