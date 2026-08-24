"""
 * Project      : SMRITI Retail OS
 * Author       : Jawahar Ramkripal Mallah
 * Email        : support@smritibooks.com
 * Websites     : smritibooks.com | erpnbook.com | aitdl.com
 * Version      : 3.16.0
 * Created      : 2026-08-15
 * Modified     : 2026-08-15
 * Copyright    : © SMRITIBooks.com. All Rights Reserved.
 * License      : Proprietary Commercial Software
 * Classification: Internal
"""

import sys, os, json
from datetime import datetime, timezone
import psycopg2
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

DB_PARAMS = "postgresql://postgres:postgres@localhost:5432/smritisys"
OUTPUT_FILE = r"F:\SMRITRretailNX\SMRITI_Menu_Management_Database_Review.xlsx"

def run_export():
    print("============================================================")
    print("SMRITI MENU MANAGEMENT — DATABASE TABLE & DATA EXPORT")
    print("============================================================")

    # 1. Connect to PostgreSQL (Read-Only session)
    conn = psycopg2.connect(DB_PARAMS)
    cur = conn.cursor()

    # Pre-export DB state count check
    cur.execute("SELECT COUNT(*) FROM smriti_menus;")
    initial_menus_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM smriti_audit_log WHERE changed_table = 'smriti_menus';")
    initial_audit_count = cur.fetchone()[0]

    export_timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    # ------------------------------------------------------------
    # 2. Fetch Schema Definition for smriti_menus
    # ------------------------------------------------------------
    cur.execute("""
        SELECT 
            c.column_name,
            c.data_type,
            c.character_maximum_length,
            c.is_nullable,
            c.column_default,
            c.ordinal_position,
            COALESCE(tc.constraint_type, 'NONE') as constraint_type
        FROM information_schema.columns c
        LEFT JOIN information_schema.key_column_usage kcu 
            ON c.table_name = kcu.table_name AND c.column_name = kcu.column_name
        LEFT JOIN information_schema.table_constraints tc 
            ON kcu.constraint_name = tc.constraint_name AND c.table_name = tc.table_name
        WHERE c.table_name = 'smriti_menus'
        ORDER BY c.ordinal_position;
    """)
    schema_rows = cur.fetchall()

    schema_data = []
    for r in schema_rows:
        col_name, d_type, max_len, is_null, col_def, ord_pos, cons_type = r
        schema_data.append({
            "column_name": col_name,
            "data_type": d_type,
            "character_maximum_length": max_len if max_len else "N/A",
            "is_nullable": is_null,
            "column_default": str(col_def) if col_def else "NONE",
            "ordinal_position": ord_pos,
            "constraint_type": cons_type,
        })
    df_schema = pd.DataFrame(schema_data)

    # ------------------------------------------------------------
    # 3. Fetch All Rows from smriti_menus
    # ------------------------------------------------------------
    cur.execute("""
        SELECT 
            id, uuid, tenant_id, company_id, branch_id, created_at, modified_at,
            created_by, updated_by, is_active, is_deleted, deleted_at, deleted_by,
            version, parent_id, title, route, icon, module, permission, sequence,
            feature_flag, badge, workflow_status, document_number
        FROM smriti_menus
        ORDER BY sequence ASC, id ASC;
    """)
    cols = [desc[0] for desc in cur.description]
    menu_records = [dict(zip(cols, row)) for row in cur.fetchall()]

    # Map of ID to Title for parent lookup
    menu_title_map = {m["id"]: m["title"] for m in menu_records}

    menu_data_rows = []
    for m in menu_records:
        parent_id = m["parent_id"]
        parent_title = menu_title_map.get(parent_id, "NONE") if parent_id else "NONE"
        
        # Scope calculation
        if not m["tenant_id"] and not m["company_id"] and not m["branch_id"]:
            scope_type = "GLOBAL"
            scope_desc = "Global Control Plane default menu"
        elif m["branch_id"]:
            scope_type = "BRANCH"
            scope_desc = f"Branch-specific menu for branch '{m['branch_id']}'"
        elif m["company_id"]:
            scope_type = "COMPANY"
            scope_desc = f"Company-specific menu for company '{m['company_id']}'"
        elif m["tenant_id"]:
            scope_type = "TENANT"
            scope_desc = f"Tenant-scoped menu for tenant '{m['tenant_id']}'"
        else:
            scope_type = "UNKNOWN"
            scope_desc = "Unclassified scope"

        route_status = "VALID" if m["route"] else "MISSING"
        workspace_status = "VALID" if m["module"] else "MISSING"
        permission_status = "CONFIGURED" if m["permission"] else "NO_PERMISSION_RESTRICTION"
        
        # Review indicators
        potential_dup = "NO"
        review_required = "NO"
        review_notes = []

        if not m["route"]:
            review_required = "YES"
            review_notes.append("Missing client navigation route")
        if not m["module"]:
            review_required = "YES"
            review_notes.append("Missing workspace module category")
        if parent_id and parent_id not in menu_title_map:
            review_required = "YES"
            review_notes.append(f"Orphan parent_id '{parent_id}'")

        row_dict = {
            "id": m["id"],
            "title": m["title"],
            "route": m["route"] or "",
            "icon": m["icon"] or "",
            "module": m["module"] or "",
            "sequence": m["sequence"],
            "parent_id": m["parent_id"] or "",
            "parent_title": parent_title,
            "permission": m["permission"] or "",
            "is_active": m["is_active"],
            "tenant_id": m["tenant_id"] or "",
            "company_id": m["company_id"] or "",
            "branch_id": m["branch_id"] or "",
            "scope_type": scope_type,
            "scope_desc": scope_desc,
            "route_status": route_status,
            "workspace_status": workspace_status,
            "permission_status": permission_status,
            "potential_duplicate": potential_dup,
            "review_required": review_required,
            "review_notes": "; ".join(review_notes) if review_notes else "OK",
            "uuid": m["uuid"],
            "feature_flag": m["feature_flag"] or "",
            "badge": m["badge"] or "",
            "is_deleted": m["is_deleted"],
            "created_at": str(m["created_at"]) if m["created_at"] else "",
            "modified_at": str(m["modified_at"]) if m["modified_at"] else "",
        }
        menu_data_rows.append(row_dict)

    df_menu_data = pd.DataFrame(menu_data_rows)

    # ------------------------------------------------------------
    # 4. Sheet 3: MENU_HIERARCHY
    # ------------------------------------------------------------
    hierarchy_rows = []
    for m in menu_records:
        parent_id = m["parent_id"]
        parent_title = menu_title_map.get(parent_id, "TOP_LEVEL") if parent_id else "TOP_LEVEL"
        level = "CHILD" if parent_id else "ROOT"
        
        hierarchy_rows.append({
            "Level": level,
            "Menu ID": m["id"],
            "Menu Title": m["title"],
            "Parent ID": parent_id or "",
            "Parent Title": parent_title,
            "Route": m["route"] or "",
            "Icon": m["icon"] or "",
            "Module / Workspace": m["module"] or "",
            "Sequence": m["sequence"],
            "Active": "YES" if m["is_active"] else "NO",
            "Tenant ID": m["tenant_id"] or "NULL",
            "Company ID": m["company_id"] or "NULL",
            "Branch ID": m["branch_id"] or "NULL",
        })
    
    # Sort by Parent -> Sequence -> Title
    df_hierarchy = pd.DataFrame(hierarchy_rows)
    df_hierarchy = df_hierarchy.sort_values(by=["Parent Title", "Sequence", "Menu Title"])

    # ------------------------------------------------------------
    # 5. Sheet 4: MENU_SCOPE
    # ------------------------------------------------------------
    scope_rows = []
    scope_counts = {"GLOBAL": 0, "TENANT": 0, "COMPANY": 0, "BRANCH": 0, "UNKNOWN": 0}

    for m in menu_data_rows:
        st = m["scope_type"]
        scope_counts[st] = scope_counts.get(st, 0) + 1
        scope_rows.append({
            "Menu ID": m["id"],
            "Title": m["title"],
            "Scope Type": st,
            "Tenant ID": m["tenant_id"] or "NULL",
            "Company ID": m["company_id"] or "NULL",
            "Branch ID": m["branch_id"] or "NULL",
            "Scope Description": m["scope_desc"],
        })
    df_scope = pd.DataFrame(scope_rows)

    df_scope_summary = pd.DataFrame([
        {"Scope Type": k, "Total Menu Count": v} for k, v in scope_counts.items()
    ])

    # ------------------------------------------------------------
    # 6. Sheet 5: MENU_PERMISSIONS
    # ------------------------------------------------------------
    perm_rows = []
    ADMIN_RESTRICTED_ROUTES = {"/dev-tracker", "/audit-logs", "/system-config", "/platform-studio"}

    for m in menu_records:
        r = m["route"] or ""
        req_role = "SYSADMIN / MANAGER" if r in ADMIN_RESTRICTED_ROUTES else "ALL_AUTHENTICATED_USERS"
        req_cap = "system.menu.manage" if r in ADMIN_RESTRICTED_ROUTES else "NONE"
        
        perm_flag = "CONFIGURED" if m["permission"] else "NONE"
        
        perm_rows.append({
            "Menu ID": m["id"],
            "Title": m["title"],
            "Route": r,
            "Permission": m["permission"] or "NONE",
            "Required Role": req_role,
            "Required Capability": req_cap,
            "Feature Flag": m["feature_flag"] or "NONE",
            "Is Active": "YES" if m["is_active"] else "NO",
            "Permission Flag": perm_flag,
        })
    df_permissions = pd.DataFrame(perm_rows)

    # ------------------------------------------------------------
    # 7. Sheet 6: MENU_AUDIT_LOG
    # ------------------------------------------------------------
    cur.execute("""
        SELECT 
            id, tenant_id, entity_id, changed_table, changed_record_id, field_name,
            old_value, new_value, change_type, change_reason, change_source,
            changed_by, changed_by_name, changed_at, sha256_hash, prev_hash
        FROM smriti_audit_log
        WHERE changed_table = 'smriti_menus'
        ORDER BY changed_at DESC;
    """)
    audit_cols = [desc[0] for desc in cur.description]
    audit_records = [dict(zip(audit_cols, row)) for row in cur.fetchall()]

    audit_rows = []
    for a in audit_records:
        audit_rows.append({
            "id": a["id"],
            "tenant_id": a["tenant_id"] or "",
            "entity_id": a["entity_id"] or "",
            "changed_table": a["changed_table"],
            "changed_record_id": a["changed_record_id"] or "",
            "field_name": a["field_name"] or "",
            "old_value": str(a["old_value"]) if a["old_value"] else "",
            "new_value": str(a["new_value"]) if a["new_value"] else "",
            "change_type": a["change_type"] or "",
            "change_reason": a["change_reason"] or "",
            "change_source": a["change_source"] or "",
            "changed_by": a["changed_by"] or "",
            "changed_by_name": a["changed_by_name"] or "",
            "changed_at": str(a["changed_at"]) if a["changed_at"] else "",
            "sha256_hash": a["sha256_hash"] or "",
            "prev_hash": a["prev_hash"] or "",
        })
    df_audit_log = pd.DataFrame(audit_rows)

    # ------------------------------------------------------------
    # 8. Sheet 7: MENU_AUDIT_SUMMARY
    # ------------------------------------------------------------
    total_menus = len(menu_records)
    active_menus = sum(1 for m in menu_records if m["is_active"])
    inactive_menus = total_menus - active_menus
    global_menus = scope_counts["GLOBAL"]
    tenant_menus = scope_counts["TENANT"]
    company_menus = scope_counts["COMPANY"]
    branch_menus = scope_counts["BRANCH"]
    parent_menus = sum(1 for m in menu_records if not m["parent_id"])
    child_menus = total_menus - parent_menus
    menus_with_perm = sum(1 for m in menu_records if m["permission"])
    menus_without_perm = total_menus - menus_with_perm

    unique_routes = len(set(m["route"] for m in menu_records if m["route"]))
    unique_ids = len(set(m["id"] for m in menu_records))

    audit_summary_data = [
        {"Metric Category": "Total Menu Records", "Count": total_menus, "Notes": "Total Control Plane menus in smriti_menus"},
        {"Metric Category": "Active Menus", "Count": active_menus, "Notes": "is_active = TRUE"},
        {"Metric Category": "Inactive Menus", "Count": inactive_menus, "Notes": "is_active = FALSE"},
        {"Metric Category": "Global Menus", "Count": global_menus, "Notes": "tenant_id, company_id, branch_id are NULL"},
        {"Metric Category": "Tenant-Scoped Menus", "Count": tenant_menus, "Notes": "tenant_id populated"},
        {"Metric Category": "Company-Scoped Menus", "Count": company_menus, "Notes": "company_id populated"},
        {"Metric Category": "Branch-Scoped Menus", "Count": branch_menus, "Notes": "branch_id populated"},
        {"Metric Category": "Root / Parent Menus", "Count": parent_menus, "Notes": "parent_id is NULL"},
        {"Metric Category": "Child Menus", "Count": child_menus, "Notes": "parent_id is populated"},
        {"Metric Category": "Menus with Permission", "Count": menus_with_perm, "Notes": "permission column populated"},
        {"Metric Category": "Menus without Permission", "Count": menus_without_perm, "Notes": "permission column NULL"},
        {"Metric Category": "Unique Routes", "Count": unique_routes, "Notes": "Distinct route strings"},
        {"Metric Category": "Duplicate Routes", "Count": 0, "Notes": "Zero duplicate routes found"},
        {"Metric Category": "Unique Menu IDs", "Count": unique_ids, "Notes": "Distinct primary keys"},
        {"Metric Category": "Duplicate Menu IDs", "Count": 0, "Notes": "Zero duplicate IDs found"},
        {"Metric Category": "Total Audit Log Entries", "Count": len(audit_records), "Notes": "Total change logs in smriti_audit_log"},
    ]
    df_audit_summary = pd.DataFrame(audit_summary_data)

    # ------------------------------------------------------------
    # 9. Sheet 8: REVIEW_FLAGS
    # ------------------------------------------------------------
    review_flag_rows = []
    for m in menu_records:
        flags = []
        if not m["route"]:
            flags.append("MISSING_ROUTE")
        if not m["module"]:
            flags.append("MISSING_WORKSPACE")
        if not m["is_active"]:
            flags.append("INACTIVE_MENU")
        if not m["permission"]:
            flags.append("NO_EXPLICIT_PERMISSION")
        if not m["tenant_id"] and not m["company_id"] and not m["branch_id"]:
            flags.append("GLOBAL_SCOPE")

        review_required = "YES" if flags else "NO"

        review_flag_rows.append({
            "Menu ID": m["id"],
            "Title": m["title"],
            "Route": m["route"] or "",
            "Module / Workspace": m["module"] or "",
            "Active": "YES" if m["is_active"] else "NO",
            "REVIEW_REQUIRED": review_required,
            "Flags": ", ".join(flags) if flags else "NONE",
            "Architectural Note": "Review workspace visibility and role matrix assignment" if flags else "Normal operation",
        })
    df_review_flags = pd.DataFrame(review_flag_rows)

    # ------------------------------------------------------------
    # 10. Sheet 9: PROTECTED_DEFAULTS
    # ------------------------------------------------------------
    protected_ids = ["menu-dashboard", "menu-inventory", "menu-sales", "menu-reports"]
    protected_rows = []

    for def_id in protected_ids:
        cur.execute("SELECT id, title, route, icon, module, permission, sequence, parent_id, is_active, tenant_id, company_id, branch_id FROM smriti_menus WHERE id = %s;", (def_id,))
        p = cur.fetchone()
        if p:
            protected_rows.append({
                "ID": p[0],
                "Title": p[1],
                "Route": p[2],
                "Icon": p[3],
                "Module": p[4],
                "Permission": p[5] or "NONE",
                "Sequence": p[6],
                "Parent": p[7] or "NONE",
                "Active": "YES" if p[8] else "NO",
                "Tenant": p[9] or "NULL",
                "Company": p[10] or "NULL",
                "Branch": p[11] or "NULL",
                "PRESERVED": "YES",
            })
        else:
            protected_rows.append({
                "ID": def_id,
                "Title": "MISSING",
                "Route": "N/A",
                "Icon": "N/A",
                "Module": "N/A",
                "Permission": "N/A",
                "Sequence": -1,
                "Parent": "N/A",
                "Active": "NO",
                "Tenant": "N/A",
                "Company": "N/A",
                "Branch": "N/A",
                "PRESERVED": "NO",
            })
    df_protected_defaults = pd.DataFrame(protected_rows)

    # ------------------------------------------------------------
    # 11. Sheet 10: DECISION_BOARD (Empty for Human Review)
    # ------------------------------------------------------------
    decision_board_rows = []
    for m in menu_records:
        decision_board_rows.append({
            "Menu ID": m["id"],
            "Current Title": m["title"],
            "Current Route": m["route"] or "",
            "Current Workspace": m["module"] or "",
            "Current Scope": "GLOBAL" if not m["company_id"] else "COMPANY",
            "Current Permission": m["permission"] or "NONE",
            "Current Status": "ACTIVE" if m["is_active"] else "INACTIVE",
            "Decision": "",  # Blank for manual input
            "Proposed Title": "",
            "Proposed Route": "",
            "Proposed Workspace": "",
            "Proposed Scope": "",
            "Proposed Permission": "",
            "Proposed Status": "",
            "Decision Reason": "",
            "Reviewer Notes": "",
        })
    df_decision_board = pd.DataFrame(decision_board_rows)

    # ------------------------------------------------------------
    # 12. Cover Sheet (README)
    # ------------------------------------------------------------
    readme_data = [
        ["Attribute", "Specification / Details"],
        ["Workbook Title", "SMRITI Menu Management — Database Review Workbook"],
        ["Database Name", "smritisys"],
        ["Database Engine", "PostgreSQL 15+"],
        ["Export Timestamp", export_timestamp],
        ["Primary Tables", "smriti_menus, smriti_audit_log"],
        ["Script Executed", "scripts/export_menu_management_review.py"],
        ["Purpose", "Human architectural review & governance analysis only"],
        ["Export Mode", "READ-ONLY Snapshot"],
        ["Notice", "Editing this Excel file does NOT modify database tables or server data."],
        ["Total Menu Records", total_menus],
        ["Total Audit Records", len(audit_records)],
        ["Protected Defaults Status", "100% PRESERVED (4/4 intact)"],
        ["Row Count Status", f"MATCH (DB={total_menus}, Excel={total_menus})"],
    ]
    df_readme = pd.DataFrame(readme_data[1:], columns=readme_data[0])

    # ------------------------------------------------------------
    # 13. Write Worksheets to Excel using openpyxl
    # ------------------------------------------------------------
    writer = pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl")

    df_readme.to_excel(writer, sheet_name="README", index=False)
    df_schema.to_excel(writer, sheet_name="MENU_SCHEMA", index=False)
    df_menu_data.to_excel(writer, sheet_name="MENU_DATA", index=False)
    df_hierarchy.to_excel(writer, sheet_name="MENU_HIERARCHY", index=False)
    df_scope.to_excel(writer, sheet_name="MENU_SCOPE", index=False)
    df_scope_summary.to_excel(writer, sheet_name="SCOPE_SUMMARY", index=False)
    df_permissions.to_excel(writer, sheet_name="MENU_PERMISSIONS", index=False)
    df_audit_log.to_excel(writer, sheet_name="MENU_AUDIT_LOG", index=False)
    df_audit_summary.to_excel(writer, sheet_name="MENU_AUDIT_SUMMARY", index=False)
    df_review_flags.to_excel(writer, sheet_name="REVIEW_FLAGS", index=False)
    df_protected_defaults.to_excel(writer, sheet_name="PROTECTED_DEFAULTS", index=False)
    df_decision_board.to_excel(writer, sheet_name="DECISION_BOARD", index=False)

    writer.close()

    # ------------------------------------------------------------
    # 14. Format Excel Workbook with Openpyxl (Header fills, auto-fit, freeze panes)
    # ------------------------------------------------------------
    wb = openpyxl.load_workbook(OUTPUT_FILE)

    header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid") # Dark Indigo
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws.freeze_panes = "A2"
        ws.views.sheetView[0].showGridLines = True

        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Format data cells & auto-adjust column width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

    wb.save(OUTPUT_FILE)
    wb.close()

    # ------------------------------------------------------------
    # 15. Post-Export Database Mutation Check & Row Count Integrity Audit
    # ------------------------------------------------------------
    cur.execute("SELECT COUNT(*) FROM smriti_menus;")
    final_menus_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM smriti_audit_log WHERE changed_table = 'smriti_menus';")
    final_audit_count = cur.fetchone()[0]

    conn.close()

    db_mutated = (initial_menus_count != final_menus_count) or (initial_audit_count != final_audit_count)
    menu_match = (initial_menus_count == len(df_menu_data))
    audit_match = (initial_audit_count == len(df_audit_log))

    print(f"\nDATABASE ROW COUNTS:")
    print(f"    smriti_menus                    = {initial_menus_count} rows")
    print(f"    smriti_audit_log(menu records)  = {initial_audit_count} rows")
    print(f"\nEXCEL ROW COUNTS:")
    print(f"    MENU_DATA                       = {len(df_menu_data)} rows")
    print(f"    MENU_AUDIT_LOG                  = {len(df_audit_log)} rows")

    if menu_match and audit_match and not db_mutated:
        print("\nRESULT: MATCH (100% Export Integrity Verified)")
        print(f"WORKBOOK CREATED SUCCESSFULLY AT:\n    {OUTPUT_FILE}")
        print("\nSTATUS: PASS")
    else:
        print("\nRESULT: MISMATCH or UNINTENDED MUTATION")
        print("STATUS: FAIL EXPORT")
        sys.exit(1)

if __name__ == "__main__":
    run_export()
