"""
Project      : SMRITI Retail OS
Author       : Jawahar Ramkripal Mallah
Designation  : Chief Systems Architect & Creator
Email        : support@smritibooks.com
Websites     : smritibooks.com | erpnbook.com | aitdl.com
Version      : 5.1.0
Created      : 2026-08-25
Modified     : 2026-08-25
Copyright    : © SMRITIBooks.com. All Rights Reserved.
License      : Proprietary Commercial Software
Classification: Internal
"""

import os
import sys
import hashlib
import json
import datetime
from decimal import Decimal
import psycopg2
import psycopg2.extras
import asyncio
from playwright.async_api import async_playwright
import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

sys.path.insert(0, r"F:\SMRITRretailNX\backend")
from app.services.invoice_pdf_service import InvoicePdfService, number_to_indian_words

DB_URL = "postgresql://postgres:postgres@localhost:5432/smriti001"

# Target SIS 8319 / S4NN info
SIS_CODE = "S4NN/8319"
SITE_NAME = "Reliance Retail Limited (Distribution Center)"
BILLING_ADDRESS = """NO 62/2,RIL BUILIDING
RICHMOND ROAD,
BANGALORE- 560025 Karnataka, INDIA"""

SHIPPING_ADDRESS = """Distribution Center
Survey No 54 1 Nandihalli Village
55th KM Stone NH 4 Tumkur Road
Oordigree Hobli Taluka
TUMKUR, Karnataka - 572101
Tel :
GSTN No : 29AABCR1718E1ZL
EMAIL : Fp_Dc_KA.TUMKUR_S4NN@zmail.ril.com"""

CUSTOMER_NAME = "Reliance Retail Limited"
CUSTOMER_GSTIN = "29AABCR1718E1ZL"
POS_STATE = "KARNATAKA"
PO_REFERENCE = "5182778158"

OUTPUT_CLIENT_DIR = r"F:\Smriti-Clients Data\Tattly Threads\24_08_26\Invoices"
OUTPUT_EXPORTS_DIR = r"F:\SMRITRretailNX\exports\tt_batch_129_137"
OUTPUT_TT_DIR = r"F:\SMRITRretailNX\TT"
DISPATCH_PATH = r"F:\Smriti-Clients Data\Tattly Threads\24_08_26\RIL_Dispatch_09-08-2026-2.xlsx"


class InvoiceItemMock:
    def __init__(self, line_no, code, name, qty, price, mrp, disc_pct, taxable_value, hsn_code, gst_rate, tax_amount, cgst_amount, sgst_amount, igst_amount, total_amount):
        self.line_no = line_no
        self.code = code
        self.name = name
        self.quantity = qty
        self.price = price
        self.mrp = mrp
        self.disc_pct = disc_pct
        self.taxable_value = taxable_value
        self.hsn_code = hsn_code
        self.gst_rate = gst_rate
        self.tax_amount = tax_amount
        self.cgst_amount = cgst_amount
        self.sgst_amount = sgst_amount
        self.igst_amount = igst_amount
        self.total_amount = total_amount


class InvoiceMock:
    def __init__(self, id, invoice_no, date, sis_code, pos_state, po_reference, eway_bill_no, customer_name, customer_gstin, site_name, billing_address, shipping_address, taxable_value, tax_total, grand_total, is_interstate, bank_name, account_no, ifsc_code, bank_branch, items):
        self.id = id
        self.invoice_no = invoice_no
        self.date = date
        self.sis_code = sis_code
        self.pos_state = pos_state
        self.po_reference = po_reference
        self.eway_bill_no = eway_bill_no
        self.customer_name = customer_name
        self.customer_gstin = customer_gstin
        self.site_name = site_name
        self.billing_address = billing_address
        self.shipping_address = shipping_address
        self.taxable_value = taxable_value
        self.tax_total = tax_total
        self.grand_total = grand_total
        self.is_interstate = is_interstate
        self.bank_name = bank_name
        self.account_no = account_no
        self.ifsc_code = ifsc_code
        self.bank_branch = bank_branch
        self.items = items
        self.status = "COMPLETED"
        self.reverse_charge = False
        self.is_reverse_charge = False
        self.irn = None
        self.signed_qr_payload = None
        self.e_invoice_status = "NOT_APPLICABLE"


async def update_bill_136():
    print("==================================================")
    print("UPDATING BILL 136 TO SIS CODE 'S4NN/8319' & ADDRESS")
    print("==================================================")

    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # 1. Fetch current invoice 136 and items
    cur.execute("SELECT * FROM sales_invoices WHERE invoice_no = 'TT2026-2027/136' OR id = 'inv-tt-2026-2027-136';")
    inv_row = cur.fetchone()
    if not inv_row:
        print("ERROR: Invoice 136 not found in smriti001!")
        return

    inv_id = inv_row['id']
    inv_no = inv_row['invoice_no']
    print(f"Found invoice: ID={inv_id}, No={inv_no}")

    cur.execute("SELECT * FROM sales_invoice_items WHERE invoice_id = %s ORDER BY line_no;", (inv_id,))
    item_rows = cur.fetchall()
    print(f"Loaded {len(item_rows)} line items.")

    item_mocks = []
    for itm in item_rows:
        item_mocks.append(InvoiceItemMock(
            line_no=itm['line_no'],
            code=itm['code'],
            name=itm['name'],
            qty=itm['quantity'],
            price=itm['price'],
            mrp=itm['mrp'],
            disc_pct=itm['disc_pct'],
            taxable_value=itm['taxable_value'],
            hsn_code=itm['hsn_code'],
            gst_rate=itm['gst_rate'],
            tax_amount=itm['tax_amount'],
            cgst_amount=itm['cgst_amount'],
            sgst_amount=itm['sgst_amount'],
            igst_amount=itm['igst_amount'],
            total_amount=itm['total_amount']
        ))

    # 2. Build Mock Model
    inv_mock = InvoiceMock(
        id=inv_id,
        invoice_no=inv_no,
        date=inv_row['date'],
        sis_code=SIS_CODE,
        pos_state=POS_STATE,
        po_reference=PO_REFERENCE,
        eway_bill_no=inv_row.get('eway_bill_no') or "",
        customer_name=CUSTOMER_NAME,
        customer_gstin=CUSTOMER_GSTIN,
        site_name=SITE_NAME,
        billing_address=BILLING_ADDRESS,
        shipping_address=SHIPPING_ADDRESS,
        taxable_value=inv_row['taxable_value'],
        tax_total=inv_row['tax_total'],
        grand_total=inv_row['grand_total'],
        is_interstate=inv_row['is_interstate'],
        bank_name=inv_row['bank_name'] or "STATE BANK OF INDIA",
        account_no=inv_row['account_no'] or "43976711765",
        ifsc_code=inv_row['ifsc_code'] or "SBIN0030425",
        bank_branch="WARDHMAN NAGAR NAGPUR",
        items=item_mocks
    )

    # 3. Generate HTML & Render PDF with Playwright
    html_content = InvoicePdfService.generate_invoice_html_from_model(
        invoice=inv_mock,
        company_name="TATTLY THREADS",
        company_gstin="27AAXFT2508H1ZR",
        extra_meta={
            "company_website": "www.tattlythreads.com",
            "dispatch_email": "dispatch@tattlythreads.com",
            "accounts_email": "accounts@tattlythreads.com"
        }
    )

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()
        await page.set_content(html_content, wait_until="networkidle")
        pdf_bytes = await page.pdf(
            format="A4",
            print_background=True,
            margin={"top": "8mm", "bottom": "10mm", "left": "8mm", "right": "8mm"}
        )
        await browser.close()

    sha256_hash = hashlib.sha256(pdf_bytes).hexdigest()
    print(f"Rendered PDF: {len(pdf_bytes)} bytes | SHA-256: {sha256_hash}")

    # File naming variations
    filenames = [
        "SIS_8319_TaxInvoice_TT2026-2027_136.pdf",
        "SIS_S4NN_8319_TaxInvoice_TT2026-2027_136.pdf",
        "S4NN_8319_TT2026-2027_136.pdf"
    ]

    for dir_path in [OUTPUT_CLIENT_DIR, OUTPUT_EXPORTS_DIR, OUTPUT_TT_DIR]:
        if os.path.exists(dir_path):
            for fn in filenames:
                fp = os.path.join(dir_path, fn)
                with open(fp, "wb") as f:
                    f.write(pdf_bytes)
                print(f"  ✓ Saved {fp}")
            # Clean up legacy TW07 file if present
            legacy_fp = os.path.join(dir_path, "SIS_TW07_TaxInvoice_TT2026-2027_136.pdf")
            if os.path.exists(legacy_fp):
                try:
                    os.remove(legacy_fp)
                    print(f"  ✓ Removed legacy file {legacy_fp}")
                except Exception as e:
                    print(f"  Notice removing legacy: {e}")

    # 4. Update Database
    primary_pdf_path = os.path.join(OUTPUT_CLIENT_DIR, "SIS_8319_TaxInvoice_TT2026-2027_136.pdf")
    cur.execute("""
        UPDATE sales_invoices
        SET sis_code = %s,
            site_name = %s,
            billing_address = %s,
            shipping_address = %s,
            customer_name = %s,
            customer_gstin = %s,
            pos_state = %s,
            po_reference = %s,
            original_pdf_sha256 = %s,
            original_pdf_path = %s,
            original_pdf_size = %s,
            modified_at = NOW()
        WHERE id = %s OR invoice_no = %s;
    """, (
        SIS_CODE, SITE_NAME, BILLING_ADDRESS, SHIPPING_ADDRESS,
        CUSTOMER_NAME, CUSTOMER_GSTIN, POS_STATE, PO_REFERENCE,
        sha256_hash, primary_pdf_path, len(pdf_bytes),
        inv_id, inv_no
    ))

    # Update artifact if exists
    cur.execute("""
        UPDATE invoice_document_artifacts
        SET storage_path = %s,
            sha256_hash = %s,
            file_size = %s,
            modified_at = NOW()
        WHERE invoice_id = %s OR invoice_no = %s;
    """, (primary_pdf_path, sha256_hash, len(pdf_bytes), inv_id, inv_no))

    conn.commit()
    conn.close()
    print("✓ Updated smriti001 database record successfully.")

    # 5. Update RIL_Dispatch_09-08-2026-2.xlsx Sheet2
    if os.path.exists(DISPATCH_PATH):
        try:
            wb_disp = openpyxl.load_workbook(DISPATCH_PATH)
            if "Sheet2" in wb_disp.sheetnames:
                ws2 = wb_disp["Sheet2"]
                for r in range(1, ws2.max_row + 1):
                    val_o = ws2.cell(r, 15).value
                    if val_o and ("136" in str(val_o) or val_o == inv_no):
                        ws2.cell(r, 1, value=8319)
                        print(f"  ✓ Updated Sheet2 Row {r} Column A to 8319")
                wb_disp.save(DISPATCH_PATH)
                print(f"✓ Saved updated dispatch workbook {DISPATCH_PATH}")
        except Exception as e:
            print(f"Notice on dispatch workbook update: {e}")

    print("==================================================")
    print("BILL 136 SUCCESSFULLY UPDATED TO SIS CODE 'S4NN/8319'")
    print("==================================================")


if __name__ == "__main__":
    asyncio.run(update_bill_136())
