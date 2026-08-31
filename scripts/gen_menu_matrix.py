"""
 * Project      : SMRITI Retail OS
 * Author       : Jawahar Ramkripal Mallah
 * Email        : support@smritibooks.com
 * Websites     : smritibooks.com | erpnbook.com | aitdl.com
 * Version      : 3.16.0
 * Created      : 2026-08-15
 * Modified     : 2026-08-15
 * Copyright    : © SMRITIBooks.com. All Rights Reserved.
 * License      : Proprietary Commercial Software
 * Classification: Internal
"""

import sys, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

FILE_PATH = r"F:\SMRITRretailNX\SMRITI_Menu_Management_Database_Review.xlsx"

# Refined Target Navigation Model v1.0 across 8 Clean Groups
TARGET_MODEL = [
    # 1. Executive Operations
    {
        "id": "menu-dashboard",
        "action": "KEEP",
        "proposed_title": "Dashboard & Executive Hub",
        "proposed_route": "/dashboard",
        "proposed_group": "Dashboard & Operations",
        "proposed_scope": "GLOBAL",
        "proposed_parent": "ROOT",
        "visibility_model": "GLOBAL",
        "req_capability": "GLOBAL_AUTHENTICATED",
        "access_perm": "DASHBOARD.ACCESS",
        "single_ws_notes": "Root entry portal. All authenticated users land here.",
        "arch_notes": "Preserved system default root menu."
    },
    {
        "id": "menu-user-profile",
        "action": "KEEP",
        "proposed_title": "My Profile Dashboard",
        "proposed_route": "/user-profile",
        "proposed_group": "Dashboard & Operations",
        "proposed_scope": "GLOBAL",
        "proposed_parent": "ROOT",
        "visibility_model": "GLOBAL",
        "req_capability": "GLOBAL_AUTHENTICATED",
        "access_perm": "PROFILE.ACCESS",
        "single_ws_notes": "Self-service account preferences & user configuration.",
        "arch_notes": "Global self-service user workspace."
    },

    # 2. System & Knowledge Base
    {
        "id": "menu-wiki",
        "action": "KEEP",
        "proposed_title": "SMRITI Gyan Kendra",
        "proposed_route": "/wiki",
        "proposed_group": "System & Knowledge Base",
        "proposed_scope": "GLOBAL",
        "proposed_parent": "ROOT",
        "visibility_model": "GLOBAL",
        "req_capability": "GLOBAL_AUTHENTICATED",
        "access_perm": "WIKI.ACCESS",
        "single_ws_notes": "Internal documentation & process training manual.",
        "arch_notes": "Available to all internal employees."
    },
    {
        "id": "menu-about-smriti",
        "action": "KEEP",
        "proposed_title": "About SMRITI Retail OS",
        "proposed_route": "/about-smriti",
        "proposed_group": "System & Knowledge Base",
        "proposed_scope": "GLOBAL",
        "proposed_parent": "ROOT",
        "visibility_model": "GLOBAL",
        "req_capability": "GLOBAL_AUTHENTICATED",
        "access_perm": "ABOUT.ACCESS",
        "single_ws_notes": "System metadata & architectural attribution.",
        "arch_notes": "Informational system drawer/footer access."
    },
    {
        "id": "menu-dev-tracker",
        "action": "HIDE",
        "proposed_title": "Dev Intelligence Center",
        "proposed_route": "/dev-tracker",
        "proposed_group": "System & Knowledge Base",
        "proposed_scope": "GLOBAL",
        "proposed_parent": "ROOT",
        "visibility_model": "CAPABILITY_BASED",
        "req_capability": "system.dev.tracker",
        "access_perm": "SYSTEM.DEV",
        "single_ws_notes": "Developer diagnostics & live build logs.",
        "arch_notes": "Hidden from standard business users. Restricted to Dev/SysAdmin."
    },

    # 3. Sales & POS Workspace Group
    {
        "id": "menu-pos",
        "action": "PARENT",
        "proposed_title": "Billing Desk (Universal POS)",
        "proposed_route": "/pos",
        "proposed_group": "Sales & POS",
        "proposed_scope": "COMPANY",
        "proposed_parent": "ROOT",
        "visibility_model": "ROLE_BASED",
        "req_capability": "pos.workspace.access",
        "access_perm": "POS.WORKSPACE.ACCESS",
        "single_ws_notes": "Single Billing Workspace parent menu. Inside-workspace actions (Create Bill, Apply Discount, Return) governed separately by granular capabilities.",
        "arch_notes": "Parent container for Sales & Billing module."
    },
    {
        "id": "menu-sales",
        "action": "CHILD",
        "proposed_title": "Sales Studio & Ledger",
        "proposed_route": "/sales",
        "proposed_group": "Sales & POS",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-pos",
        "visibility_model": "ROLE_BASED",
        "req_capability": "sales.studio.access",
        "access_perm": "SALES.WORKSPACE.ACCESS",
        "single_ws_notes": "Sales invoice ledger & history view.",
        "arch_notes": "Child item under Sales & POS hierarchy."
    },
    {
        "id": "menu-customer-master",
        "action": "CHILD",
        "proposed_title": "Customer Master Directory",
        "proposed_route": "/customer-master",
        "proposed_group": "Sales & POS",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-pos",
        "visibility_model": "ROLE_BASED",
        "req_capability": "customer.master.access",
        "access_perm": "CUSTOMER.WORKSPACE.ACCESS",
        "single_ws_notes": "Customer CRM directory.",
        "arch_notes": "Child item under Sales & POS hierarchy."
    },
    {
        "id": "menu-crm",
        "action": "CHILD",
        "proposed_title": "CRM & Engagement Studio",
        "proposed_route": "/crm",
        "proposed_group": "Sales & POS",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-pos",
        "visibility_model": "ROLE_BASED",
        "req_capability": "crm.studio.access",
        "access_perm": "CRM.WORKSPACE.ACCESS",
        "single_ws_notes": "Customer relationship management.",
        "arch_notes": "Child item under Sales & POS hierarchy."
    },
    {
        "id": "menu-loyalty",
        "action": "CHILD",
        "proposed_title": "Loyalty & Rewards Studio",
        "proposed_route": "/loyalty",
        "proposed_group": "Sales & POS",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-pos",
        "visibility_model": "ROLE_BASED",
        "req_capability": "loyalty.studio.access",
        "access_perm": "LOYALTY.WORKSPACE.ACCESS",
        "single_ws_notes": "Loyalty program administration.",
        "arch_notes": "Child item under Sales & POS hierarchy."
    },
    {
        "id": "menu-profiles",
        "action": "CHILD",
        "proposed_title": "POS Terminal Profiles",
        "proposed_route": "/profiles",
        "proposed_group": "Sales & POS",
        "proposed_scope": "BRANCH",
        "proposed_parent": "menu-pos",
        "visibility_model": "ROLE_BASED",
        "req_capability": "pos.terminals.manage",
        "access_perm": "TERMINALS.MANAGE",
        "single_ws_notes": "Hardware terminal setup for specific physical store branches.",
        "arch_notes": "Future applicability decision for branch-scoped terminal setup."
    },

    # 4. Inventory & Purchase Workspace Group
    {
        "id": "menu-inventory",
        "action": "PARENT",
        "proposed_title": "Inventory Workspace",
        "proposed_route": "/inventory",
        "proposed_group": "Inventory & Purchase",
        "proposed_scope": "COMPANY",
        "proposed_parent": "ROOT",
        "visibility_model": "ROLE_BASED",
        "req_capability": "inventory.workspace.access",
        "access_perm": "INVENTORY.WORKSPACE.ACCESS",
        "single_ws_notes": "Single Inventory Workspace parent menu. Group root for Stock, Items, Barcodes.",
        "arch_notes": "Preserved system default parent container."
    },
    {
        "id": "menu-item-master",
        "action": "CHILD",
        "proposed_title": "Item Master Catalog",
        "proposed_route": "/item-master",
        "proposed_group": "Inventory & Purchase",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-inventory",
        "visibility_model": "ROLE_BASED",
        "req_capability": "item.master.access",
        "access_perm": "ITEM.WORKSPACE.ACCESS",
        "single_ws_notes": "Product catalog and variant setup.",
        "arch_notes": "Child item under Inventory hierarchy."
    },
    {
        "id": "menu-barcode",
        "action": "CHILD",
        "proposed_title": "Barcode Studio & Generator",
        "proposed_route": "/barcode",
        "proposed_group": "Inventory & Purchase",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-inventory",
        "visibility_model": "ROLE_BASED",
        "req_capability": "barcode.studio.access",
        "access_perm": "BARCODE.WORKSPACE.ACCESS",
        "single_ws_notes": "Barcode printing & mapping utility.",
        "arch_notes": "Child item under Inventory hierarchy."
    },
    {
        "id": "menu-stock-ledger",
        "action": "CHILD",
        "proposed_title": "Stock Movements & Ledger",
        "proposed_route": "/stock-ledger",
        "proposed_group": "Inventory & Purchase",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-inventory",
        "visibility_model": "ROLE_BASED",
        "req_capability": "stock.ledger.access",
        "access_perm": "STOCK.WORKSPACE.ACCESS",
        "single_ws_notes": "Physical stock movements, GRN, and adjustments.",
        "arch_notes": "Child item under Inventory hierarchy."
    },
    {
        "id": "menu-purchase",
        "action": "CHILD",
        "proposed_title": "Purchase Studio & Orders",
        "proposed_route": "/purchase",
        "proposed_group": "Inventory & Purchase",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-inventory",
        "visibility_model": "ROLE_BASED",
        "req_capability": "purchase.workspace.access",
        "access_perm": "PURCHASE.WORKSPACE.ACCESS",
        "single_ws_notes": "Single Purchase Workspace PO workflow.",
        "arch_notes": "Child item under Inventory / Purchase hierarchy."
    },
    {
        "id": "menu-supplier-mgmt",
        "action": "CHILD",
        "proposed_title": "Supplier / Person Master",
        "proposed_route": "/supplier-mgmt",
        "proposed_group": "Inventory & Purchase",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-inventory",
        "visibility_model": "ROLE_BASED",
        "req_capability": "supplier.master.access",
        "access_perm": "SUPPLIER.WORKSPACE.ACCESS",
        "single_ws_notes": "Universal Person Workspace supplier & vendor directory.",
        "arch_notes": "Renamed label to Supplier / Person to reflect Universal Person Workspace architecture."
    },

    # 5. Accounts & Financial Sync
    {
        "id": "menu-business-ledger",
        "action": "KEEP",
        "proposed_title": "Business Ledger & Statements",
        "proposed_route": "/business-ledger",
        "proposed_group": "Accounts",
        "proposed_scope": "COMPANY",
        "proposed_parent": "ROOT",
        "visibility_model": "ROLE_BASED",
        "req_capability": "accounts.ledger.access",
        "access_perm": "ACCOUNTS.WORKSPACE.ACCESS",
        "single_ws_notes": "Financial ledger statements and trial balance.",
        "arch_notes": "Restricted to Accounting & Managerial roles."
    },
    {
        "id": "menu-accounting-sync",
        "action": "KEEP",
        "proposed_title": "Tally / ERP Accounting Sync",
        "proposed_route": "/accounting-sync",
        "proposed_group": "Accounts",
        "proposed_scope": "COMPANY",
        "proposed_parent": "ROOT",
        "visibility_model": "ROLE_BASED",
        "req_capability": "accounts.sync.execute",
        "access_perm": "ACCOUNTS.SYNC.EXECUTE",
        "single_ws_notes": "Third-party ERP integration sync gateway.",
        "arch_notes": "Restricted to SYSADMIN & Senior Accountant."
    },

    # 6. Reports & Analytics
    {
        "id": "menu-reports",
        "action": "PARENT",
        "proposed_title": "Reports Portal & Analytics",
        "proposed_route": "/reports",
        "proposed_group": "Reports",
        "proposed_scope": "COMPANY",
        "proposed_parent": "ROOT",
        "visibility_model": "ROLE_BASED",
        "req_capability": "reports.portal.access",
        "access_perm": "REPORT.WORKSPACE.ACCESS",
        "single_ws_notes": "Centralized reporting hub.",
        "arch_notes": "Parent container for custom report builder."
    },
    {
        "id": "menu-report-designer",
        "action": "CHILD",
        "proposed_title": "Visual Report Designer",
        "proposed_route": "/report-designer",
        "proposed_group": "Reports",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-reports",
        "visibility_model": "ROLE_BASED",
        "req_capability": "reports.designer.access",
        "access_perm": "REPORT.DESIGN.ACCESS",
        "single_ws_notes": "Custom report layout & query authoring studio.",
        "arch_notes": "Child of Reports Portal."
    },

    # 7. Configuration & Governance (Renamed from Master Framework)
    {
        "id": "menu-masters",
        "action": "PARENT",
        "proposed_title": "Configuration & Governance Hub",
        "proposed_route": "/masters",
        "proposed_group": "Configuration & Governance",
        "proposed_scope": "COMPANY",
        "proposed_parent": "ROOT",
        "visibility_model": "ROLE_BASED",
        "req_capability": "config.governance.access",
        "access_perm": "CONFIG.GOVERNANCE.ACCESS",
        "single_ws_notes": "Renamed parent group from Master Framework to Configuration & Governance for clean architectural taxonomy.",
        "arch_notes": "Parent container for configuration, engines, and system rules."
    },
    {
        "id": "menu-ufe",
        "action": "CHILD",
        "proposed_title": "Universal Field Explorer (UFE)",
        "proposed_route": "/ufe",
        "proposed_group": "Configuration & Governance",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-masters",
        "visibility_model": "ROLE_BASED",
        "req_capability": "ufe.explorer.access",
        "access_perm": "UFE.ACCESS",
        "single_ws_notes": "Dynamic form field attribute registry.",
        "arch_notes": "Child of Configuration & Governance."
    },
    {
        "id": "menu-formulas",
        "action": "CHILD",
        "proposed_title": "Formula & KPI Registry",
        "proposed_route": "/formulas",
        "proposed_group": "Configuration & Governance",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-masters",
        "visibility_model": "ROLE_BASED",
        "req_capability": "kpi.formulas.manage",
        "access_perm": "FORMULA.MANAGE",
        "single_ws_notes": "Calculated KPI metric formula definitions.",
        "arch_notes": "Child of Configuration & Governance."
    },
    {
        "id": "menu-psv",
        "action": "CHILD",
        "proposed_title": "Channel Visibility Matrix (PSV)",
        "proposed_route": "/psv",
        "proposed_group": "Configuration & Governance",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-masters",
        "visibility_model": "ROLE_BASED",
        "req_capability": "psv.visibility.manage",
        "access_perm": "PSV.MANAGE",
        "single_ws_notes": "Omnichannel price & visibility rules.",
        "arch_notes": "Child of Configuration & Governance."
    },
    {
        "id": "menu-document-series",
        "action": "CHILD",
        "proposed_title": "Numbering Engine & Series",
        "proposed_route": "/document-series",
        "proposed_group": "Configuration & Governance",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-masters",
        "visibility_model": "ROLE_BASED",
        "req_capability": "numbering.series.manage",
        "access_perm": "NUMBERING.MANAGE",
        "single_ws_notes": "Invoice/GRN document sequence counter engine.",
        "arch_notes": "Child of Configuration & Governance."
    },
    {
        "id": "menu-print-studio",
        "action": "CHILD",
        "proposed_title": "Print Studio & Template Designer",
        "proposed_route": "/print-studio",
        "proposed_group": "Configuration & Governance",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-masters",
        "visibility_model": "ROLE_BASED",
        "req_capability": "print.templates.manage",
        "access_perm": "PRINT.MANAGE",
        "single_ws_notes": "Thermal bill & invoice print format designer.",
        "arch_notes": "Child of Configuration & Governance."
    },
    {
        "id": "menu-print-history",
        "action": "CHILD",
        "proposed_title": "Print Audit & History Logs",
        "proposed_route": "/print-history",
        "proposed_group": "Configuration & Governance",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-masters",
        "visibility_model": "ROLE_BASED",
        "req_capability": "print.history.view",
        "access_perm": "PRINT.LOG.ACCESS",
        "single_ws_notes": "Re-print tracking & audit trail.",
        "arch_notes": "Child of Configuration & Governance."
    },
    {
        "id": "menu-terms-engine",
        "action": "CHILD",
        "proposed_title": "Terms & Conditions Engine",
        "proposed_route": "/terms-engine",
        "proposed_group": "Configuration & Governance",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-masters",
        "visibility_model": "ROLE_BASED",
        "req_capability": "terms.engine.manage",
        "access_perm": "TERMS.MANAGE",
        "single_ws_notes": "Legal boilerplate & invoice terms manager.",
        "arch_notes": "Child of Configuration & Governance."
    },
    {
        "id": "menu-data-exchange",
        "action": "CHILD",
        "proposed_title": "Data Exchange & Migration Hub",
        "proposed_route": "/data-exchange",
        "proposed_group": "Configuration & Governance",
        "proposed_scope": "COMPANY",
        "proposed_parent": "menu-masters",
        "visibility_model": "ROLE_BASED",
        "req_capability": "data.exchange.import",
        "access_perm": "DATA.IMPORT.ACCESS",
        "single_ws_notes": "CSV/Excel bulk data import & export hub.",
        "arch_notes": "Child of Configuration & Governance."
    },

    # 8. Administration Workspace Group
    {
        "id": "menu-staff-management",
        "action": "RESTRICT",
        "proposed_title": "Staff Management & Payroll",
        "proposed_route": "/staff-management",
        "proposed_group": "Administration",
        "proposed_scope": "COMPANY",
        "proposed_parent": "ROOT",
        "visibility_model": "CAPABILITY_BASED",
        "req_capability": "staff.management.access",
        "access_perm": "STAFF.WORKSPACE.ACCESS",
        "single_ws_notes": "Employee records & cashier assignments. Restricted to HR / Store Manager.",
        "arch_notes": "Restricted administrative module."
    },
    {
        "id": "menu-approval-matrix",
        "action": "RESTRICT",
        "proposed_title": "Approval Matrix Governance",
        "proposed_route": "/approval-matrix",
        "proposed_group": "Administration",
        "proposed_scope": "COMPANY",
        "proposed_parent": "ROOT",
        "visibility_model": "CAPABILITY_BASED",
        "req_capability": "approval.matrix.manage",
        "access_perm": "APPROVAL.MANAGE",
        "single_ws_notes": "Multi-tier discount & credit limit approval rules.",
        "arch_notes": "Restricted to General Manager / SYSADMIN."
    },
    {
        "id": "menu-company-setup",
        "action": "RESTRICT",
        "proposed_title": "Company Setup & Branch Config",
        "proposed_route": "/company-setup",
        "proposed_group": "Administration",
        "proposed_scope": "TENANT",
        "proposed_parent": "ROOT",
        "visibility_model": "CAPABILITY_BASED",
        "req_capability": "system.company.setup",
        "access_perm": "COMPANY.SETUP.ACCESS",
        "single_ws_notes": "Onboarding & enterprise legal entity configuration.",
        "arch_notes": "Restricted to Tenant Administrator."
    },
    {
        "id": "menu-audit-logs",
        "action": "RESTRICT",
        "proposed_title": "System Audit Trail & Security Logs",
        "proposed_route": "/audit-logs",
        "proposed_group": "Administration",
        "proposed_scope": "COMPANY",
        "proposed_parent": "ROOT",
        "visibility_model": "CAPABILITY_BASED",
        "req_capability": "system.audit.view",
        "access_perm": "AUDIT.WORKSPACE.ACCESS",
        "single_ws_notes": "Control Plane audit log inspection hub.",
        "arch_notes": "Restricted to Compliance Auditor & SYSADMIN."
    },
]

def populate_decision_matrix():
    assert os.path.exists(FILE_PATH), f"Excel file missing at {FILE_PATH}"

    wb = openpyxl.load_workbook(FILE_PATH)
    
    sheet_name = "TARGET_MODEL_V1"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    headers = [
        "Menu ID", "Current Title", "Current Route", "Target Navigation Group (8 Groups)",
        "Recommended Action", "Proposed Title", "Proposed Route",
        "Hierarchy (Parent ID)", "Scope Applicability", "Visibility Model",
        "Required Capability", "Workspace Access Permission",
        "Single Workspace Alignment Notes", "Architectural Governance Notes"
    ]

    ws.append(headers)

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Slate
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rec_map = {r["id"]: r for r in TARGET_MODEL}

    ws_data = wb["MENU_DATA"]
    for row in ws_data.iter_rows(min_row=2, values_only=True):
        menu_id = row[0]
        title = row[1]
        route = row[2]

        rec = rec_map.get(menu_id, {
            "action": "KEEP", "proposed_title": title, "proposed_route": route,
            "proposed_group": "Dashboard & Operations", "proposed_scope": "GLOBAL",
            "proposed_parent": "ROOT", "visibility_model": "GLOBAL",
            "req_capability": "GLOBAL_AUTHENTICATED", "access_perm": "GENERAL.ACCESS",
            "single_ws_notes": "Default retention", "arch_notes": "Standard menu"
        })

        ws.append([
            menu_id,
            title,
            route,
            rec["proposed_group"],
            rec["action"],
            rec["proposed_title"],
            rec["proposed_route"],
            rec["proposed_parent"],
            rec["proposed_scope"],
            rec["visibility_model"],
            rec["req_capability"],
            rec["access_perm"],
            rec["single_ws_notes"],
            rec["arch_notes"]
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 48)

    wb.save(FILE_PATH)
    wb.close()

    print("============================================================")
    print("SMRITI MENU GOVERNANCE TARGET MODEL v1.0 GENERATED")
    print("============================================================")
    print(f"Excel Workbook Updated : {FILE_PATH}")
    print(f"Worksheet Created      : {sheet_name} (34 Target Model Specifications)")
    print(f"Database Status        : ZERO MUTATIONS (100% Read-Only Baseline Maintained)")

if __name__ == "__main__":
    populate_decision_matrix()
