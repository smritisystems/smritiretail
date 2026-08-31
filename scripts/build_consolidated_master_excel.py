"""
Project      : SMRITI Retail OS
Author       : Jawahar Ramkripal Mallah
Designation  : Chief Systems Architect & Creator
Email        : support@smritibooks.com
Websites     : smritibooks.com | erpnbook.com | aitdl.com
Version      : 3.28.0
Created      : 2026-08-25
Modified     : 2026-08-25
Copyright    : © SMRITIBooks.com. All Rights Reserved.
License      : Proprietary Commercial Software
Classification: Internal
"""

import os
import re
from decimal import Decimal
import pypdf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

FOLDER_PATH = r"F:\Smriti-Clients Data\Tattly Threads\Invoice\Tattly Threads"

OUTPUT_WORKBOOK_PATH = r"F:\Smriti-Clients Data\Tattly Threads\Invoice\Tattly Threads\MASTER_CONSOLIDATED_ALL_60_INVOICES_TATTLY_THREADS.xlsx"
BACKUP_WORKBOOK_PATH = r"F:\Smriti-Clients Data\Tattly Threads\MASTER_CONSOLIDATED_ALL_60_INVOICES_TATTLY_THREADS.xlsx"
EXPORTS_WORKBOOK_PATH = r"F:\SMRITRretailNX\exports\MASTER_CONSOLIDATED_ALL_60_INVOICES_TATTLY_THREADS.xlsx"
MASTER_CSV_PATH = r"F:\Smriti-Clients Data\Tattly Threads\Invoice\Tattly Threads\MASTER_ALL_18036_ITEMS_REGISTER.csv"

def parse_all_pos(folder_path):
    pdf_files = sorted([f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")])
    print(f"Starting extraction for {len(pdf_files)} PDF Purchase Orders...")
    
    po_summaries = []
    all_items = []
    
    for idx, pdf_name in enumerate(pdf_files, start=1):
        pdf_path = os.path.join(folder_path, pdf_name)
        reader = pypdf.PdfReader(pdf_path)
        num_pages = len(reader.pages)
        
        p1_text = reader.pages[0].extract_text() or ""
        
        m_po = re.search(r'PO\s*NO\.?\s*:\s*(\d+)', p1_text, re.IGNORECASE)
        po_num = m_po.group(1) if m_po else pdf_name.replace(".pdf", "")
        
        m_podate = re.search(r'PO\s*Date\s*:\s*(\d{2}\.\d{2}\.\d{4})', p1_text, re.IGNORECASE)
        po_date = m_podate.group(1) if m_podate else "31.07.2026"
        
        m_site = re.search(r'Site\s*:\s*([A-Z0-9]{4})', p1_text)
        site_code = m_site.group(1) if m_site else "S4NN"
        
        m_deldate = re.search(r'DELIVERY\s*DATE\s*:\s*(\d{2}\.\d{2}\.\d{4})', p1_text, re.IGNORECASE)
        del_date = m_deldate.group(1) if m_deldate else "15.09.2026"
        
        m_basic = re.search(r'TOTAL\s+BASIC\s+VALUE\s+INR\s+([\d,]+\.?\d*)', p1_text, re.IGNORECASE)
        hdr_basic = Decimal(m_basic.group(1).replace(",", "")) if m_basic else Decimal("0")
        
        m_igst = re.search(r'TOTAL\s+IGST\s+INR\s+([\d,]+\.?\d*)', p1_text, re.IGNORECASE)
        hdr_igst = Decimal(m_igst.group(1).replace(",", "")) if m_igst else Decimal("0")
        
        m_tot = re.search(r'Total\s+Order\s+Value\s*:\s*INR\s+([\d,]+\.?\d*)', p1_text, re.IGNORECASE)
        hdr_tot = Decimal(m_tot.group(1).replace(",", "")) if m_tot else Decimal("0")
        
        po_item_count = 0
        po_qty = Decimal("0")
        po_calc_base = Decimal("0")
        po_calc_igst = Decimal("0")
        
        for p_idx in range(1, num_pages):
            p_text = reader.pages[p_idx].extract_text()
            if not p_text:
                continue
            if "General Conditions of Purchase" in p_text or "Anti-Bribery and Corruption" in p_text or "Site Site Name Address Format" in p_text:
                continue
            
            lines = [l.strip() for l in p_text.split("\n") if l.strip()]
            i = 0
            while i < len(lines):
                line = lines[i]
                m_start = re.match(r'^(\d+)\s+(\d{10,14})$', line)
                if m_start:
                    sr_no = int(m_start.group(1))
                    article_no = m_start.group(2)
                    
                    hsn = lines[i+1] if i+1 < len(lines) else ""
                    ean = lines[i+2] if i+2 < len(lines) else ""
                    vendor_art = lines[i+3] if i+3 < len(lines) else ""
                    desc = lines[i+4] if i+4 < len(lines) else ""
                    
                    item_del_date = del_date
                    item_site = site_code
                    qty = Decimal("0")
                    uom = "EA"
                    mrp = Decimal("0")
                    base_cost = Decimal("0")
                    igst_pct = Decimal("5.00")
                    igst_amt = Decimal("0")
                    total_base_val = Decimal("0")
                    
                    for j in range(i+4, min(i+18, len(lines))):
                        if re.match(r'^\d{2}\.\d{2}\.\d{4}$', lines[j]):
                            item_del_date = lines[j]
                            if j+1 < len(lines) and re.match(r'^[A-Z0-9]{4}$', lines[j+1]):
                                item_site = lines[j+1]
                        
                        m_qty = re.search(r'([\d,]+\.?\d*)\s+(EA|PRS|SET|PCS|PAIR)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)', lines[j])
                        if m_qty:
                            qty = Decimal(m_qty.group(1).replace(",", ""))
                            uom = m_qty.group(2)
                            mrp = Decimal(m_qty.group(3).replace(",", ""))
                            base_cost = Decimal(m_qty.group(4).replace(",", ""))
                            igst_pct = Decimal(m_qty.group(5).replace(",", ""))
                        
                        if "Sub Total of Qty" in lines[j]:
                            val_lines = [lines[k] for k in range(max(i, j-6), j) if re.match(r'^[\d,]+\.\d{2}$', lines[k])]
                            if len(val_lines) >= 2:
                                try:
                                    igst_amt = Decimal(val_lines[-2].replace(",", ""))
                                    total_base_val = Decimal(val_lines[-1].replace(",", ""))
                                except:
                                    pass
                            elif len(val_lines) == 1:
                                try:
                                    total_base_val = Decimal(val_lines[0].replace(",", ""))
                                except:
                                    pass
                            break
                    
                    if total_base_val == Decimal("0") and qty > 0 and base_cost > 0:
                        total_base_val = (qty * base_cost).quantize(Decimal("0.01"))
                    if igst_amt == Decimal("0") and total_base_val > 0 and igst_pct > 0:
                        igst_amt = (total_base_val * igst_pct / Decimal("100")).quantize(Decimal("0.01"))
                    
                    art_name = desc
                    color = ""
                    size = ""
                    parts = [p.strip() for p in desc.split(",")]
                    if len(parts) >= 3:
                        art_name = parts[0]
                        color = parts[1]
                        size = parts[2]
                    elif len(parts) == 2:
                        art_name = parts[0]
                        color = parts[1]
                    
                    item_dict = {
                        "po_number": po_num,
                        "po_date": po_date,
                        "delivery_date": item_del_date,
                        "site": item_site,
                        "sr_no": sr_no,
                        "article_no": article_no,
                        "hsn": hsn,
                        "ean": ean,
                        "vendor_art": vendor_art,
                        "description": desc,
                        "art_name": art_name,
                        "color": color,
                        "size": size,
                        "qty": float(qty),
                        "uom": uom,
                        "mrp": float(mrp),
                        "base_cost": float(base_cost),
                        "igst_pct": float(igst_pct),
                        "igst_amt": float(igst_amt),
                        "total_base_val": float(total_base_val),
                        "total_gross_val": float(total_base_val + igst_amt)
                    }
                    all_items.append(item_dict)
                    po_item_count += 1
                    po_qty += qty
                    po_calc_base += total_base_val
                    po_calc_igst += igst_amt
                i += 1
        
        po_summaries.append({
            "po_number": po_num,
            "po_date": po_date,
            "delivery_date": del_date,
            "site": site_code,
            "pages": num_pages,
            "item_count": po_item_count,
            "total_qty": float(po_qty),
            "hdr_basic": float(hdr_basic if hdr_basic > 0 else po_calc_base),
            "calc_basic": float(po_calc_base),
            "hdr_igst": float(hdr_igst if hdr_igst > 0 else po_calc_igst),
            "calc_igst": float(po_calc_igst),
            "hdr_total": float(hdr_tot if hdr_tot > 0 else (po_calc_base + po_calc_igst)),
            "calc_total": float(po_calc_base + po_calc_igst)
        })
        
        if idx % 15 == 0 or idx == len(pdf_files):
            print(f"  Processed {idx}/{len(pdf_files)}: PO {po_num} | {po_item_count} items | Qty={po_qty:,.0f} | Total=INR {po_calc_base+po_calc_igst:,.2f}")
            
    return po_summaries, all_items

def generate_master_workbook(po_summaries, all_items):
    print("Building master consolidated openpyxl workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Theme Fonts & Fills
    f_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    f_sub = Font(name="Calibri", size=10, italic=True, color="E2E8F0")
    f_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    f_data = Font(name="Calibri", size=10)
    f_mono = Font(name="Consolas", size=9)
    f_total = Font(name="Calibri", size=10, bold=True, color="000000")
    
    fill_navy = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_blue = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    fill_teal = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    fill_indigo = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    fill_purple = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    fill_slate = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_total = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    total_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    # ─────────────────────────────────────────────────────────────
    # SHEET 1: Master PO Summary (All 60 Purchase Orders)
    # ─────────────────────────────────────────────────────────────
    ws_summary = wb.create_sheet(title="Master 60 POs Summary")
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Banner
    ws_summary.merge_cells("A1:K2")
    ws_summary["A1"] = "RELIANCE RETAIL LIMITED — MASTER 60 PURCHASE ORDERS REGISTER"
    ws_summary["A1"].font = f_title
    ws_summary["A1"].fill = fill_navy
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws_summary.merge_cells("A3:K3")
    ws_summary["A3"] = f"Vendor: TATTLY THREADS (32071140) | Total POs: {len(po_summaries)} | Destination DC: S4NN Tumkur NDC | State: Karnataka (29)"
    ws_summary["A3"].font = f_sub
    ws_summary["A3"].fill = fill_blue
    ws_summary["A3"].alignment = Alignment(horizontal="center", vertical="center")
    
    summary_headers = [
        "S.No", "PO Number", "PO Date", "Delivery Date", "NDC Site",
        "PDF Pages", "Total Items", "Total Qty (Pairs)", "Basic Value (₹)",
        "IGST 5% (₹)", "Total PO Value (₹)"
    ]
    
    ws_summary.row_dimensions[5].height = 28
    for c_idx, h in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(5, c_idx, h)
        cell.font = f_header
        cell.fill = fill_navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    s_row = 6
    for idx, p in enumerate(po_summaries, start=1):
        ws_summary.row_dimensions[s_row].height = 18
        row_fill = fill_zebra if s_row % 2 == 0 else PatternFill(fill_type=None)
        
        ws_summary.cell(s_row, 1, idx).alignment = Alignment(horizontal="center")
        ws_summary.cell(s_row, 2, p["po_number"]).font = Font(name="Calibri", size=10, bold=True)
        ws_summary.cell(s_row, 2).alignment = Alignment(horizontal="center")
        ws_summary.cell(s_row, 3, p["po_date"]).alignment = Alignment(horizontal="center")
        ws_summary.cell(s_row, 4, p["delivery_date"]).alignment = Alignment(horizontal="center")
        ws_summary.cell(s_row, 5, p["site"]).alignment = Alignment(horizontal="center")
        ws_summary.cell(s_row, 6, p["pages"]).alignment = Alignment(horizontal="center")
        
        c_items = ws_summary.cell(s_row, 7, p["item_count"])
        c_items.number_format = '#,##0'
        c_items.alignment = Alignment(horizontal="right")
        
        c_qty = ws_summary.cell(s_row, 8, p["total_qty"])
        c_qty.number_format = '#,##0'
        c_qty.font = Font(name="Calibri", size=10, bold=True)
        c_qty.alignment = Alignment(horizontal="right")
        
        c_basic = ws_summary.cell(s_row, 9, p["calc_basic"])
        c_basic.number_format = '₹ #,##0.00'
        c_basic.alignment = Alignment(horizontal="right")
        
        c_igst = ws_summary.cell(s_row, 10, p["calc_igst"])
        c_igst.number_format = '₹ #,##0.00'
        c_igst.alignment = Alignment(horizontal="right")
        
        c_tot = ws_summary.cell(s_row, 11, p["calc_total"])
        c_tot.number_format = '₹ #,##0.00'
        c_tot.font = Font(name="Calibri", size=10, bold=True)
        c_tot.alignment = Alignment(horizontal="right")
        
        for c in range(1, 12):
            c_cell = ws_summary.cell(s_row, c)
            if row_fill.fill_type:
                c_cell.fill = row_fill
            c_cell.border = thin_border
            if not c_cell.font.name:
                c_cell.font = f_data
        
        s_row += 1
        
    # Summary Totals Row
    ws_summary.row_dimensions[s_row].height = 22
    ws_summary.cell(s_row, 1, "TOTAL").font = f_total
    ws_summary.cell(s_row, 1).alignment = Alignment(horizontal="center")
    ws_summary.cell(s_row, 2, f"{len(po_summaries)} POs").font = f_total
    
    c_tot_it = ws_summary.cell(s_row, 7, f"=SUM(G6:G{s_row-1})")
    c_tot_it.font = f_total
    c_tot_it.number_format = '#,##0'
    c_tot_it.alignment = Alignment(horizontal="right")
    
    c_tot_qt = ws_summary.cell(s_row, 8, f"=SUM(H6:H{s_row-1})")
    c_tot_qt.font = f_total
    c_tot_qt.number_format = '#,##0'
    c_tot_qt.alignment = Alignment(horizontal="right")
    
    c_tot_b = ws_summary.cell(s_row, 9, f"=SUM(I6:I{s_row-1})")
    c_tot_b.font = f_total
    c_tot_b.number_format = '₹ #,##0.00'
    c_tot_b.alignment = Alignment(horizontal="right")
    
    c_tot_ig = ws_summary.cell(s_row, 10, f"=SUM(J6:J{s_row-1})")
    c_tot_ig.font = f_total
    c_tot_ig.number_format = '₹ #,##0.00'
    c_tot_ig.alignment = Alignment(horizontal="right")
    
    c_tot_all = ws_summary.cell(s_row, 11, f"=SUM(K6:K{s_row-1})")
    c_tot_all.font = f_total
    c_tot_all.number_format = '₹ #,##0.00'
    c_tot_all.alignment = Alignment(horizontal="right")
    
    for c in range(1, 12):
        c_cell = ws_summary.cell(s_row, c)
        c_cell.fill = fill_total
        c_cell.border = total_border
        
    ws_summary.freeze_panes = "A6"
    ws_summary.auto_filter.ref = f"A5:K{s_row-1}"
    
    # ─────────────────────────────────────────────────────────────
    # SHEET 2: All Line Items Register (18,036 Items)
    # ─────────────────────────────────────────────────────────────
    ws_all = wb.create_sheet(title="All POs Items Master")
    ws_all.views.sheetView[0].showGridLines = True
    
    all_headers = [
        "PO Number", "PO Date", "Delivery Date", "NDC Site", "Line No",
        "Article No (SAP)", "HSN Code", "EAN / Barcode", "Vendor Style / Art",
        "Material Description", "Article Name", "Color", "Size",
        "Quantity (Pairs)", "UOM", "MRP (₹)", "Base Rate (₹)",
        "IGST (%)", "IGST Amt (₹)", "Total Base Value (₹)", "Total Value (₹)"
    ]
    
    ws_all.row_dimensions[1].height = 28
    for c_idx, h in enumerate(all_headers, start=1):
        cell = ws_all.cell(1, c_idx, h)
        cell.font = f_header
        cell.fill = fill_navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    print(f"Writing {len(all_items)} lines to Sheet 2...")
    for r_idx, it in enumerate(all_items, start=2):
        ws_all.cell(r_idx, 1, it["po_number"]).alignment = Alignment(horizontal="center")
        ws_all.cell(r_idx, 2, it["po_date"]).alignment = Alignment(horizontal="center")
        ws_all.cell(r_idx, 3, it["delivery_date"]).alignment = Alignment(horizontal="center")
        ws_all.cell(r_idx, 4, it["site"]).alignment = Alignment(horizontal="center")
        ws_all.cell(r_idx, 5, it["sr_no"]).alignment = Alignment(horizontal="center")
        ws_all.cell(r_idx, 6, it["article_no"]).font = f_mono
        ws_all.cell(r_idx, 7, it["hsn"]).alignment = Alignment(horizontal="center")
        ws_all.cell(r_idx, 8, it["ean"]).font = f_mono
        ws_all.cell(r_idx, 9, it["vendor_art"]).font = Font(name="Calibri", size=10, bold=True)
        ws_all.cell(r_idx, 10, it["description"])
        ws_all.cell(r_idx, 11, it["art_name"])
        ws_all.cell(r_idx, 12, it["color"]).alignment = Alignment(horizontal="center")
        ws_all.cell(r_idx, 13, it["size"]).alignment = Alignment(horizontal="center")
        
        c_qty = ws_all.cell(r_idx, 14, it["qty"])
        c_qty.number_format = '#,##0'
        c_qty.alignment = Alignment(horizontal="right")
        
        ws_all.cell(r_idx, 15, it["uom"]).alignment = Alignment(horizontal="center")
        
        c_mrp = ws_all.cell(r_idx, 16, it["mrp"])
        c_mrp.number_format = '₹ #,##0.00'
        c_mrp.alignment = Alignment(horizontal="right")
        
        c_cost = ws_all.cell(r_idx, 17, it["base_cost"])
        c_cost.number_format = '₹ #,##0.00'
        c_cost.alignment = Alignment(horizontal="right")
        
        c_igstp = ws_all.cell(r_idx, 18, it["igst_pct"] / 100.0)
        c_igstp.number_format = '0.00%'
        c_igstp.alignment = Alignment(horizontal="right")
        
        c_igst = ws_all.cell(r_idx, 19, it["igst_amt"])
        c_igst.number_format = '₹ #,##0.00'
        c_igst.alignment = Alignment(horizontal="right")
        
        c_base = ws_all.cell(r_idx, 20, it["total_base_val"])
        c_base.number_format = '₹ #,##0.00'
        c_base.alignment = Alignment(horizontal="right")
        
        c_gross = ws_all.cell(r_idx, 21, it["total_gross_val"])
        c_gross.number_format = '₹ #,##0.00'
        c_gross.alignment = Alignment(horizontal="right")
        
        for c in range(1, 22):
            c_cell = ws_all.cell(r_idx, c)
            c_cell.border = thin_border
            if not c_cell.font.name:
                c_cell.font = f_data
                
    # Items Totals Row
    all_tot_row = len(all_items) + 2
    ws_all.row_dimensions[all_tot_row].height = 22
    ws_all.cell(all_tot_row, 1, "TOTAL").font = f_total
    ws_all.cell(all_tot_row, 1).alignment = Alignment(horizontal="center")
    ws_all.cell(all_tot_row, 2, f"{len(all_items)} Items").font = f_total
    
    c_all_qty = ws_all.cell(all_tot_row, 14, f"=SUM(N2:N{all_tot_row-1})")
    c_all_qty.font = f_total
    c_all_qty.number_format = '#,##0'
    c_all_qty.alignment = Alignment(horizontal="right")
    
    c_all_igst = ws_all.cell(all_tot_row, 19, f"=SUM(S2:S{all_tot_row-1})")
    c_all_igst.font = f_total
    c_all_igst.number_format = '₹ #,##0.00'
    c_all_igst.alignment = Alignment(horizontal="right")
    
    c_all_base = ws_all.cell(all_tot_row, 20, f"=SUM(T2:T{all_tot_row-1})")
    c_all_base.font = f_total
    c_all_base.number_format = '₹ #,##0.00'
    c_all_base.alignment = Alignment(horizontal="right")
    
    c_all_gross = ws_all.cell(all_tot_row, 21, f"=SUM(U2:U{all_tot_row-1})")
    c_all_gross.font = f_total
    c_all_gross.number_format = '₹ #,##0.00'
    c_all_gross.alignment = Alignment(horizontal="right")
    
    for c in range(1, 22):
        c_cell = ws_all.cell(all_tot_row, c)
        c_cell.fill = fill_total
        c_cell.border = total_border
        
    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = f"A1:U{all_tot_row-1}"
    
    # ─────────────────────────────────────────────────────────────
    # SHEET 3: Consolidated Size Matrix (All 60 POs Combined)
    # ─────────────────────────────────────────────────────────────
    ws_cmat = wb.create_sheet(title="Consolidated Size Matrix")
    ws_cmat.views.sheetView[0].showGridLines = True
    
    all_sizes = ["35", "36", "37", "38", "39", "40", "41", "42"]
    cons_matrix = {}
    
    for it in all_items:
        key = (it["vendor_art"], it["art_name"], it["color"], it["mrp"], it["base_cost"])
        if key not in cons_matrix:
            cons_matrix[key] = {s: 0 for s in all_sizes}
        sz = str(it["size"]).strip()
        if sz in cons_matrix[key]:
            cons_matrix[key][sz] += it["qty"]
        else:
            cons_matrix[key][sz] = it["qty"]
            
    cmat_headers = [
        "Style / Vendor Art", "Article Description", "Color", "Base Rate (₹)", "MRP (₹)",
        "Sz 35", "Sz 36", "Sz 37", "Sz 38", "Sz 39", "Sz 40", "Sz 41", "Sz 42",
        "Total Pairs", "Base Value (₹)", "IGST 5% (₹)", "Gross Value (₹)"
    ]
    
    ws_cmat.row_dimensions[1].height = 28
    for c_idx, h in enumerate(cmat_headers, start=1):
        cell = ws_cmat.cell(1, c_idx, h)
        cell.font = f_header
        cell.fill = fill_teal
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    cm_row = 2
    for (v_art, a_name, col, mrp_val, base_rate), size_dict in sorted(cons_matrix.items()):
        ws_cmat.row_dimensions[cm_row].height = 18
        row_fill = fill_zebra if cm_row % 2 == 0 else PatternFill(fill_type=None)
        
        ws_cmat.cell(cm_row, 1, v_art).font = Font(name="Calibri", size=10, bold=True)
        ws_cmat.cell(cm_row, 2, a_name)
        ws_cmat.cell(cm_row, 3, col).alignment = Alignment(horizontal="center")
        
        c_rate = ws_cmat.cell(cm_row, 4, base_rate)
        c_rate.number_format = '₹ #,##0.00'
        c_rate.alignment = Alignment(horizontal="right")
        
        c_mrp = ws_cmat.cell(cm_row, 5, mrp_val)
        c_mrp.number_format = '₹ #,##0.00'
        c_mrp.alignment = Alignment(horizontal="right")
        
        for s_idx, sz_code in enumerate(all_sizes, start=6):
            s_qty = size_dict.get(sz_code, 0)
            c_s = ws_cmat.cell(cm_row, s_idx, s_qty if s_qty > 0 else "-")
            c_s.alignment = Alignment(horizontal="center")
            if s_qty > 0:
                c_s.font = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
                
        c_tot_p = ws_cmat.cell(cm_row, 14, f"=SUM(F{cm_row}:M{cm_row})")
        c_tot_p.number_format = '#,##0'
        c_tot_p.font = Font(name="Calibri", size=10, bold=True)
        c_tot_p.alignment = Alignment(horizontal="right")
        
        c_bval = ws_cmat.cell(cm_row, 15, f"=N{cm_row}*D{cm_row}")
        c_bval.number_format = '₹ #,##0.00'
        c_bval.alignment = Alignment(horizontal="right")
        
        c_igst = ws_cmat.cell(cm_row, 16, f"=O{cm_row}*0.05")
        c_igst.number_format = '₹ #,##0.00'
        c_igst.alignment = Alignment(horizontal="right")
        
        c_gval = ws_cmat.cell(cm_row, 17, f"=O{cm_row}+P{cm_row}")
        c_gval.number_format = '₹ #,##0.00'
        c_gval.alignment = Alignment(horizontal="right")
        
        for c in range(1, 18):
            c_cell = ws_cmat.cell(cm_row, c)
            if row_fill.fill_type:
                c_cell.fill = row_fill
            c_cell.border = thin_border
            if not c_cell.font.name:
                c_cell.font = f_data
        cm_row += 1
        
    # Matrix Total Row
    ws_cmat.row_dimensions[cm_row].height = 22
    ws_cmat.cell(cm_row, 1, "TOTAL").font = f_total
    ws_cmat.cell(cm_row, 1).alignment = Alignment(horizontal="center")
    ws_cmat.cell(cm_row, 2, f"{len(cons_matrix)} SKU Variations").font = f_total
    
    for s_idx, col_let in enumerate(["F", "G", "H", "I", "J", "K", "L", "M"], start=6):
        c_sz_tot = ws_cmat.cell(cm_row, s_idx, f"=SUM({col_let}2:{col_let}{cm_row-1})")
        c_sz_tot.font = f_total
        c_sz_tot.number_format = '#,##0'
        c_sz_tot.alignment = Alignment(horizontal="center")
        
    c_mqty = ws_cmat.cell(cm_row, 14, f"=SUM(N2:N{cm_row-1})")
    c_mqty.font = f_total
    c_mqty.number_format = '#,##0'
    c_mqty.alignment = Alignment(horizontal="right")
    
    c_mbval = ws_cmat.cell(cm_row, 15, f"=SUM(O2:O{cm_row-1})")
    c_mbval.font = f_total
    c_mbval.number_format = '₹ #,##0.00'
    c_mbval.alignment = Alignment(horizontal="right")
    
    c_migst = ws_cmat.cell(cm_row, 16, f"=SUM(P2:P{cm_row-1})")
    c_migst.font = f_total
    c_migst.number_format = '₹ #,##0.00'
    c_migst.alignment = Alignment(horizontal="right")
    
    c_mgval = ws_cmat.cell(cm_row, 17, f"=SUM(Q2:Q{cm_row-1})")
    c_mgval.font = f_total
    c_mgval.number_format = '₹ #,##0.00'
    c_mgval.alignment = Alignment(horizontal="right")
    
    for c in range(1, 18):
        c_cell = ws_cmat.cell(cm_row, c)
        c_cell.fill = fill_total
        c_cell.border = total_border
        
    ws_cmat.freeze_panes = "A2"
    ws_cmat.auto_filter.ref = f"A1:Q{cm_row-1}"
    
    # ─────────────────────────────────────────────────────────────
    # SHEET 4: Style-Wise Summary (All 60 POs Combined)
    # ─────────────────────────────────────────────────────────────
    ws_cstyle = wb.create_sheet(title="Style-Wise Master Summary")
    ws_cstyle.views.sheetView[0].showGridLines = True
    
    style_headers = [
        "Vendor Style / Article", "Article Description", "Colors Available", "SKU Lines",
        "Total Pairs", "Base Rate (₹)", "MRP (₹)", "Total Base Value (₹)",
        "IGST 5% (₹)", "Total Gross Value (₹)", "% of Master POs"
    ]
    
    ws_cstyle.row_dimensions[1].height = 28
    for col_num, h_name in enumerate(style_headers, start=1):
        cell = ws_cstyle.cell(1, col_num, h_name)
        cell.font = f_header
        cell.fill = fill_indigo
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    style_agg = {}
    for it in all_items:
        st = it["vendor_art"]
        if st not in style_agg:
            style_agg[st] = {
                "desc": it["art_name"],
                "colors": set(),
                "skus": 0,
                "qty": 0,
                "base_cost": it["base_cost"],
                "mrp": it["mrp"],
                "base_val": 0,
                "igst": 0,
                "gross": 0
            }
        style_agg[st]["colors"].add(it["color"])
        style_agg[st]["skus"] += 1
        style_agg[st]["qty"] += it["qty"]
        style_agg[st]["base_val"] += it["total_base_val"]
        style_agg[st]["igst"] += it["igst_amt"]
        style_agg[st]["gross"] += it["total_gross_val"]
        
    tot_all_gross = sum(s["gross"] for s in style_agg.values()) or 1
    
    st_row = 2
    for st_code, s_info in sorted(style_agg.items()):
        ws_cstyle.row_dimensions[st_row].height = 18
        row_fill = fill_zebra if st_row % 2 == 0 else PatternFill(fill_type=None)
        
        ws_cstyle.cell(st_row, 1, st_code).font = Font(name="Calibri", size=10, bold=True)
        ws_cstyle.cell(st_row, 2, s_info["desc"])
        ws_cstyle.cell(st_row, 3, ", ".join(sorted(s_info["colors"])))
        ws_cstyle.cell(st_row, 4, s_info["skus"]).alignment = Alignment(horizontal="center")
        
        c_sqty = ws_cstyle.cell(st_row, 5, s_info["qty"])
        c_sqty.number_format = '#,##0'
        c_sqty.font = Font(name="Calibri", size=10, bold=True)
        c_sqty.alignment = Alignment(horizontal="right")
        
        c_srate = ws_cstyle.cell(st_row, 6, s_info["base_cost"])
        c_srate.number_format = '₹ #,##0.00'
        c_srate.alignment = Alignment(horizontal="right")
        
        c_smrp = ws_cstyle.cell(st_row, 7, s_info["mrp"])
        c_smrp.number_format = '₹ #,##0.00'
        c_smrp.alignment = Alignment(horizontal="right")
        
        c_sbval = ws_cstyle.cell(st_row, 8, s_info["base_val"])
        c_sbval.number_format = '₹ #,##0.00'
        c_sbval.alignment = Alignment(horizontal="right")
        
        c_sigst = ws_cstyle.cell(st_row, 9, s_info["igst"])
        c_sigst.number_format = '₹ #,##0.00'
        c_sigst.alignment = Alignment(horizontal="right")
        
        c_sgross = ws_cstyle.cell(st_row, 10, s_info["gross"])
        c_sgross.number_format = '₹ #,##0.00'
        c_sgross.alignment = Alignment(horizontal="right")
        
        c_spct = ws_cstyle.cell(st_row, 11, s_info["gross"] / tot_all_gross)
        c_spct.number_format = '0.00%'
        c_spct.alignment = Alignment(horizontal="right")
        
        for c in range(1, 12):
            c_cell = ws_cstyle.cell(st_row, c)
            if row_fill.fill_type:
                c_cell.fill = row_fill
            c_cell.border = thin_border
            if not c_cell.font.name:
                c_cell.font = f_data
        st_row += 1
        
    # Style Total Row
    ws_cstyle.row_dimensions[st_row].height = 22
    ws_cstyle.cell(st_row, 1, "TOTAL").font = f_total
    ws_cstyle.cell(st_row, 1).alignment = Alignment(horizontal="center")
    ws_cstyle.cell(st_row, 2, f"{len(style_agg)} Styles").font = f_total
    ws_cstyle.cell(st_row, 4, f"=SUM(D2:D{st_row-1})").font = f_total
    ws_cstyle.cell(st_row, 4).alignment = Alignment(horizontal="center")
    
    c_st_tot_qty = ws_cstyle.cell(st_row, 5, f"=SUM(E2:E{st_row-1})")
    c_st_tot_qty.font = f_total
    c_st_tot_qty.number_format = '#,##0'
    c_st_tot_qty.alignment = Alignment(horizontal="right")
    
    c_st_tot_bval = ws_cstyle.cell(st_row, 8, f"=SUM(H2:H{st_row-1})")
    c_st_tot_bval.font = f_total
    c_st_tot_bval.number_format = '₹ #,##0.00'
    c_st_tot_bval.alignment = Alignment(horizontal="right")
    
    c_st_tot_igst = ws_cstyle.cell(st_row, 9, f"=SUM(I2:I{st_row-1})")
    c_st_tot_igst.font = f_total
    c_st_tot_igst.number_format = '₹ #,##0.00'
    c_st_tot_igst.alignment = Alignment(horizontal="right")
    
    c_st_tot_gross = ws_cstyle.cell(st_row, 10, f"=SUM(J2:J{st_row-1})")
    c_st_tot_gross.font = f_total
    c_st_tot_gross.number_format = '₹ #,##0.00'
    c_st_tot_gross.alignment = Alignment(horizontal="right")
    
    c_st_tot_pct = ws_cstyle.cell(st_row, 11, f"=SUM(K2:K{st_row-1})")
    c_st_tot_pct.font = f_total
    c_st_tot_pct.number_format = '0.00%'
    c_st_tot_pct.alignment = Alignment(horizontal="right")
    
    for c in range(1, 12):
        c_cell = ws_cstyle.cell(st_row, c)
        c_cell.fill = fill_total
        c_cell.border = total_border
        
    ws_cstyle.freeze_panes = "A2"
    ws_cstyle.auto_filter.ref = f"A1:K{st_row-1}"
    
    # ─────────────────────────────────────────────────────────────
    # SHEET 5: Site & NDC Distribution Breakdown
    # ─────────────────────────────────────────────────────────────
    ws_site = wb.create_sheet(title="NDC Site Distribution")
    ws_site.views.sheetView[0].showGridLines = True
    
    site_headers = [
        "Site Code", "NDC Location / Distribution Center", "PO Count", "Total Line Items",
        "Total Quantity (Pairs)", "Total Basic Value (₹)", "Total IGST (₹)", "Total PO Value (₹)", "% Share"
    ]
    
    ws_site.row_dimensions[1].height = 28
    for col_num, h_name in enumerate(site_headers, start=1):
        cell = ws_site.cell(1, col_num, h_name)
        cell.font = f_header
        cell.fill = fill_purple
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    site_map = {
        "S4NN": "RRL FOOTPRINT Tumkur NDC, Survey No 54 1 Nandihalli Village, 55th KM Stone NH 4 Tumkur Road, TUMKUR KAR-572101"
    }
    
    site_agg = {}
    for p in po_summaries:
        s_code = p["site"]
        if s_code not in site_agg:
            site_agg[s_code] = {
                "desc": site_map.get(s_code, f"Reliance Retail NDC ({s_code})"),
                "pos": 0,
                "items": 0,
                "qty": 0,
                "basic": 0,
                "igst": 0,
                "total": 0
            }
        site_agg[s_code]["pos"] += 1
        site_agg[s_code]["items"] += p["item_count"]
        site_agg[s_code]["qty"] += p["total_qty"]
        site_agg[s_code]["basic"] += p["calc_basic"]
        site_agg[s_code]["igst"] += p["calc_igst"]
        site_agg[s_code]["total"] += p["calc_total"]
        
    site_row = 2
    for sc, s_data in sorted(site_agg.items()):
        ws_site.row_dimensions[site_row].height = 20
        ws_site.cell(site_row, 1, sc).font = Font(name="Calibri", size=10, bold=True)
        ws_site.cell(site_row, 1).alignment = Alignment(horizontal="center")
        ws_site.cell(site_row, 2, s_data["desc"])
        
        c_p = ws_site.cell(site_row, 3, s_data["pos"])
        c_p.number_format = '#,##0'
        c_p.alignment = Alignment(horizontal="center")
        
        c_i = ws_site.cell(site_row, 4, s_data["items"])
        c_i.number_format = '#,##0'
        c_i.alignment = Alignment(horizontal="right")
        
        c_q = ws_site.cell(site_row, 5, s_data["qty"])
        c_q.number_format = '#,##0'
        c_q.font = Font(name="Calibri", size=10, bold=True)
        c_q.alignment = Alignment(horizontal="right")
        
        c_b = ws_site.cell(site_row, 6, s_data["basic"])
        c_b.number_format = '₹ #,##0.00'
        c_b.alignment = Alignment(horizontal="right")
        
        c_ig = ws_site.cell(site_row, 7, s_data["igst"])
        c_ig.number_format = '₹ #,##0.00'
        c_ig.alignment = Alignment(horizontal="right")
        
        c_t = ws_site.cell(site_row, 8, s_data["total"])
        c_t.number_format = '₹ #,##0.00'
        c_t.font = Font(name="Calibri", size=10, bold=True)
        c_t.alignment = Alignment(horizontal="right")
        
        c_pct = ws_site.cell(site_row, 9, s_data["total"] / (tot_all_gross or 1))
        c_pct.number_format = '0.00%'
        c_pct.alignment = Alignment(horizontal="right")
        
        for c in range(1, 10):
            c_cell = ws_site.cell(site_row, c)
            c_cell.border = thin_border
            if not c_cell.font.name:
                c_cell.font = f_data
        site_row += 1
        
    # Auto-fit Column Widths across all sheets (ignoring very large sheets for speed)
    for ws in [ws_summary, ws_cmat, ws_cstyle, ws_site]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if cell.number_format and "₹" in cell.number_format:
                    val_str += "    "
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
    # Sheet 2 column widths manually set for speed
    col_widths = {
        "A": 14, "B": 12, "C": 14, "D": 10, "E": 9,
        "F": 16, "G": 12, "H": 16, "I": 18, "J": 38,
        "K": 26, "L": 14, "M": 8, "N": 16, "O": 8,
        "P": 14, "Q": 14, "R": 10, "S": 14, "T": 18, "U": 18
    }
    for col_l, w in col_widths.items():
        ws_all.column_dimensions[col_l].width = w
        
    return wb

def main():
    start_time = time.time()
    po_summaries, all_items = parse_all_pos(FOLDER_PATH)
    
    print(f"\nSuccessfully extracted all {len(po_summaries)} POs ({len(all_items)} total line items) in {time.time()-start_time:.2f}s.")
    print(f"Total Quantity: {sum(p['total_qty'] for p in po_summaries):,.0f} Pairs")
    print(f"Total Basic: INR {sum(p['calc_basic'] for p in po_summaries):,.2f}")
    print(f"Total IGST: INR {sum(p['calc_igst'] for p in po_summaries):,.2f}")
    print(f"Total Order Value: INR {sum(p['calc_total'] for p in po_summaries):,.2f}")
    
    wb = generate_master_workbook(po_summaries, all_items)
    
    # Save Workbook to target paths
    for target in [OUTPUT_WORKBOOK_PATH, BACKUP_WORKBOOK_PATH, EXPORTS_WORKBOOK_PATH]:
        os.makedirs(os.path.dirname(target), exist_ok=True)
        try:
            wb.save(target)
            print(f"[SUCCESS] Saved Master Excel Workbook: {target} ({os.path.getsize(target):,} bytes)")
        except Exception as e:
            print(f"Error saving to {target}: {e}")
            
    # Save Master CSV for high performance data pipelines
    import csv
    fieldnames = [
        "po_number", "po_date", "delivery_date", "site", "sr_no", "article_no",
        "hsn", "ean", "vendor_art", "description", "art_name", "color", "size",
        "qty", "uom", "mrp", "base_cost", "igst_pct", "igst_amt", "total_base_val", "total_gross_val"
    ]
    with open(MASTER_CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for it in all_items:
            writer.writerow(it)
    print(f"[SUCCESS] Saved Master CSV Register: {MASTER_CSV_PATH} ({os.path.getsize(MASTER_CSV_PATH):,} bytes)")
    print(f"All processing completed in {time.time()-start_time:.2f} seconds.")

if __name__ == "__main__":
    main()
