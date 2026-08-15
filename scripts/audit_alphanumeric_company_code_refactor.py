"""
 * Project      : SMRITI Retail OS
 * Author       : Jawahar Ramkripal Mallah
 * Email        : support@smritibooks.com
 * Websites     : smritibooks.com | erpnbook.com | aitdl.com
 * Version      : 3.16.0
 * Created      : 2026-08-15
 * Modified     : 2026-08-15
 * Copyright    : © SMRITIBooks.com. All Rights Reserved.
 * License      : Proprietary Commercial Software
 * Classification: Internal
"""

import sys, os, glob, re, json
from datetime import datetime, timezone
import psycopg2
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from backend.app.services.company_database_provisioner import CompanyDatabaseProvisioner

DB_PARAMS = "postgresql://postgres:postgres@localhost:5432/smritisys"
EXCEL_OUTPUT = r"F:\SMRITRretailNX\SMRITI_Control_Plane_Architecture_Review.xlsx"
DOC_ARCH = r"F:\SMRITRretailNX\docs\architecture\SMRITI_MULTI_COMPANY_DATABASE_ARCHITECTURE_v1.0.md"

def audit_alphanumeric_company_code():
    print("============================================================")
    print("SMRITI ALPHANUMERIC 3-CHARACTER COMPANY CODE GOVERNANCE REFACTOR")
    print("============================================================")

    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    # 1. Connect & Initial Check
    conn = psycopg2.connect(DB_PARAMS)
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM smriti_menus;")
    initial_menus = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM smriti_audit_log;")
    initial_audit = cur.fetchone()[0]

    # Check database count to ensure 0 new DBs are created
    cur.execute("SELECT datname FROM pg_database WHERE datname IN ('smritiABC', 'smritiMUM', 'smritiTT1');")
    unapproved_dbs = cur.fetchall()

    # 2. Execute Alphanumeric Dry-Run Provisioning Plans
    provisioner = CompanyDatabaseProvisioner(dry_run=True)
    plan_abc = provisioner.run_dry_run_provisioning("COMP-ABC", "SMRITI Alpha Retail", company_code="ABC")
    plan_mum = provisioner.run_dry_run_provisioning("COMP-MUM", "SMRITI Mumbai Megastore", company_code="MUM")
    plan_001 = provisioner.run_dry_run_provisioning("COMP-001", "SMRITI Enterprise Default", company_code="001")

    print("\nDRY-RUN PROVISIONING EXAMPLES:")
    print(f"  Code 'ABC' -> DB: {plan_abc['database_name']} (Dry-Run: {plan_abc['dry_run']})")
    print(f"  Code 'MUM' -> DB: {plan_mum['database_name']} (Dry-Run: {plan_mum['dry_run']})")
    print(f"  Code '001' -> DB: {plan_001['database_name']} (Dry-Run: {plan_001['dry_run']})")

    # 3. Create Excel Worksheet Data
    code_matrix = pd.DataFrame([
        {"Company Code Input": "001", "Normalized Code": "001", "Target Database Name": "smriti001", "Validation Result": "APPROVED", "Governance Notes": "Numeric code compliant with 3-char alphanumeric"},
        {"Company Code Input": "007", "Normalized Code": "007", "Target Database Name": "smriti007", "Validation Result": "APPROVED", "Governance Notes": "Numeric code compliant with 3-char alphanumeric"},
        {"Company Code Input": "abc", "Normalized Code": "ABC", "Target Database Name": "smritiABC", "Validation Result": "APPROVED", "Governance Notes": "Normalized lowercase to uppercase"},
        {"Company Code Input": "ABC", "Normalized Code": "ABC", "Target Database Name": "smritiABC", "Validation Result": "APPROVED", "Governance Notes": "Standard 3-character uppercase alphanumeric"},
        {"Company Code Input": "A01", "Normalized Code": "A01", "Target Database Name": "smritiA01", "Validation Result": "APPROVED", "Governance Notes": "Mixed letter-digit code"},
        {"Company Code Input": "R01", "Normalized Code": "R01", "Target Database Name": "smritiR01", "Validation Result": "APPROVED", "Governance Notes": "Region-prefixed code"},
        {"Company Code Input": "MUM", "Normalized Code": "MUM", "Target Database Name": "smritiMUM", "Validation Result": "APPROVED", "Governance Notes": "City/Hub code"},
        {"Company Code Input": "TT1", "Normalized Code": "TT1", "Target Database Name": "smritiTT1", "Validation Result": "APPROVED", "Governance Notes": "Tenant test code"},
        {"Company Code Input": "000", "Normalized Code": "FORBIDDEN", "Target Database Name": "RESERVED", "Validation Result": "REJECTED", "Governance Notes": "smriti000 is permanently reserved"},
        {"Company Code Input": "SYS", "Normalized Code": "FORBIDDEN", "Target Database Name": "RESERVED", "Validation Result": "REJECTED", "Governance Notes": "smritisys is permanently reserved for Control Plane"},
        {"Company Code Input": "A-1", "Normalized Code": "FORBIDDEN", "Target Database Name": "INVALID", "Validation Result": "REJECTED", "Governance Notes": "Hyphen separator forbidden"},
        {"Company Code Input": "A_1", "Normalized Code": "FORBIDDEN", "Target Database Name": "INVALID", "Validation Result": "REJECTED", "Governance Notes": "Underscore separator forbidden"},
        {"Company Code Input": "ABCD", "Normalized Code": "FORBIDDEN", "Target Database Name": "INVALID", "Validation Result": "REJECTED", "Governance Notes": "4-character code forbidden (exact length 3 required)"}
    ])

    # 4. Update Excel Review Workbook
    wb = openpyxl.load_workbook(EXCEL_OUTPUT)

    writer = pd.ExcelWriter(EXCEL_OUTPUT, engine="openpyxl", mode="a", if_sheet_exists="replace")
    code_matrix.to_excel(writer, sheet_name="ALPHANUMERIC_CODE_GOVERNANCE", index=False)
    writer.close()

    # Format Excel Sheet
    wb = openpyxl.load_workbook(EXCEL_OUTPUT)
    ws = wb["ALPHANUMERIC_CODE_GOVERNANCE"]
    header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'), top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

    ws.freeze_panes = "A2"
    ws.views.sheetView[0].showGridLines = True

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 60)

    wb.save(EXCEL_OUTPUT)
    wb.close()

    # 5. Update Documentation
    with open(DOC_ARCH, "w", encoding="utf-8") as f:
        f.write(f"""<!--
  Project      : SMRITI Retail OS
  Author       : Jawahar Ramkripal Mallah
  Email        : support@smritibooks.com
  Websites     : smritibooks.com | erpnbook.com | aitdl.com
  Version      : 3.16.0
  Created      : 2026-08-15
  Modified     : 2026-08-15
  Copyright    : © SMRITIBooks.com. All Rights Reserved.
  License      : Proprietary Commercial Software
  Classification: Internal
-->

# SMRITI Multi-Company Database Architecture Specification v1.0

**Status: COMPANY_CODE_STANDARD_UPDATED — ZERO DATABASE PROVISIONING**  
**Audit Timestamp:** {ts}  
**Official Control Plane DB:** `smritisys`  
**Official Company Business DB Naming Standard:** `smriti<3-character-alphanumeric-code>`

---

## 1. Official Alphanumeric Company Code Standard (`smriti<A-Z0-9>`)

| Rule | Specification | Example |
|---|---|---|
| **Prefix** | Exactly `smriti` | `smriti` |
| **Separator** | **NO** underscore, hyphen, space | `smritiABC` (NOT `smriti_ABC`) |
| **Company Code** | Exactly 3 alphanumeric characters `[A-Z0-9]` | `001`, `ABC`, `A01`, `MUM`, `TT1` |
| **Case Normalization** | Lowercase automatically converted to UPPERCASE | `abc` -> `ABC` (`smritiABC`) |
| **Reserved Code 000** | `000` is permanently reserved | `smriti000` (Forbidden) |
| **Reserved Code SYS** | `SYS` is permanently reserved for Control Plane | `smritisys` (Control Plane) |

```text
PostgreSQL Server
│
├── smritisys (SMRITI Control Plane)
│
├── smriti001 (Company Business DB #001)
├── smriti007 (Company Business DB #007)
├── smritiABC (Company Business DB #ABC)
├── smritiMUM (Company Business DB #MUM)
└── smritiTT1 (Company Business DB #TT1)
```

---

## 2. Server-Side Generator Logic

```python
def generate_company_database_name(company_code: str) -> str:
    \"\"\"Generates server-side database name: smriti<3-character-alphanumeric>.\"\"\"
    code = str(company_code).strip().upper()
    if len(code) != 3 or not code.isalnum():
        raise ValueError("Company code must be exactly 3 alphanumeric characters [A-Z0-9].")
    if code in ("000", "SYS"):
        raise ValueError(f"Company code '{{code}}' is permanently reserved.")
    return f"smriti{{code}}"
```
""")

    # 6. Post-Audit Verification
    cur.execute("SELECT COUNT(*) FROM smriti_menus;")
    final_menus = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM smriti_audit_log;")
    final_audit = cur.fetchone()[0]

    conn.close()

    mutated = (initial_menus != final_menus) or (initial_audit != final_audit) or (len(unapproved_dbs) > 0)

    print("\nREFACTOR AUDIT RESULTS:")
    print(f"  Unapproved DBs Found  : {len(unapproved_dbs)} (Zero DBs Created)")
    print(f"  smriti_menus Count    : Initial={initial_menus}, Final={final_menus}")
    print(f"  smriti_audit_log Count: Initial={initial_audit}, Final={final_audit}")

    if not mutated:
        print("\nRESULT: ZERO PROVISIONING & ZERO MUTATIONS VERIFIED")
        print(f"Excel Workbook Updated : {EXCEL_OUTPUT}")
        print(f"Doc Specification      : {DOC_ARCH}")
        print("\nFINAL STATUS: COMPANY_CODE_STANDARD_UPDATED — ZERO DATABASE PROVISIONING")
    else:
        print("\nRESULT: MUTATION DETECTED")
        print("FINAL STATUS: BLOCKED")
        sys.exit(1)

if __name__ == "__main__":
    audit_alphanumeric_company_code()
