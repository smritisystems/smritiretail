#!/usr/bin/env python3
"""
Read Tax Invoices Excel file - Item Level Details
"""
import openpyxl
import pandas as pd

file_path = 'Tax_Invoices_TT_18_to_137_Article_Color_Size_Split.xlsx'

print("=" * 80)
print(f"READING: {file_path}")
print("=" * 80)

wb = openpyxl.load_workbook(file_path)

# Try reading "Item Level Details" sheet
if 'Item Level Details' in wb.sheetnames:
    ws = wb['Item Level Details']
    print(f"\n✓ Reading sheet: {ws.title}")
    print(f"✓ Dimensions: {ws.dimensions}")
    
    # Get headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
        
    print(f"\n✓ Headers ({len(headers)}):")
    for i, h in enumerate(headers, 1):
        print(f"   {i}. {h}")
    
    # Show first 10 rows
    print(f"\n✓ First 10 data rows:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), 2):
        print(f"   {row_idx}: {row}")
    
    # Count total rows
    total_rows = ws.max_row - 1
    print(f"\n✓ Total data rows: {total_rows}")
    
    # Get unique values for analysis
    print(f"\n✓ Sample unique values:")
    
    # Check if we have SKU, Barcode, Article, Color, Size columns
    col_names = {h: i+1 for i, h in enumerate(headers) if h}
    print(f"\n   Column mapping: {col_names}")

wb.close()

print("\n" + "=" * 80)
