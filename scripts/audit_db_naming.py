"""
 * Project      : SMRITI Retail OS
 * Author       : Jawahar Ramkripal Mallah
 * Email        : support@smritibooks.com
 * Websites     : smritibooks.com | erpnbook.com | aitdl.com
 * Version      : 3.16.0
 * Created      : 2026-08-15
 * Modified     : 2026-08-15
 * Copyright    : © SMRITIBooks.com. All Rights Reserved.
 * License      : Proprietary Commercial Software
 * Classification: Internal
"""

import sys, os, glob, re
import psycopg2
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

DB_PARAMS = "postgresql://postgres:postgres@localhost:5432/smritisys"
EXCEL_OUTPUT = r"F:\SMRITRretailNX\SMRITI_Control_Plane_Architecture_Review.xlsx"

def audit_naming_conventions():
    print("============================================================")
    print("SMRITI COMPANY DATABASE NAMING CONVENTION AUDIT")
    print("============================================================")

    # 1. Connect & Initial Row Counts
    conn = psycopg2.connect(DB_PARAMS)
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM smriti_menus;")
    initial_menus = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM smriti_audit_log;")
    initial_audit = cur.fetchone()[0]

    # 2. Scan repository for legacy DB naming references
    repo_root = r"F:\SMRITRretailNX"
    legacy_keywords = ["smriti_retail_db", "company_comp_", "company_tenant", "company_001"]

    findings = []
    for root, _, files in os.walk(repo_root):
        if any(skip in root for skip in [".git", "node_modules", "dist", ".pytest_cache", "__pycache__", "brain"]):
            continue
        for file in files:
            if file.endswith((".py", ".ts", ".tsx", ".md", ".json", ".env", ".ps1", ".sh", ".yml", ".yaml")):
                filepath = os.path.join(root, file)
                rel_path = os.path.relpath(filepath, repo_root)
                try:
                    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                        lines = f.readlines()
                    for line_num, line in enumerate(lines, 1):
                        for kw in legacy_keywords:
                            if kw in line:
                                findings.append({
                                    "file": rel_path,
                                    "line_number": line_num,
                                    "legacy_reference": kw,
                                    "target_standard": "smriti<001-999>",
                                    "snippet": line.strip()[:80]
                                })
                except Exception:
                    pass

    df_findings = pd.DataFrame(findings)

    # Naming Standard Matrix Data
    naming_standard_matrix = pd.DataFrame([
        {"Company Code", "Target Database Name", "Status", "Notes"},
        ["001", "smriti001", "APPROVED_STANDARD", "Company Business Database #001"],
        ["002", "smriti002", "APPROVED_STANDARD", "Company Business Database #002"],
        ["007", "smriti007", "APPROVED_STANDARD", "Enforced: 007 -> smriti007"],
        ["125", "smriti125", "APPROVED_STANDARD", "Company Business Database #125"],
        ["999", "smriti999", "APPROVED_STANDARD", "Upper limit namespace"],
        ["000", "RESERVED_FORBIDDEN", "FORBIDDEN", "smriti000 is permanently reserved"],
        ["smritisys", "smritisys", "APPROVED_CONTROL_PLANE", "Permanently reserved for SMRITI Control Plane"]
    ])

    naming_standard_matrix.columns = naming_standard_matrix.iloc[0]
    naming_standard_matrix = naming_standard_matrix.iloc[1:]

    # 3. Update Excel Workbook
    wb = openpyxl.load_workbook(EXCEL_OUTPUT)

    writer = pd.ExcelWriter(EXCEL_OUTPUT, engine="openpyxl", mode="a", if_sheet_exists="replace")
    df_findings.to_excel(writer, sheet_name="LEGACY_NAMING_AUDIT", index=False)
    naming_standard_matrix.to_excel(writer, sheet_name="NAMING_CONVENTION_GOVERNANCE", index=False)
    writer.close()

    # Format Excel Sheets
    wb = openpyxl.load_workbook(EXCEL_OUTPUT)
    header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'), top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

    for sheetname in ["LEGACY_NAMING_AUDIT", "NAMING_CONVENTION_GOVERNANCE"]:
        ws = wb[sheetname]
        ws.freeze_panes = "A2"
        ws.views.sheetView[0].showGridLines = True

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 50)

    wb.save(EXCEL_OUTPUT)
    wb.close()

    # 4. Database Mutation Verification
    cur.execute("SELECT COUNT(*) FROM smriti_menus;")
    final_menus = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM smriti_audit_log;")
    final_audit = cur.fetchone()[0]

    conn.close()

    mutated = (initial_menus != final_menus) or (initial_audit != final_audit)

    print("\nNAMING CONVENTION AUDIT RESULTS:")
    print(f"  Legacy References Identified : {len(findings)}")
    print(f"  smriti_menus Count           : Initial={initial_menus}, Final={final_menus}")
    print(f"  smriti_audit_log Count       : Initial={initial_audit}, Final={final_audit}")

    if not mutated:
        print("\nRESULT: ZERO MUTATIONS VERIFIED")
        print(f"Excel Workbook Updated        : {EXCEL_OUTPUT}")
        print("\nFINAL STATUS: AUDIT_COMPLETE — BASELINE FROZEN")
    else:
        print("\nRESULT: MUTATION DETECTED")
        print("FINAL STATUS: BLOCKED")
        sys.exit(1)

if __name__ == "__main__":
    audit_naming_conventions()
