#!/usr/bin/env python3
"""
Import items from Tax Invoices Excel to smriti001 itemmaster
Extracts unique Article-Color-Size combinations and creates products
"""
import openpyxl
import psycopg2
from psycopg2.extras import execute_batch
import json
from datetime import datetime
from uuid import uuid4

file_path = 'Tax_Invoices_TT_18_to_137_Article_Color_Size_Split.xlsx'

print("=" * 80)
print("IMPORT: Tax Invoice Items to smriti001 Itemmaster")
print("=" * 80)

# Read Excel file
wb = openpyxl.load_workbook(file_path)
ws = wb['Item Level Details']

# Extract unique products
unique_products = {}
row_count = 0

print(f"\n✓ Reading {ws.max_row - 1} rows from Excel...")

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    if row[12] is None:  # SKU Code column (index 12)
        continue
    
    row_count += 1
    
    sku_code = str(row[12]).strip()  # SKU Code (Item / SKU Code)
    article = str(row[9]).strip() if row[9] else ''  # Article
    color = str(row[10]).strip() if row[10] else ''  # Color
    size = str(row[11]).strip() if row[11] else ''  # Size
    description = str(row[13]).strip() if row[13] else ''  # Item Description
    hsn_code = str(row[14]).strip() if row[14] else ''  # HSN / SAC Code
    mrp = float(row[16]) if row[16] else 0  # MRP
    
    if sku_code not in unique_products:
        unique_products[sku_code] = {
            'sku': sku_code,
            'article': article,
            'color': color,
            'size': size,
            'description': description,
            'hsn_code': hsn_code,
            'mrp': mrp,
            'name': f"{article} {color} {size}"
        }

print(f"   Read {row_count} rows from Excel")
print(f"   Found {len(unique_products)} unique SKU codes")

# Connect to smriti001
conn = psycopg2.connect('postgresql://postgres:postgres@localhost:5432/smriti001')
cur = conn.cursor()

print(f"\n✓ Connected to smriti001")

# Get company ID for Tattly Threads
cur.execute("SELECT id FROM companies WHERE name LIKE '%Tattly%' LIMIT 1")
result = cur.fetchone()
if result:
    company_id = result[0]
    print(f"   Company: {company_id}")
else:
    cur.execute("SELECT id FROM companies LIMIT 1")
    company_id = cur.fetchone()[0]
    print(f"   Company: {company_id} (default)")

# Get a branch ID
cur.execute("SELECT id FROM branches LIMIT 1")
branch_id = cur.fetchone()[0]
print(f"   Branch: {branch_id}")

# Check existing products
cur.execute("SELECT COUNT(*) FROM products")
existing_count = cur.fetchone()[0]
print(f"   Existing products: {existing_count}")

# Check if SKUs already exist
cur.execute("""
    SELECT sku FROM products WHERE is_deleted = FALSE
""")
existing_skus = {row[0] for row in cur.fetchall() if row[0]}
print(f"   Existing SKUs: {len(existing_skus)}")

# Also check existing article-color-size combinations
cur.execute("""
    SELECT DISTINCT LOWER(style_code), LOWER(color), LOWER(size)
    FROM products
    WHERE company_id = %s AND is_deleted = FALSE
""", (company_id,))
existing_variants = {(row[0], row[1], row[2]) for row in cur.fetchall()}
print(f"   Existing variants: {len(existing_variants)}")

# Filter unique_products
new_products = {}
skipped_sku = 0
skipped_variant = 0

for k, v in unique_products.items():
    if k in existing_skus:
        skipped_sku += 1
        continue
    
    variant_key = (
        v['article'].lower(),
        v['color'].lower(),
        v['size'].lower()
    )
    if variant_key in existing_variants:
        skipped_variant += 1
        continue
    
    new_products[k] = v

print(f"   Skipped (SKU exists): {skipped_sku}")
print(f"   Skipped (variant exists): {skipped_variant}")
print(f"   After filtering: {len(new_products)} truly new products")

# Prepare insert data
insert_data = []
for sku, product in new_products.items():
    # Generate a numeric code from SKU
    code = sku.replace('-', '')[:20]  # Use SKU as code
    
    insert_data.append((
        code,  # code (NOT NULL, PK)
        product['name'],  # name (NOT NULL)
        float(product['mrp']),  # price (NOT NULL) - use MRP as selling price
        0,  # stock (NOT NULL) - default 0
        'FOOTWEAR',  # category (NOT NULL) - inferred from data
        f"{sku}-barcode",  # barcode (NOT NULL) - generated from SKU
        float(product['mrp']),  # mrp (NOT NULL)
        5.0,  # gst_percentage (NOT NULL) - from Excel data
        product['sku'],  # sku (optional)
        product['article'],  # style_code (optional)
        float(product['mrp']) * 0.75,  # cost_price (optional) - assume 25% margin
        product['color'],  # color (optional)
        product['size'],  # size (optional)
        product['hsn_code'],  # hsn_code (NOT NULL)
        0,  # reserved_stock (NOT NULL)
        str(uuid4()),  # id
        str(uuid4()),  # uuid
        company_id,  # company_id (use fetched company)
        branch_id,  # branch_id (use fetched branch)
        datetime.now().isoformat(),  # created_at
        datetime.now().isoformat(),  # modified_at
        'SYSTEM',  # created_by
        'SYSTEM',  # updated_by
        True,  # is_active
        False,  # is_deleted
        None,  # deleted_at
        None,  # deleted_by
        1,  # version
    ))

# Insert products
if insert_data:
    print(f"\n✓ Inserting {len(insert_data)} new products...")
    
    insert_sql = '''
        INSERT INTO products (
            code, name, price, stock, category, barcode, mrp, gst_percentage,
            sku, style_code, cost_price, color, size, hsn_code, reserved_stock,
            id, uuid, company_id, branch_id, created_at, modified_at,
            created_by, updated_by, is_active, is_deleted, deleted_at,
            deleted_by, version
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        )
    '''
    
    try:
        execute_batch(cur, insert_sql, insert_data, page_size=100)
        conn.commit()
        print(f"   ✅ Successfully inserted {len(insert_data)} products")
    except Exception as e:
        conn.rollback()
        print(f"   ❌ Error: {e}")
        cur.close()
        conn.close()
        exit(1)
else:
    print("\n✓ All products already imported")

# Verify import
cur.execute("SELECT COUNT(*) FROM products")
final_count = cur.fetchone()[0]
print(f"\n✓ Final product count: {final_count}")
print(f"   Added: {final_count - existing_count}")

cur.close()
conn.close()

print("\n" + "=" * 80)
print("✅ IMPORT COMPLETE")
print("=" * 80)
