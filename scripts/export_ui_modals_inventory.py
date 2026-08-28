from pathlib import Path
import re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "reports" / "frontend_ui_modals_inventory.xlsx"

patterns = ["*Modal.tsx", "*Modal.ts", "*Dialog.tsx", "*Dlg.tsx", "*Dialog.ts", "*Dlg.ts"]
files = sorted({p for pattern in patterns for p in (ROOT / "src").rglob(pattern)})

def humanize(value: str) -> str:
    value = re.sub(r"(Modal|Dialog|Dlg)$", "", value)
    value = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", value)
    value = value.replace("CRM", "CRM").replace("Pdt", "PDT").replace("PO", "PO")
    return value.strip() or "UI Dialog"

def component_name(text: str, path: Path) -> str:
    matches = re.findall(r"export\s+(?:const|function|default\s+function)\s+([A-Z][A-Za-z0-9_]*)", text)
    if matches:
        return matches[0]
    matches = re.findall(r"(?:const|function)\s+([A-Z][A-Za-z0-9_]*)\s*[:=(]", text)
    return matches[0] if matches else path.stem

def props_summary(text: str) -> str:
    match = re.search(r"interface\s+\w*Props\s*\{(.*?)\}", text, re.S)
    if not match:
        match = re.search(r"type\s+\w*Props\s*=\s*\{(.*?)\}", text, re.S)
    if not match:
        return "Not declared"
    props = []
    for line in match.group(1).splitlines():
        item = re.match(r"\s*([A-Za-z0-9_]+)\??\s*:", line)
        if item:
            props.append(item.group(1))
    return ", ".join(props) if props else "Declared inline"

def domain_for(path: Path) -> str:
    relative = path.relative_to(ROOT / "src").parts
    return relative[1] if len(relative) > 1 else "components"

source_texts = {}
for source in (ROOT / "src").rglob("*.ts"):
    try:
        source_texts[source] = source.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        pass
for source in (ROOT / "src").rglob("*.tsx"):
    try:
        source_texts[source] = source.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        pass

def references_for(path: Path) -> int:
    needle = path.stem
    return sum(1 for source, text in source_texts.items() if source != path and needle in text)

rows = []
for path in files:
    text = path.read_text(encoding="utf-8", errors="ignore")
    relative = path.relative_to(ROOT).as_posix()
    kind = "Modal" if path.name.endswith("Modal.tsx") or path.name.endswith("Modal.ts") else ("Dialog" if "Dialog" in path.name else "Dialog-style")
    references = references_for(path)
    rows.append({
        "UI Type": kind,
        "Domain": domain_for(path),
        "Component": component_name(text, path),
        "File": relative,
        "Purpose": humanize(path.stem),
        "Props": props_summary(text),
        "Source References": references,
        "Status": "Referenced" if references > 0 else "Needs wiring review",
    })

wb = Workbook()
ws = wb.active
ws.title = "UI Modals"
headers = list(rows[0].keys()) if rows else ["UI Type", "Domain", "Component", "File", "Purpose", "Props", "Source References", "Status"]
ws.append(headers)
for row in rows:
    ws.append([row[h] for h in headers])

header_fill = PatternFill("solid", fgColor="1F4E78")
for cell in ws[1]:
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
widths = {"A": 14, "B": 18, "C": 34, "D": 65, "E": 34, "F": 48, "G": 18, "H": 22}
for col, width in widths.items():
    ws.column_dimensions[col].width = width
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

summary = wb.create_sheet("Summary")
summary.append(["Metric", "Value"])
summary.append(["Total UI modal/dialog files", len(rows)])
summary.append(["Modal files", sum(1 for r in rows if r["UI Type"] == "Modal")])
summary.append(["Dialog-style files", sum(1 for r in rows if r["UI Type"] != "Modal")])
summary.append(["Referenced files", sum(1 for r in rows if r["Source References"] > 0)])
summary.append(["Needs wiring review", sum(1 for r in rows if r["Source References"] == 0)])
summary.append([])
summary.append(["Domain", "Count"])
for domain, count in sorted(Counter(r["Domain"] for r in rows).items()):
    summary.append([domain, count])
for cell in summary[1]:
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = header_fill
summary.freeze_panes = "A2"
summary.column_dimensions["A"].width = 32
summary.column_dimensions["B"].width = 18

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(f"Wrote {OUTPUT}")
print(f"Inventory rows: {len(rows)}")
