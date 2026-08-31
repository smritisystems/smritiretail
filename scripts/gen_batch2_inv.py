"""
Project      : SMRITI Retail OS
Author       : Jawahar Ramkripal Mallah
Designation  : Chief Systems Architect & Creator
Email        : support@smritibooks.com
Websites     : smritibooks.com | erpnbook.com | aitdl.com
Version      : 5.0.0
Created      : 2026-08-25
Modified     : 2026-08-25
Copyright    : © SMRITIBooks.com. All Rights Reserved.
License      : Proprietary Commercial Software
Classification: Internal
"""

import os
import sys
import io
import re
import json
import base64
import uuid
import datetime
import hashlib
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import psycopg2
import psycopg2.extras
import asyncio

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

sys.path.insert(0, r"F:\SMRITRretailNX\backend")

from app.services.invoice_pdf_service import InvoicePdfService, number_to_indian_words

# Paths
DISPATCH_PATH = r"F:\Smriti-Clients Data\Tattly Threads\24_08_26\RIL_Dispatch_09-08-2026-2.xlsx"
PO_PATH = r"F:\Smriti-Clients Data\Tattly Threads\All Stores Po.xlsx"
CUST_LIST_PATH = r"F:\Smriti-Clients Data\Tattly Threads\Invoice\RIL FINAL LIST.xlsx"

OUTPUT_CLIENT_DIR = r"F:\Smriti-Clients Data\Tattly Threads\24_08_26\Invoices"
OUTPUT_EXPORTS_DIR = r"F:\SMRITRretailNX\exports\tt_batch_129_137"
OUTPUT_TT_DIR = r"F:\SMRITRretailNX\TT"
CONFIRMATION_EXCEL_PATH = r"F:\Smriti-Clients Data\Tattly Threads\24_08_26\Tax_Invoice_Store_PO_Address_Confirmation_Batch2_2026-08-25.xlsx"

os.makedirs(OUTPUT_CLIENT_DIR, exist_ok=True)
os.makedirs(OUTPUT_EXPORTS_DIR, exist_ok=True)
os.makedirs(OUTPUT_TT_DIR, exist_ok=True)

INVOICE_DATE_DB = datetime.date(2026, 8, 24)
INVOICE_DATE_STR = "24-08-2026"
START_INVOICE_NUM = 129
DB_URL = "postgresql://postgres:postgres@localhost:5432/smriti001"


class InvoiceItemMock:
    def __init__(self, line_no, code, name, qty, price, mrp, disc_pct, taxable_value, hsn_code, gst_rate, tax_amount, cgst_amount, sgst_amount, igst_amount, total_amount):
        self.line_no = line_no
        self.code = code
        self.name = name
        self.quantity = qty
        self.price = price
        self.mrp = mrp
        self.disc_pct = disc_pct
        self.taxable_value = taxable_value
        self.hsn_code = hsn_code
        self.gst_rate = gst_rate
        self.tax_amount = tax_amount
        self.cgst_amount = cgst_amount
        self.sgst_amount = sgst_amount
        self.igst_amount = igst_amount
        self.total_amount = total_amount


class InvoiceMock:
    def __init__(self, id, invoice_no, date, sis_code, pos_state, po_reference, eway_bill_no, customer_name, customer_gstin, site_name, billing_address, shipping_address, taxable_value, tax_total, grand_total, is_interstate, bank_name, account_no, ifsc_code, bank_branch, items):
        self.id = id
        self.invoice_no = invoice_no
        self.date = date
        self.sis_code = sis_code
        self.pos_state = pos_state
        self.po_reference = po_reference
        self.eway_bill_no = eway_bill_no
        self.customer_name = customer_name
        self.customer_gstin = customer_gstin
        self.site_name = site_name
        self.billing_address = billing_address
        self.shipping_address = shipping_address
        self.taxable_value = taxable_value
        self.tax_total = tax_total
        self.grand_total = grand_total
        self.is_interstate = is_interstate
        self.bank_name = bank_name
        self.account_no = account_no
        self.ifsc_code = ifsc_code
        self.bank_branch = bank_branch
        self.items = items
        self.status = "COMPLETED"
        self.reverse_charge = False
        self.is_reverse_charge = False
        self.irn = None
        self.signed_qr_payload = None
        self.e_invoice_status = "NOT_APPLICABLE"


def generate_confirmation_workbook(store_groups, po_lookup, cust_lookup, state_gstin_map):
    out_wb = openpyxl.Workbook()

    # Sheet 1: Store_PO_Address_Confirmation
    ws_out1 = out_wb.active
    ws_out1.title = "Store_PO_Address_Confirmation"
    ws_out1.views.sheetView[0].showGridLines = True

    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    title_font = Font(name="Segoe UI", size=13, bold=True, color="1E3A8A")
    sub_font = Font(name="Segoe UI", size=9.5, italic=True, color="4B5563")
    data_font = Font(name="Segoe UI", size=9)
    mono_font = Font(name="Consolas", size=9)

    border_thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    ws_out1['A1'] = "SMRITI RETAIL OS — TATTLY THREADS TAX INVOICE PRE-GENERATION CONFIRMATION REGISTER (BATCH 2)"
    ws_out1['A1'].font = title_font
    ws_out1['A2'] = f"Invoice Date: 24-Aug-2026 | Total Stores to Invoice: {len(store_groups)} | Source: RIL_Dispatch_09-08-2026-2.xlsx (Sheet2)"
    ws_out1['A2'].font = sub_font

    headers = [
        "Sr.",
        "Invoice No.",
        "Store / SIS Code",
        "Purchase Order No.",
        "Store / Site Name",
        "State",
        "Customer GSTIN",
        "Billing Address (Recipient)",
        "Shipping Address (Delivery Site)",
        "Line Items",
        "Total Pairs (Qty)",
        "Carton No(s)"
    ]

    ws_out1.row_dimensions[4].height = 28
    for c_idx, h in enumerate(headers, start=1):
        cell = ws_out1.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    row_idx = 5
    total_all_pairs = 0
    total_all_items = 0
    curr_seq = START_INVOICE_NUM

    for sr, (sis_code, sdata) in enumerate(store_groups.items(), start=1):
        inv_no = f"TT2026-2027/{curr_seq}"
        po_info = po_lookup.get(sis_code, {})
        cust_info = cust_lookup.get(sis_code, {})

        po_no = po_info.get("po_number") or ("5182778158" if sis_code == "TW07" else "5182778151")
        site_name = cust_info.get("site_name") or po_info.get("site_name") or f"Reliance Retail Store {sis_code}"
        site_addr = cust_info.get("address") or po_info.get("site_address") or f"Reliance Retail Store, SIS Code {sis_code}"
        state = cust_info.get("state") or "KARNATAKA"
        gstin = cust_info.get("gstin") or state_gstin_map.get(state.upper(), "29AABCR1718E1ZL")

        billing_addr = f"Reliance Retail Limited, {site_addr}"
        shipping_addr = f"Reliance Retail Limited ({site_name}), {site_addr}"

        item_count = len(sdata["raw_rows"])
        pairs = sdata["total_pairs"]
        cartons_str = ", ".join(sorted(list(sdata["cartons"]))) if sdata["cartons"] else "N/A"

        total_all_pairs += pairs
        total_all_items += item_count

        ws_out1.cell(row=row_idx, column=1, value=sr).alignment = Alignment(horizontal="center", vertical="center")
        ws_out1.cell(row=row_idx, column=2, value=inv_no).alignment = Alignment(horizontal="center", vertical="center")
        ws_out1.cell(row=row_idx, column=3, value=sis_code).alignment = Alignment(horizontal="center", vertical="center")
        ws_out1.cell(row=row_idx, column=4, value=po_no).alignment = Alignment(horizontal="center", vertical="center")
        ws_out1.cell(row=row_idx, column=5, value=site_name).alignment = Alignment(horizontal="left", vertical="center")
        ws_out1.cell(row=row_idx, column=6, value=state).alignment = Alignment(horizontal="center", vertical="center")
        ws_out1.cell(row=row_idx, column=7, value=gstin).alignment = Alignment(horizontal="center", vertical="center")
        ws_out1.cell(row=row_idx, column=8, value=billing_addr).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_out1.cell(row=row_idx, column=9, value=shipping_addr).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_out1.cell(row=row_idx, column=10, value=item_count).alignment = Alignment(horizontal="right", vertical="center")
        ws_out1.cell(row=row_idx, column=11, value=pairs).alignment = Alignment(horizontal="right", vertical="center")
        ws_out1.cell(row=row_idx, column=12, value=cartons_str).alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 13):
            c = ws_out1.cell(row=row_idx, column=col)
            c.font = data_font
            c.border = border_thin
            if col in [1, 2, 3, 4, 7, 10, 11]:
                c.font = mono_font

        ws_out1.row_dimensions[row_idx].height = 36
        row_idx += 1
        curr_seq += 1

    # Total Summary Row
    total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    ws_out1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
    tot_cell = ws_out1.cell(row=row_idx, column=1, value="GRAND TOTAL")
    tot_cell.font = Font(name="Segoe UI", size=10, bold=True)
    tot_cell.alignment = Alignment(horizontal="right", vertical="center")

    c_items = ws_out1.cell(row=row_idx, column=10, value=total_all_items)
    c_items.font = Font(name="Consolas", size=10, bold=True)
    c_items.alignment = Alignment(horizontal="right", vertical="center")

    c_pairs = ws_out1.cell(row=row_idx, column=11, value=total_all_pairs)
    c_pairs.font = Font(name="Consolas", size=10, bold=True)
    c_pairs.alignment = Alignment(horizontal="right", vertical="center")

    for col in range(1, 13):
        c = ws_out1.cell(row=row_idx, column=col)
        c.fill = total_fill
        c.border = border_thin

    ws_out1.row_dimensions[row_idx].height = 24

    col_widths = {1: 6, 2: 18, 3: 16, 4: 16, 5: 28, 6: 16, 7: 18, 8: 42, 9: 42, 10: 12, 11: 16, 12: 16}
    for col, width in col_widths.items():
        ws_out1.column_dimensions[get_column_letter(col)].width = width

    # Sheet 2: Items Breakdown
    ws_out2 = out_wb.create_sheet(title="Dispatch_Items_Breakdown")
    ws_out2.views.sheetView[0].showGridLines = True

    headers_items = ["Invoice No.", "Store / SIS Code", "PO Number", "Article", "Color", "MRP", "36", "37", "38", "39", "40", "41", "42", "Total Pairs", "Packing Date", "Carton No"]
    ws_out2.row_dimensions[1].height = 24
    for c_idx, h in enumerate(headers_items, start=1):
        cell = ws_out2.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin

    item_row_idx = 2
    curr_seq = START_INVOICE_NUM
    for sis_code, sdata in store_groups.items():
        inv_no = f"TT2026-2027/{curr_seq}"
        po_no = po_lookup.get(sis_code, {}).get("po_number") or ("5182778158" if sis_code == "TW07" else "5182778151")
        for itm in sdata["raw_rows"]:
            ws_out2.cell(row=item_row_idx, column=1, value=inv_no).alignment = Alignment(horizontal="center")
            ws_out2.cell(row=item_row_idx, column=2, value=sis_code).alignment = Alignment(horizontal="center")
            ws_out2.cell(row=item_row_idx, column=3, value=po_no).alignment = Alignment(horizontal="center")
            ws_out2.cell(row=item_row_idx, column=4, value=itm["article"]).alignment = Alignment(horizontal="left")
            ws_out2.cell(row=item_row_idx, column=5, value=itm["color"]).alignment = Alignment(horizontal="left")
            ws_out2.cell(row=item_row_idx, column=6, value=itm["mrp"]).alignment = Alignment(horizontal="right")
            ws_out2.cell(row=item_row_idx, column=7, value=itm["sizes"][36]).alignment = Alignment(horizontal="right")
            ws_out2.cell(row=item_row_idx, column=8, value=itm["sizes"][37]).alignment = Alignment(horizontal="right")
            ws_out2.cell(row=item_row_idx, column=9, value=itm["sizes"][38]).alignment = Alignment(horizontal="right")
            ws_out2.cell(row=item_row_idx, column=10, value=itm["sizes"][39]).alignment = Alignment(horizontal="right")
            ws_out2.cell(row=item_row_idx, column=11, value=itm["sizes"][40]).alignment = Alignment(horizontal="right")
            ws_out2.cell(row=item_row_idx, column=12, value=itm["sizes"][41]).alignment = Alignment(horizontal="right")
            ws_out2.cell(row=item_row_idx, column=13, value=itm["sizes"][42]).alignment = Alignment(horizontal="right")
            ws_out2.cell(row=item_row_idx, column=14, value=itm["total_qty"]).alignment = Alignment(horizontal="right")

            pdate_str = str(itm["packing_date"])[:10] if itm["packing_date"] else ""
            ws_out2.cell(row=item_row_idx, column=15, value=pdate_str).alignment = Alignment(horizontal="center")
            ws_out2.cell(row=item_row_idx, column=16, value=str(itm["carton_no"] or "")).alignment = Alignment(horizontal="center")

            for c in range(1, 17):
                ws_out2.cell(row=item_row_idx, column=c).border = border_thin
                ws_out2.cell(row=item_row_idx, column=c).font = data_font

            item_row_idx += 1
        curr_seq += 1

    for col in range(1, 17):
        ws_out2.column_dimensions[get_column_letter(col)].width = 14

    out_wb.save(CONFIRMATION_EXCEL_PATH)
    print(f"Confirmation workbook saved: {CONFIRMATION_EXCEL_PATH}")


async def main():
    print("=" * 80)
    print("SMRITI RETAIL OS — BATCH 2 TAX INVOICE GENERATOR (INVOICES 129 TO 137)")
    print("=" * 80)

    # 1. Connect to DB
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # 2. Load Customer Master and PO Registers
    wb_cust = openpyxl.load_workbook(CUST_LIST_PATH, data_only=True)
    ws_cust = wb_cust['Sheet1']
    cust_lookup = {}
    state_gstin_map = {}

    for r in range(2, ws_cust.max_row + 1):
        sis = str(ws_cust.cell(row=r, column=3).value or '').strip()
        sname = str(ws_cust.cell(row=r, column=4).value or '').strip()
        state = str(ws_cust.cell(row=r, column=5).value or '').strip()
        city = str(ws_cust.cell(row=r, column=6).value or '').strip()
        pincode = str(ws_cust.cell(row=r, column=8).value or '').strip()
        addr = str(ws_cust.cell(row=r, column=9).value or '').strip()
        gstin = str(ws_cust.cell(row=r, column=10).value or '').strip()

        if state and gstin and gstin != '#N/A':
            state_gstin_map[state.upper()] = gstin

        if sis and sis != '#N/A':
            cust_lookup[sis] = {
                "sis_code": sis,
                "site_name": sname if sname != '#N/A' else '',
                "state": state if state != '#N/A' else '',
                "city": city if city != '#N/A' else '',
                "pincode": pincode if pincode != '#N/A' else '',
                "address": addr if addr != '#N/A' else '',
                "gstin": gstin if gstin != '#N/A' else ''
            }

    wb_po = openpyxl.load_workbook(PO_PATH, data_only=True)
    ws_po = wb_po['PO_Summary_Register']
    po_lookup = {}
    for r in range(2, ws_po.max_row + 1):
        po_no = str(ws_po.cell(row=r, column=1).value or '').strip()
        sis_code = str(ws_po.cell(row=r, column=2).value or '').strip()
        site_name = str(ws_po.cell(row=r, column=3).value or '').strip()
        site_address = str(ws_po.cell(row=r, column=4).value or '').strip()
        buyer = str(ws_po.cell(row=r, column=6).value or '').strip()

        if sis_code:
            po_lookup[sis_code] = {
                "po_number": po_no,
                "sis_code": sis_code,
                "site_name": site_name,
                "site_address": site_address,
                "buyer": buyer
            }

    # 3. Read Dispatch Matrix from Sheet2
    wb_disp = openpyxl.load_workbook(DISPATCH_PATH, data_only=False)
    ws_disp = wb_disp['Sheet2']

    store_groups = {}
    size_cols = [36, 37, 38, 39, 40, 41, 42]

    for r in range(2, ws_disp.max_row + 1):
        sis_val = ws_disp.cell(row=r, column=1).value
        if sis_val is None or str(sis_val).strip() == '':
            continue
        sis_code = str(sis_val).strip()
        article = str(ws_disp.cell(row=r, column=2).value or '').strip().upper()
        color = str(ws_disp.cell(row=r, column=3).value or '').strip().upper()
        mrp_val = ws_disp.cell(row=r, column=4).value or 0
        mrp = Decimal(str(mrp_val))

        qty_36 = int(float(ws_disp.cell(row=r, column=5).value or 0))
        qty_37 = int(float(ws_disp.cell(row=r, column=6).value or 0))
        qty_38 = int(float(ws_disp.cell(row=r, column=7).value or 0))
        qty_39 = int(float(ws_disp.cell(row=r, column=8).value or 0))
        qty_40 = int(float(ws_disp.cell(row=r, column=9).value or 0))
        qty_41 = int(float(ws_disp.cell(row=r, column=10).value or 0))
        qty_42 = int(float(ws_disp.cell(row=r, column=11).value or 0))
        total_qty = int(float(ws_disp.cell(row=r, column=12).value or (qty_36+qty_37+qty_38+qty_39+qty_40+qty_41+qty_42)))

        p_date = ws_disp.cell(row=r, column=13).value
        carton = ws_disp.cell(row=r, column=14).value

        if sis_code not in store_groups:
            store_groups[sis_code] = {
                "sis_code": sis_code,
                "items": [],
                "raw_rows": [],
                "total_pairs": 0,
                "cartons": set()
            }

        store_groups[sis_code]["raw_rows"].append({
            "article": article,
            "color": color,
            "mrp": mrp,
            "sizes": {36: qty_36, 37: qty_37, 38: qty_38, 39: qty_39, 40: qty_40, 41: qty_41, 42: qty_42},
            "total_qty": total_qty,
            "packing_date": p_date,
            "carton_no": carton,
            "excel_row": r
        })
        store_groups[sis_code]["total_pairs"] += total_qty
        if carton:
            store_groups[sis_code]["cartons"].add(str(carton))

        for sz in size_cols:
            c_idx = size_cols.index(sz) + 5
            qty_val = ws_disp.cell(row=r, column=c_idx).value
            if qty_val is not None and float(qty_val) > 0:
                qty = int(float(qty_val))
                store_groups[sis_code]["items"].append({
                    "article": article,
                    "color": color,
                    "size": str(sz),
                    "qty": qty,
                    "mrp": mrp
                })

    # Generate Confirmation Workbook
    generate_confirmation_workbook(store_groups, po_lookup, cust_lookup, state_gstin_map)

    # 4. Generate Invoices & PDFs
    current_seq = START_INVOICE_NUM
    rendered_invoices = []

    from playwright.async_api import async_playwright
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage", "--disable-gpu"]
        )
        page = await browser.new_page()

        for sis_code, sdata in store_groups.items():
            inv_no = f"TT2026-2027/{current_seq}"
            inv_id = f"inv-tt-2026-2027-{current_seq}"

            po_info = po_lookup.get(sis_code, {})
            cust_info = cust_lookup.get(sis_code, {})

            po_no = po_info.get("po_number") or ("5182778158" if sis_code == "TW07" else "5182778151")
            site_name = cust_info.get("site_name") or po_info.get("site_name") or f"Reliance Retail Site {sis_code}"
            site_addr = cust_info.get("address") or po_info.get("site_address") or f"Reliance Retail Store, SIS Code {sis_code}"
            state = cust_info.get("state") or "KARNATAKA"
            gstin = cust_info.get("gstin") or state_gstin_map.get(state.upper(), "29AABCR1718E1ZL")

            is_interstate = not gstin.startswith("27")

            subtotal_taxable = Decimal("0.00")
            total_tax = Decimal("0.00")
            cgst_total = Decimal("0.00")
            sgst_total = Decimal("0.00")
            igst_total = Decimal("0.00")
            total_qty = 0

            item_mocks = []
            for l_idx, item in enumerate(sdata["items"], start=1):
                art = item["article"]
                col = item["color"]
                sz = item["size"]
                qty = item["qty"]
                mrp = item["mrp"]

                sku_code = f"{art}-{col}-{sz}".replace(" ", "-")
                prod_name = f"{art} {col} {sz}"
                hsn = "64041990"

                unit_taxable_rate = (mrp * Decimal("0.5624")).quantize(Decimal("0.01"))
                line_taxable = (unit_taxable_rate * Decimal(qty)).quantize(Decimal("0.01"))
                subtotal_taxable += line_taxable
                total_qty += qty

                if is_interstate:
                    line_tax = (line_taxable * Decimal("0.05")).quantize(Decimal("0.01"))
                    line_cgst = Decimal("0.00")
                    line_sgst = Decimal("0.00")
                    line_igst = line_tax
                else:
                    line_cgst = (line_taxable * Decimal("0.025")).quantize(Decimal("0.01"))
                    line_sgst = (line_taxable * Decimal("0.025")).quantize(Decimal("0.01"))
                    line_igst = Decimal("0.00")
                    line_tax = line_cgst + line_sgst

                line_total = line_taxable + line_tax
                total_tax += line_tax
                cgst_total += line_cgst
                sgst_total += line_sgst
                igst_total += line_igst

                item_mocks.append(InvoiceItemMock(
                    line_no=l_idx,
                    code=sku_code,
                    name=prod_name,
                    qty=qty,
                    price=unit_taxable_rate,
                    mrp=mrp,
                    disc_pct=Decimal("43.76"),
                    taxable_value=line_taxable,
                    hsn_code=hsn,
                    gst_rate=Decimal("5.00"),
                    tax_amount=line_tax,
                    cgst_amount=line_cgst,
                    sgst_amount=line_sgst,
                    igst_amount=line_igst,
                    total_amount=line_total
                ))

            pre_round = subtotal_taxable + total_tax
            grand_total = round(pre_round)
            round_adj = grand_total - pre_round

            billing_full = f"Reliance Retail Limited\n{site_addr}"
            shipping_full = f"Reliance Retail Limited ({site_name})\n{site_addr}"

            inv_mock = InvoiceMock(
                id=inv_id,
                invoice_no=inv_no,
                date=INVOICE_DATE_DB,
                sis_code=sis_code,
                pos_state=state,
                po_reference=po_no,
                eway_bill_no="",
                customer_name="Reliance Retail Limited",
                customer_gstin=gstin,
                site_name=site_name,
                billing_address=billing_full,
                shipping_address=shipping_full,
                taxable_value=subtotal_taxable,
                tax_total=total_tax,
                grand_total=grand_total,
                is_interstate=is_interstate,
                bank_name="STATE BANK OF INDIA",
                account_no="43976711765",
                ifsc_code="SBIN0030425",
                bank_branch="WARDHMAN NAGAR NAGPUR",
                items=item_mocks
            )

            # Generate plain Tax Invoice HTML with VERIFY INVOICE QR (matching earlier batches)
            html_content = InvoicePdfService.generate_invoice_html_from_model(
                invoice=inv_mock,
                company_name="TATTLY THREADS",
                company_gstin="27AAXFT2508H1ZR",
                extra_meta={
                    "company_website": "www.tattlythreads.com",
                    "dispatch_email": "dispatch@tattlythreads.com",
                    "accounts_email": "accounts@tattlythreads.com"
                }
            )

            # Render PDF with Playwright
            await page.set_content(html_content, wait_until="networkidle")
            pdf_bytes = await page.pdf(
                format="A4",
                print_background=True,
                margin={"top": "8mm", "bottom": "10mm", "left": "8mm", "right": "8mm"}
            )

            # Compute SHA-256
            sha256_hash = hashlib.sha256(pdf_bytes).hexdigest()
            pdf_filename = f"SIS_{sis_code}_TaxInvoice_TT2026-2027_{current_seq}.pdf"

            path_client = os.path.join(OUTPUT_CLIENT_DIR, pdf_filename)
            path_exports = os.path.join(OUTPUT_EXPORTS_DIR, pdf_filename)
            path_tt = os.path.join(OUTPUT_TT_DIR, pdf_filename)

            with open(path_client, "wb") as f:
                f.write(pdf_bytes)
            with open(path_exports, "wb") as f:
                f.write(pdf_bytes)
            with open(path_tt, "wb") as f:
                f.write(pdf_bytes)

            # Update Database smriti001
            cur.execute("""
                DELETE FROM sales_invoice_items WHERE invoice_id = %s;
                DELETE FROM invoice_document_artifacts WHERE invoice_id = %s;
                DELETE FROM sales_invoices WHERE id = %s OR invoice_no = %s;
            """, (inv_id, inv_id, inv_id, inv_no))

            cur.execute("""
                INSERT INTO sales_invoices (
                    id, uuid, invoice_no, date, customer_id, customer_name, customer_gstin,
                    billing_address, shipping_address, site_name, sis_code, pos_state,
                    po_reference, taxable_value, tax_total, grand_total, rounding_amount,
                    amount_in_words, is_interstate, status, company_id, branch_id,
                    bank_name, account_no, ifsc_code, original_pdf_sha256, original_pdf_path,
                    original_pdf_size, original_pdf_pages, irn, signed_qr_payload, e_invoice_status,
                    rule_snapshots, created_at, modified_at
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, NULL, NULL, 'NOT_APPLICABLE',
                    %s, NOW(), NOW()
                )
            """, (
                inv_id, str(uuid.uuid4()), inv_no, INVOICE_DATE_DB, 'cust-rrl-192b561d', 'Reliance Retail Limited', gstin,
                billing_full, shipping_full, site_name, sis_code, state,
                po_no, subtotal_taxable, total_tax, grand_total, round_adj,
                number_to_indian_words(float(grand_total)),
                is_interstate, 'COMPLETED', 'COMP-001', 'MAIN',
                "STATE BANK OF INDIA", "43976711765", "SBIN0030425", sha256_hash, path_client,
                len(pdf_bytes), 1, json.dumps({"engine": "InvoicePdfService", "version": "5.0.0"})
            ))

            for itm in item_mocks:
                cur.execute("""
                    INSERT INTO sales_invoice_items (
                        invoice_id, code, name, quantity, price, mrp, disc_pct,
                        taxable_value, hsn_code, gst_rate, tax_amount, cgst_amount,
                        sgst_amount, igst_amount, total_amount, line_no
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s
                    )
                """, (
                    inv_id, itm.code, itm.name, itm.quantity, itm.price, itm.mrp, itm.disc_pct,
                    itm.taxable_value, itm.hsn_code, itm.gst_rate, itm.tax_amount, itm.cgst_amount,
                    itm.sgst_amount, itm.igst_amount, itm.total_amount, itm.line_no
                ))

            # Insert document artifact
            artifact_id = str(uuid.uuid4())
            cur.execute("""
                INSERT INTO invoice_document_artifacts (
                    id, uuid, company_id, branch_id, invoice_id, invoice_no,
                    document_type, template_code, template_version, template_status,
                    storage_path, sha256_hash, file_size, page_count,
                    generated_at, created_at, modified_at, is_active, is_deleted
                ) VALUES (
                    %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    NOW(), NOW(), NOW(), TRUE, FALSE
                )
            """, (
                artifact_id, str(uuid.uuid4()), 'COMP-001', 'MAIN', inv_id, inv_no,
                'TAX_INVOICE_A4_PDF', 'TAX_INVOICE_TATTLY_THREADS_CANONICAL_V1', '1.0.0', 'FROZEN',
                path_client, sha256_hash, len(pdf_bytes), 1
            ))

            conn.commit()

            # Record Dispatch Status in workbook Sheet2
            for raw_row in sdata["raw_rows"]:
                ws_disp.cell(row=raw_row["excel_row"], column=15, value=inv_no)

            print(f"✓ Created {inv_no} | SIS: {sis_code:6s} | PO: {po_no} | {total_qty:3d} Pairs | Rs {grand_total:,.2f} | File: {pdf_filename} ({len(pdf_bytes)} bytes)")
            rendered_invoices.append(inv_no)
            current_seq += 1

        # Header for column 15 in Sheet2 if not present
        if ws_disp.cell(row=1, column=15).value is None:
            ws_disp.cell(row=1, column=15, value="Dispatch Status")

        wb_disp.save(DISPATCH_PATH)
        print(f"✓ Updated Dispatch Status in {DISPATCH_PATH} (Sheet2)")

        await browser.close()

    conn.close()
    print("=" * 80)
    print(f"Successfully generated all {len(rendered_invoices)} Tax Invoices (TT2026-2027/129 to TT2026-2027/137)!")
    print("=" * 80)

if __name__ == "__main__":
    asyncio.run(main())
