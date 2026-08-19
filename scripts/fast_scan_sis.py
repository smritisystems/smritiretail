"""
Project      : SMRITI Retail OS
Author       : Jawahar Ramkripal Mallah
Designation  : Chief Systems Architect & Creator
Email        : support@smritibooks.com
Websites     : smritibooks.com | erpnbook.com | aitdl.com
Version      : 3.16.0
Created      : 2026-08-19
Modified     : 2026-08-19
Copyright    : © SMRITIBooks.com. All Rights Reserved.
License      : Proprietary Commercial Software
"""

import os
import openpyxl

base = r"F:\Smriti-Clients Data\Tattly Threads\Invoice"
excel_files = [
    "Tattly_Threads_Consolidated_PO_Extraction.xlsx",
    "Tattly_Threads_Consolidated_PO_Extraction2.xlsx",
    "All_86_Bills_Audit_12_and_14_Aug.xlsx",
    "RIL_Dispatch_09-08-2026.xlsx",
    "RIL_Dispatch_09-08-2026-1.xlsx"
]

target_sis = ['TW07', 'TUK5', 'TYAC']

for ef in excel_files:
    fp = os.path.join(base, ef)
    if os.path.exists(fp):
        print(f"\nScanning {ef}...")
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            header = None
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    header = [str(c) if c is not None else '' for c in row]
                    continue
                row_str = " ".join([str(c) for c in row if c is not None])
                for s in target_sis:
                    if s in row_str:
                        print(f"  [{ef} -> {sheetname}] MATCH {s}:")
                        row_dict = {h: v for h, v in zip(header, row) if v is not None}
                        print("   ", row_dict)
