"""
Project      : SMRITI Retail OS
Author       : Jawahar Ramkripal Mallah
Designation  : Chief Systems Architect & Creator
Email        : support@smritibooks.com
Websites     : smritibooks.com | erpnbook.com | aitdl.com
Version      : 3.16.0
Created      : 2026-08-19
Modified     : 2026-08-19
Copyright    : © SMRITIBooks.com. All Rights Reserved.
License      : Proprietary Commercial Software
"""

import os
import openpyxl

fp = r"F:\Smriti-Clients Data\Tattly Threads\Invoice\Tattly_Threads_Consolidated_PO_Extraction.xlsx"
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)

for sn in wb.sheetnames:
    ws = wb[sn]
    count = 0
    for row in ws.iter_rows(values_only=True):
        row_str = " ".join([str(c) for c in row if c is not None])
        if "TW07" in row_str or "Commercial" in row_str or "Prestige" in row_str:
            print(f"[{sn}] MATCH: {row}")
            count += 1
            if count > 5:
                break
    print(f"Sheet {sn} total matches: {count}")
