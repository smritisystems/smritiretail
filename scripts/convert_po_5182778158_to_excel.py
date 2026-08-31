"""
Project      : SMRITI Retail OS
Author       : Jawahar Ramkripal Mallah
Designation  : Chief Systems Architect & Creator
Email        : support@smritibooks.com
Websites     : smritibooks.com | erpnbook.com | aitdl.com
Version      : 3.28.0
Created      : 2026-08-25
Modified     : 2026-08-25
Copyright    : © SMRITIBooks.com. All Rights Reserved.
License      : Proprietary Commercial Software
Classification: Internal
"""

import os
import re
from decimal import Decimal
import pypdf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PDF_PATH = r"F:\Smriti-Clients Data\Tattly Threads\Invoice\Tattly Threads\5182778158.pdf"

OUTPUT_PATHS = [
    r"F:\Smriti-Clients Data\Tattly Threads\Invoice\Tattly Threads\PO_5182778158_Reliance_Retail_Tax_Invoice_Matrix.xlsx",
    r"F:\Smriti-Clients Data\Tattly Threads\5182778158_PO_Full_Details.xlsx",
    r"F:\SMRITRretailNX\exports\PO_5182778158_Reliance_Retail_Tax_Invoice_Matrix.xlsx",
]

def parse_po_pdf(pdf_path):
    reader = pypdf.PdfReader(pdf_path)
    items = []
    
    for page_idx in range(1, 63):
        text = reader.pages[page_idx].extract_text()
        if not text:
            continue
        
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        i = 0
        while i < len(lines):
            line = lines[i]
            m_start = re.match(r'^(\d+)\s+(\d{10,14})$', line)
            if m_start:
                sr_no = int(m_start.group(1))
                article_no = m_start.group(2)
                
                hsn = lines[i+1] if i+1 < len(lines) else ""
                ean = lines[i+2] if i+2 < len(lines) else ""
                vendor_art = lines[i+3] if i+3 < len(lines) else ""
                desc = lines[i+4] if i+4 < len(lines) else ""
                
                delivery_date = "15.09.2026"
                site = "S4NN"
                qty = Decimal("0")
                uom = "EA"
                mrp = Decimal("0")
                base_cost = Decimal("0")
                igst_pct = Decimal("5.00")
                igst_amt = Decimal("0")
                total_base_val = Decimal("0")
                
                for j in range(i+4, min(i+18, len(lines))):
                    if re.match(r'^\d{2}\.\d{2}\.\d{4}$', lines[j]):
                        delivery_date = lines[j]
                        if j+1 < len(lines) and re.match(r'^[A-Z0-9]{4}$', lines[j+1]):
                            site = lines[j+1]
                    
                    m_qty = re.search(r'([\d,]+\.?\d*)\s+(EA|PRS|SET|PCS|PAIR)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)', lines[j])
                    if m_qty:
                        qty = Decimal(m_qty.group(1).replace(",", ""))
                        uom = m_qty.group(2)
                        mrp = Decimal(m_qty.group(3).replace(",", ""))
                        base_cost = Decimal(m_qty.group(4).replace(",", ""))
                        igst_pct = Decimal(m_qty.group(5).replace(",", ""))
                    
                    if "Sub Total of Qty" in lines[j]:
                        val_lines = [lines[k] for k in range(max(i, j-6), j) if re.match(r'^[\d,]+\.\d{2}$', lines[k])]
                        if len(val_lines) >= 2:
                            try:
                                igst_amt = Decimal(val_lines[-2].replace(",", ""))
                                total_base_val = Decimal(val_lines[-1].replace(",", ""))
                            except:
                                pass
                        elif len(val_lines) == 1:
                            try:
                                total_base_val = Decimal(val_lines[0].replace(",", ""))
                            except:
                                pass
                        break
                
                if total_base_val == Decimal("0") and qty > 0 and base_cost > 0:
                    total_base_val = (qty * base_cost).quantize(Decimal("0.01"))
                if igst_amt == Decimal("0") and total_base_val > 0 and igst_pct > 0:
                    igst_amt = (total_base_val * igst_pct / Decimal("100")).quantize(Decimal("0.01"))
                
                art_name = desc
                color = ""
                size = ""
                parts = [p.strip() for p in desc.split(",")]
                if len(parts) >= 3:
                    art_name = parts[0]
                    color = parts[1]
                    size = parts[2]
                elif len(parts) == 2:
                    art_name = parts[0]
                    color = parts[1]
                
                items.append({
                    "sr_no": sr_no,
                    "article_no": article_no,
                    "hsn": hsn,
                    "ean": ean,
                    "vendor_art": vendor_art,
                    "description": desc,
                    "art_name": art_name,
                    "color": color,
                    "size": size,
                    "delivery_date": delivery_date,
                    "site": site,
                    "qty": float(qty),
                    "uom": uom,
                    "mrp": float(mrp),
                    "base_cost": float(base_cost),
                    "igst_pct": float(igst_pct),
                    "igst_amt": float(igst_amt),
                    "total_base_val": float(total_base_val),
                    "total_gross_val": float(total_base_val + igst_amt)
                })
            i += 1
    return items

def build_excel(items):
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    # Styles
    f_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    f_sub = Font(name="Calibri", size=10, italic=True, color="E2E8F0")
    f_card_label = Font(name="Calibri", size=9, bold=True, color="475569")
    f_card_val = Font(name="Calibri", size=11, bold=True, color="0F172A")
    f_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    f_data = Font(name="Calibri", size=10)
    f_mono = Font(name="Consolas", size=9)
    f_total = Font(name="Calibri", size=10, bold=True, color="000000")
    
    fill_navy = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_blue = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    fill_teal = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    fill_indigo = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    fill_slate = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_card = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_total = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid") # Soft yellow
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    total_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    # ─────────────────────────────────────────────────────────────
    # SHEET 1: PO Summary & Metadata
    # ─────────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title="PO Summary & Details")
    ws_sum.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws_sum.merge_cells("A1:G2")
    ws_sum["A1"] = "RELIANCE RETAIL LIMITED — PURCHASE ORDER LEDGER"
    ws_sum["A1"].font = f_title
    ws_sum["A1"].fill = fill_navy
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws_sum.merge_cells("A3:G3")
    ws_sum["A3"] = "Official Statutory PO Master: 5182778158 | Delivery NDC: S4NN Tumkur | Seller: Tattly Threads"
    ws_sum["A3"].font = f_sub
    ws_sum["A3"].fill = fill_blue
    ws_sum["A3"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Summary Info Cards
    info_rows = [
        ("PO Number", "5182778158", "Vendor Code", "32071140 (TATTLY THREADS)"),
        ("PO Date", "31.07.2026", "Vendor GSTIN", "27AAXFT2508H1ZR"),
        ("Delivery Date", "15.09.2026", "Vendor PAN", "AAXFT2508H"),
        ("Delivery Site", "S4NN (RRL FOOTPRINT Tumkur NDC)", "Vendor Email", "TATTLYTHREADSOPERATIONS@GMAIL.COM"),
        ("Customer Name", "Reliance Retail Limited", "Buyer Contact", "Prakash d Jagatap (prakash.jagatap@ril.com)"),
        ("Customer GSTIN", "29AABCR1718E1ZL", "Place of Supply", "Karnataka (29) — Inter-State"),
        ("Billing Address", "NO 62/2, RIL BUILIDING, RICHMOND ROAD, BANGALORE - 560025", "Delivery Address", "Distribution Center, Survey No 54 1 Nandihalli Village, 55th KM Stone NH 4 Tumkur Road, Oordigree Hobli Taluka, TUMKUR, KAR - 572101"),
    ]
    
    for idx, (lbl1, val1, lbl2, val2) in enumerate(info_rows, start=5):
        ws_sum.cell(idx, 1, lbl1).font = f_card_label
        ws_sum.cell(idx, 1).fill = fill_card
        ws_sum.cell(idx, 1).border = thin_border
        
        ws_sum.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=3)
        ws_sum.cell(idx, 2, val1).font = f_card_val
        ws_sum.cell(idx, 2).border = thin_border
        ws_sum.cell(idx, 3).border = thin_border
        
        ws_sum.cell(idx, 4, lbl2).font = f_card_label
        ws_sum.cell(idx, 4).fill = fill_card
        ws_sum.cell(idx, 4).border = thin_border
        
        ws_sum.merge_cells(start_row=idx, start_column=5, end_row=idx, end_column=7)
        ws_sum.cell(idx, 5, val2).font = f_card_val
        ws_sum.cell(idx, 5).border = thin_border
        for c in range(5, 8):
            ws_sum.cell(idx, c).border = thin_border
    
    # Financial KPI Cards
    kpis = [
        ("TOTAL ITEMS", f"{len(items)} Lines", fill_navy),
        ("TOTAL QUANTITY", f"{sum(it['qty'] for it in items):,.0f} Pairs", fill_blue),
        ("TOTAL BASIC VALUE", f"₹ {sum(it['total_base_val'] for it in items):,.2f}", fill_teal),
        ("TOTAL IGST (5%)", f"₹ {sum(it['igst_amt'] for it in items):,.2f}", fill_indigo),
        ("TOTAL PO VALUE", f"₹ {sum(it['total_gross_val'] for it in items):,.2f}", fill_navy),
    ]
    
    start_kpi_row = 14
    for idx, (label, val, fill_k) in enumerate(kpis, start=1):
        c = ws_sum.cell(start_kpi_row, idx + 1)
        c.value = label
        c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        c.fill = fill_k
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
        
        v = ws_sum.cell(start_kpi_row + 1, idx + 1)
        v.value = val
        v.font = Font(name="Calibri", size=12, bold=True, color="0F172A")
        v.fill = fill_slate
        v.alignment = Alignment(horizontal="center", vertical="center")
        v.border = thin_border
    
    # ─────────────────────────────────────────────────────────────
    # SHEET 2: PO Items Register (All 426 Lines)
    # ─────────────────────────────────────────────────────────────
    ws_items = wb.create_sheet(title="PO Items Register")
    ws_items.views.sheetView[0].showGridLines = True
    
    item_headers = [
        "Sr No", "Article No (SAP)", "HSN Code", "EAN / Barcode", "Vendor Style / Art",
        "Material Description", "Article Name", "Color", "Size", "Delivery Date",
        "Site", "Quantity (Pairs)", "UOM", "MRP (₹)", "Base Rate (₹)",
        "IGST (%)", "IGST Amt (₹)", "Total Base Value (₹)", "Total Value (₹)"
    ]
    
    ws_items.row_dimensions[1].height = 28
    for col_num, h_name in enumerate(item_headers, start=1):
        cell = ws_items.cell(1, col_num, h_name)
        cell.font = f_header
        cell.fill = fill_navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    for r_idx, it in enumerate(items, start=2):
        ws_items.row_dimensions[r_idx].height = 18
        row_fill = fill_zebra if r_idx % 2 == 0 else PatternFill(fill_type=None)
        
        ws_items.cell(r_idx, 1, it["sr_no"]).alignment = Alignment(horizontal="center")
        ws_items.cell(r_idx, 2, it["article_no"]).font = f_mono
        ws_items.cell(r_idx, 3, it["hsn"]).alignment = Alignment(horizontal="center")
        ws_items.cell(r_idx, 4, it["ean"]).font = f_mono
        ws_items.cell(r_idx, 5, it["vendor_art"]).font = Font(name="Calibri", size=10, bold=True)
        ws_items.cell(r_idx, 6, it["description"])
        ws_items.cell(r_idx, 7, it["art_name"])
        ws_items.cell(r_idx, 8, it["color"]).alignment = Alignment(horizontal="center")
        ws_items.cell(r_idx, 9, it["size"]).alignment = Alignment(horizontal="center")
        ws_items.cell(r_idx, 10, it["delivery_date"]).alignment = Alignment(horizontal="center")
        ws_items.cell(r_idx, 11, it["site"]).alignment = Alignment(horizontal="center")
        
        c_qty = ws_items.cell(r_idx, 12, it["qty"])
        c_qty.number_format = '#,##0'
        c_qty.alignment = Alignment(horizontal="right")
        
        ws_items.cell(r_idx, 13, it["uom"]).alignment = Alignment(horizontal="center")
        
        c_mrp = ws_items.cell(r_idx, 14, it["mrp"])
        c_mrp.number_format = '₹ #,##0.00'
        c_mrp.alignment = Alignment(horizontal="right")
        
        c_cost = ws_items.cell(r_idx, 15, it["base_cost"])
        c_cost.number_format = '₹ #,##0.00'
        c_cost.alignment = Alignment(horizontal="right")
        
        c_igstp = ws_items.cell(r_idx, 16, it["igst_pct"] / 100.0)
        c_igstp.number_format = '0.00%'
        c_igstp.alignment = Alignment(horizontal="right")
        
        c_igst = ws_items.cell(r_idx, 17, it["igst_amt"])
        c_igst.number_format = '₹ #,##0.00'
        c_igst.alignment = Alignment(horizontal="right")
        
        c_base = ws_items.cell(r_idx, 18, it["total_base_val"])
        c_base.number_format = '₹ #,##0.00'
        c_base.alignment = Alignment(horizontal="right")
        
        c_gross = ws_items.cell(r_idx, 19, it["total_gross_val"])
        c_gross.number_format = '₹ #,##0.00'
        c_gross.alignment = Alignment(horizontal="right")
        
        for c in range(1, 20):
            c_cell = ws_items.cell(r_idx, c)
            if row_fill.fill_type:
                c_cell.fill = row_fill
            c_cell.border = thin_border
            if not c_cell.font.name:
                c_cell.font = f_data
    
    # Total Row
    tot_row = len(items) + 2
    ws_items.row_dimensions[tot_row].height = 22
    ws_items.cell(tot_row, 1, "TOTAL").font = f_total
    ws_items.cell(tot_row, 1).alignment = Alignment(horizontal="center")
    ws_items.cell(tot_row, 2, f"{len(items)} Items").font = f_total
    
    c_tot_qty = ws_items.cell(tot_row, 12, f"=SUM(L2:L{tot_row-1})")
    c_tot_qty.font = f_total
    c_tot_qty.number_format = '#,##0'
    c_tot_qty.alignment = Alignment(horizontal="right")
    
    c_tot_igst = ws_items.cell(tot_row, 17, f"=SUM(Q2:Q{tot_row-1})")
    c_tot_igst.font = f_total
    c_tot_igst.number_format = '₹ #,##0.00'
    c_tot_igst.alignment = Alignment(horizontal="right")
    
    c_tot_base = ws_items.cell(tot_row, 18, f"=SUM(R2:R{tot_row-1})")
    c_tot_base.font = f_total
    c_tot_base.number_format = '₹ #,##0.00'
    c_tot_base.alignment = Alignment(horizontal="right")
    
    c_tot_gross = ws_items.cell(tot_row, 19, f"=SUM(S2:S{tot_row-1})")
    c_tot_gross.font = f_total
    c_tot_gross.number_format = '₹ #,##0.00'
    c_tot_gross.alignment = Alignment(horizontal="right")
    
    for c in range(1, 20):
        c_cell = ws_items.cell(tot_row, c)
        c_cell.fill = fill_total
        c_cell.border = total_border
    
    ws_items.freeze_panes = "A2"
    ws_items.auto_filter.ref = f"A1:S{tot_row-1}"
    
    # ─────────────────────────────────────────────────────────────
    # SHEET 3: Article & Color & Size Matrix
    # ─────────────────────────────────────────────────────────────
    ws_mat = wb.create_sheet(title="Article Color Size Matrix")
    ws_mat.views.sheetView[0].showGridLines = True
    
    # Group items by (vendor_art, art_name, color)
    matrix_map = {}
    all_sizes = ["35", "36", "37", "38", "39", "40", "41", "42"]
    
    for it in items:
        key = (it["vendor_art"], it["art_name"], it["color"], it["mrp"], it["base_cost"])
        if key not in matrix_map:
            matrix_map[key] = {s: 0 for s in all_sizes}
        sz = str(it["size"]).strip()
        if sz in matrix_map[key]:
            matrix_map[key][sz] += it["qty"]
        else:
            matrix_map[key][sz] = it["qty"]
    
    mat_headers = [
        "Style / Vendor Art", "Article Description", "Color", "Base Rate (₹)", "MRP (₹)",
        "Sz 35", "Sz 36", "Sz 37", "Sz 38", "Sz 39", "Sz 40", "Sz 41", "Sz 42",
        "Total Pairs", "Base Value (₹)", "IGST 5% (₹)", "Gross Value (₹)"
    ]
    
    ws_mat.row_dimensions[1].height = 28
    for col_num, h_name in enumerate(mat_headers, start=1):
        cell = ws_mat.cell(1, col_num, h_name)
        cell.font = f_header
        cell.fill = fill_teal
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    mat_row = 2
    for (v_art, a_name, col, mrp_val, base_rate), size_dict in sorted(matrix_map.items()):
        ws_mat.row_dimensions[mat_row].height = 18
        row_fill = fill_zebra if mat_row % 2 == 0 else PatternFill(fill_type=None)
        
        ws_mat.cell(mat_row, 1, v_art).font = Font(name="Calibri", size=10, bold=True)
        ws_mat.cell(mat_row, 2, a_name)
        ws_mat.cell(mat_row, 3, col).alignment = Alignment(horizontal="center")
        
        c_rate = ws_mat.cell(mat_row, 4, base_rate)
        c_rate.number_format = '₹ #,##0.00'
        c_rate.alignment = Alignment(horizontal="right")
        
        c_mrp = ws_mat.cell(mat_row, 5, mrp_val)
        c_mrp.number_format = '₹ #,##0.00'
        c_mrp.alignment = Alignment(horizontal="right")
        
        row_qty = 0
        for s_idx, sz_code in enumerate(all_sizes, start=6):
            s_qty = size_dict.get(sz_code, 0)
            row_qty += s_qty
            c_s = ws_mat.cell(mat_row, s_idx, s_qty if s_qty > 0 else "-")
            c_s.alignment = Alignment(horizontal="center")
            if s_qty > 0:
                c_s.font = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
        
        c_tot_p = ws_mat.cell(mat_row, 14, f"=SUM(F{mat_row}:M{mat_row})")
        c_tot_p.number_format = '#,##0'
        c_tot_p.font = Font(name="Calibri", size=10, bold=True)
        c_tot_p.alignment = Alignment(horizontal="right")
        
        c_bval = ws_mat.cell(mat_row, 15, f"=N{mat_row}*D{mat_row}")
        c_bval.number_format = '₹ #,##0.00'
        c_bval.alignment = Alignment(horizontal="right")
        
        c_igst = ws_mat.cell(mat_row, 16, f"=O{mat_row}*0.05")
        c_igst.number_format = '₹ #,##0.00'
        c_igst.alignment = Alignment(horizontal="right")
        
        c_gval = ws_mat.cell(mat_row, 17, f"=O{mat_row}+P{mat_row}")
        c_gval.number_format = '₹ #,##0.00'
        c_gval.alignment = Alignment(horizontal="right")
        
        for c in range(1, 18):
            c_cell = ws_mat.cell(mat_row, c)
            if row_fill.fill_type:
                c_cell.fill = row_fill
            c_cell.border = thin_border
            if not c_cell.font.name:
                c_cell.font = f_data
        
        mat_row += 1
    
    # Matrix Total Row
    ws_mat.row_dimensions[mat_row].height = 22
    ws_mat.cell(mat_row, 1, "TOTAL").font = f_total
    ws_mat.cell(mat_row, 1).alignment = Alignment(horizontal="center")
    ws_mat.cell(mat_row, 2, f"{len(matrix_map)} SKU Variations").font = f_total
    
    for s_idx, col_let in enumerate(["F", "G", "H", "I", "J", "K", "L", "M"], start=6):
        c_sz_tot = ws_mat.cell(mat_row, s_idx, f"=SUM({col_let}2:{col_let}{mat_row-1})")
        c_sz_tot.font = f_total
        c_sz_tot.number_format = '#,##0'
        c_sz_tot.alignment = Alignment(horizontal="center")
    
    c_mqty = ws_mat.cell(mat_row, 14, f"=SUM(N2:N{mat_row-1})")
    c_mqty.font = f_total
    c_mqty.number_format = '#,##0'
    c_mqty.alignment = Alignment(horizontal="right")
    
    c_mbval = ws_mat.cell(mat_row, 15, f"=SUM(O2:O{mat_row-1})")
    c_mbval.font = f_total
    c_mbval.number_format = '₹ #,##0.00'
    c_mbval.alignment = Alignment(horizontal="right")
    
    c_migst = ws_mat.cell(mat_row, 16, f"=SUM(P2:P{mat_row-1})")
    c_migst.font = f_total
    c_migst.number_format = '₹ #,##0.00'
    c_migst.alignment = Alignment(horizontal="right")
    
    c_mgval = ws_mat.cell(mat_row, 17, f"=SUM(Q2:Q{mat_row-1})")
    c_mgval.font = f_total
    c_mgval.number_format = '₹ #,##0.00'
    c_mgval.alignment = Alignment(horizontal="right")
    
    for c in range(1, 18):
        c_cell = ws_mat.cell(mat_row, c)
        c_cell.fill = fill_total
        c_cell.border = total_border
    
    ws_mat.freeze_panes = "A2"
    ws_mat.auto_filter.ref = f"A1:Q{mat_row-1}"
    
    # ─────────────────────────────────────────────────────────────
    # SHEET 4: Style-Wise Aggregation Summary
    # ─────────────────────────────────────────────────────────────
    ws_style = wb.create_sheet(title="Style-Wise Summary")
    ws_style.views.sheetView[0].showGridLines = True
    
    style_headers = [
        "Vendor Style / Article", "Article Description", "Colors Available", "SKU Variations",
        "Total Pairs", "Base Rate (₹)", "MRP (₹)", "Total Base Value (₹)",
        "IGST 5% (₹)", "Total Gross Value (₹)", "% of PO Value"
    ]
    
    ws_style.row_dimensions[1].height = 28
    for col_num, h_name in enumerate(style_headers, start=1):
        cell = ws_style.cell(1, col_num, h_name)
        cell.font = f_header
        cell.fill = fill_indigo
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    style_agg = {}
    for it in items:
        st = it["vendor_art"]
        if st not in style_agg:
            style_agg[st] = {
                "desc": it["art_name"],
                "colors": set(),
                "skus": 0,
                "qty": 0,
                "base_cost": it["base_cost"],
                "mrp": it["mrp"],
                "base_val": 0,
                "igst": 0,
                "gross": 0
            }
        style_agg[st]["colors"].add(it["color"])
        style_agg[st]["skus"] += 1
        style_agg[st]["qty"] += it["qty"]
        style_agg[st]["base_val"] += it["total_base_val"]
        style_agg[st]["igst"] += it["igst_amt"]
        style_agg[st]["gross"] += it["total_gross_val"]
    
    tot_all_gross = sum(s["gross"] for s in style_agg.values()) or 1
    
    st_row = 2
    for st_code, s_info in sorted(style_agg.items()):
        ws_style.row_dimensions[st_row].height = 18
        row_fill = fill_zebra if st_row % 2 == 0 else PatternFill(fill_type=None)
        
        ws_style.cell(st_row, 1, st_code).font = Font(name="Calibri", size=10, bold=True)
        ws_style.cell(st_row, 2, s_info["desc"])
        ws_style.cell(st_row, 3, ", ".join(sorted(s_info["colors"])))
        ws_style.cell(st_row, 4, s_info["skus"]).alignment = Alignment(horizontal="center")
        
        c_sqty = ws_style.cell(st_row, 5, s_info["qty"])
        c_sqty.number_format = '#,##0'
        c_sqty.font = Font(name="Calibri", size=10, bold=True)
        c_sqty.alignment = Alignment(horizontal="right")
        
        c_srate = ws_style.cell(st_row, 6, s_info["base_cost"])
        c_srate.number_format = '₹ #,##0.00'
        c_srate.alignment = Alignment(horizontal="right")
        
        c_smrp = ws_style.cell(st_row, 7, s_info["mrp"])
        c_smrp.number_format = '₹ #,##0.00'
        c_smrp.alignment = Alignment(horizontal="right")
        
        c_sbval = ws_style.cell(st_row, 8, s_info["base_val"])
        c_sbval.number_format = '₹ #,##0.00'
        c_sbval.alignment = Alignment(horizontal="right")
        
        c_sigst = ws_style.cell(st_row, 9, s_info["igst"])
        c_sigst.number_format = '₹ #,##0.00'
        c_sigst.alignment = Alignment(horizontal="right")
        
        c_sgross = ws_style.cell(st_row, 10, s_info["gross"])
        c_sgross.number_format = '₹ #,##0.00'
        c_sgross.alignment = Alignment(horizontal="right")
        
        c_spct = ws_style.cell(st_row, 11, s_info["gross"] / tot_all_gross)
        c_spct.number_format = '0.00%'
        c_spct.alignment = Alignment(horizontal="right")
        
        for c in range(1, 12):
            c_cell = ws_style.cell(st_row, c)
            if row_fill.fill_type:
                c_cell.fill = row_fill
            c_cell.border = thin_border
            if not c_cell.font.name:
                c_cell.font = f_data
        
        st_row += 1
    
    # Total Row for Style Summary
    ws_style.row_dimensions[st_row].height = 22
    ws_style.cell(st_row, 1, "TOTAL").font = f_total
    ws_style.cell(st_row, 1).alignment = Alignment(horizontal="center")
    ws_style.cell(st_row, 2, f"{len(style_agg)} Styles").font = f_total
    ws_style.cell(st_row, 4, f"=SUM(D2:D{st_row-1})").font = f_total
    ws_style.cell(st_row, 4).alignment = Alignment(horizontal="center")
    
    c_st_tot_qty = ws_style.cell(st_row, 5, f"=SUM(E2:E{st_row-1})")
    c_st_tot_qty.font = f_total
    c_st_tot_qty.number_format = '#,##0'
    c_st_tot_qty.alignment = Alignment(horizontal="right")
    
    c_st_tot_bval = ws_style.cell(st_row, 8, f"=SUM(H2:H{st_row-1})")
    c_st_tot_bval.font = f_total
    c_st_tot_bval.number_format = '₹ #,##0.00'
    c_st_tot_bval.alignment = Alignment(horizontal="right")
    
    c_st_tot_igst = ws_style.cell(st_row, 9, f"=SUM(I2:I{st_row-1})")
    c_st_tot_igst.font = f_total
    c_st_tot_igst.number_format = '₹ #,##0.00'
    c_st_tot_igst.alignment = Alignment(horizontal="right")
    
    c_st_tot_gross = ws_style.cell(st_row, 10, f"=SUM(J2:J{st_row-1})")
    c_st_tot_gross.font = f_total
    c_st_tot_gross.number_format = '₹ #,##0.00'
    c_st_tot_gross.alignment = Alignment(horizontal="right")
    
    c_st_tot_pct = ws_style.cell(st_row, 11, f"=SUM(K2:K{st_row-1})")
    c_st_tot_pct.font = f_total
    c_st_tot_pct.number_format = '0.00%'
    c_st_tot_pct.alignment = Alignment(horizontal="right")
    
    for c in range(1, 12):
        c_cell = ws_style.cell(st_row, c)
        c_cell.fill = fill_total
        c_cell.border = total_border
    
    ws_style.freeze_panes = "A2"
    ws_style.auto_filter.ref = f"A1:K{st_row-1}"
    
    # Auto-fit Column Widths across all sheets
    for ws in [ws_sum, ws_items, ws_mat, ws_style]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if cell.number_format and "₹" in cell.number_format:
                    val_str += "    "
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
    
    return wb

def main():
    print(f"Parsing Reliance PO PDF: {PDF_PATH}...")
    items = parse_po_pdf(PDF_PATH)
    print(f"Parsed {len(items)} line items successfully.")
    
    wb = build_excel(items)
    
    for out_path in OUTPUT_PATHS:
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        try:
            wb.save(out_path)
            print(f"[SUCCESS] Saved Excel Workbook: {out_path} ({os.path.getsize(out_path):,} bytes)")
        except Exception as e:
            print(f"Error saving to {out_path}: {e}")

if __name__ == "__main__":
    main()
