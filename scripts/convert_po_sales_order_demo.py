import csv
import json
from pathlib import Path
from openpyxl import load_workbook

root = Path(__file__).resolve().parent.parent
xlsx_path = root / 'exports' / 'PO_5182778158_Reliance_Retail_Tax_Invoice_Matrix.xlsx'
out_json = root / 'exports' / 'PO_5182778158_Reliance_Retail_Tax_Invoice_Matrix_sales_order_demo.json'
out_csv = root / 'exports' / 'PO_5182778158_Reliance_Retail_Tax_Invoice_Matrix_sales_order_demo.csv'

wb = load_workbook(xlsx_path, read_only=True, data_only=True)
ws = wb['PO Items Register']
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
records = []
for row in rows[1:]:
    rec = dict(zip(header, row))
    if not rec.get('Article No (SAP)'):
        continue
    records.append(rec)

if not records:
    raise SystemExit('No data rows found in PO Items Register sheet.')

sample = records[0]
payload = {
    'docPrefix': 'SO',
    'docNumber': 'SO-2026-0001',
    'docDate': '2026-09-01',
    'docTime': '10:30',
    'referenceNo': 'PO-5182778158',
    'deliveryTerms': 'Door Delivery',
    'paymentTerms': 'Net 30',
    'orderStatus': 'Open',
    'customerId': 'RIL-001',
    'customerCode': 'RIL-001',
    'customerName': 'Reliance Retail Limited',
    'salesStaff': 'Sales Desk',
    'remarks': 'Imported from PO_5182778158_Reliance_Retail_Tax_Invoice_Matrix.xlsx',
    'items': [{
        'id': f"PO-5182778158-{1:02d}",
        'stockNo': str(sample.get('Article No (SAP)', '')).strip(),
        'description': str(sample.get('Material Description', '')).strip(),
        'hsn': str(sample.get('HSN Code', '')).strip(),
        'rate': float(sample.get('Base Rate (₹)', 0) or 0),
        'quantity': float(sample.get('Quantity (Pairs)', 0) or 0),
        'uom': str(sample.get('UOM', 'EA')).strip() or 'EA',
        'value': float(sample.get('Total Base Value (₹)', 0) or 0),
        'discPercent': 0,
        'discAmount': 0,
        'taxPercent': float(sample.get('IGST (%)', 0) or 0),
        'taxAmount': float(sample.get('IGST Amt (₹)', 0) or 0),
        'total': float(sample.get('Total Value (₹)', 0) or 0),
        'salesStaff': 'Sales Desk',
    }]
}

out_json.write_text(json.dumps(payload, indent=2), encoding='utf-8')

csv_fieldnames = ['docPrefix','docNumber','docDate','docTime','referenceNo','deliveryTerms','paymentTerms','orderStatus','customerId','customerCode','customerName','salesStaff','remarks','stockNo','description','hsn','rate','quantity','uom','value','discPercent','discAmount','taxPercent','taxAmount','total','salesStaffItem']
with out_csv.open('w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=csv_fieldnames)
    writer.writeheader()
    writer.writerow({
        'docPrefix': payload['docPrefix'],
        'docNumber': payload['docNumber'],
        'docDate': payload['docDate'],
        'docTime': payload['docTime'],
        'referenceNo': payload['referenceNo'],
        'deliveryTerms': payload['deliveryTerms'],
        'paymentTerms': payload['paymentTerms'],
        'orderStatus': payload['orderStatus'],
        'customerId': payload['customerId'],
        'customerCode': payload['customerCode'],
        'customerName': payload['customerName'],
        'salesStaff': payload['salesStaff'],
        'remarks': payload['remarks'],
        'stockNo': payload['items'][0]['stockNo'],
        'description': payload['items'][0]['description'],
        'hsn': payload['items'][0]['hsn'],
        'rate': payload['items'][0]['rate'],
        'quantity': payload['items'][0]['quantity'],
        'uom': payload['items'][0]['uom'],
        'value': payload['items'][0]['value'],
        'discPercent': payload['items'][0]['discPercent'],
        'discAmount': payload['items'][0]['discAmount'],
        'taxPercent': payload['items'][0]['taxPercent'],
        'taxAmount': payload['items'][0]['taxAmount'],
        'total': payload['items'][0]['total'],
        'salesStaffItem': payload['items'][0]['salesStaff'],
    })

print(f'Generated {out_json.name} and {out_csv.name} from {xlsx_path.name}')
print(f'Imported item: {payload["items"][0]["stockNo"]} / {payload["items"][0]["description"]} / Qty {payload["items"][0]["quantity"]}')
